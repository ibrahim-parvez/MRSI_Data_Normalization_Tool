from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from copy import copy
import re
from datetime import datetime
import settings

# --- Helper Functions for He/CO2 Logic ---

def extract_run_number(identifier):
    if not identifier:
        return None, None
    match = re.search(r'(?i)[rR](\d+)(?:\.(\d+))?', str(identifier))
    if match:
        major = int(match.group(1))
        minor = int(match.group(2)) if match.group(2) else 0
        return major, minor
    return None, None

def get_valid_heco2_indices(ws):
    valid_rows = set()
    seen_majors = {} 
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=3).value
        if not val:
            continue
        s_val = str(val).strip().lower()
        if "heco2" in s_val or "co2" in s_val:
            major, minor = extract_run_number(s_val)
            if major is None or major == 1: 
                continue
            if major not in seen_majors:
                seen_majors[major] = (minor, r)
            else:
                prev_minor, prev_r = seen_majors[major]
                if minor < prev_minor:
                    seen_majors[major] = (minor, r)
    for _, row_idx in seen_majors.values():
        valid_rows.add(row_idx)
    return valid_rows

def step6_normalization_water(file_path: str):
    wb = load_workbook(file_path)
    # Load again to get calculated values for logic
    wb_values = load_workbook(file_path, data_only=True)

    if "Group_DNT" not in wb.sheetnames:
        print("❌ Sheet 'Group_DNT' not found.")
        return 0

    group_ws = wb["Group_DNT"]
    group_values_ws = wb_values["Group_DNT"]

    if "Normalization_DNT" in wb.sheetnames:
        del wb["Normalization_DNT"]

    summary_ws = wb.create_sheet("Normalization_DNT", wb.index(group_ws))

    # --- Styles Definition ---
    light_green_fill = PatternFill(start_color="DBF2D0", end_color="DBF2D0", fill_type="solid")
    box_blue_fill = PatternFill(start_color="DAE8F9", end_color="DAE8F9", fill_type="solid")
    box_peach_fill = PatternFill(start_color="FBE2D5", end_color="FBE2D5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    gray_box_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    heco2_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Updated He/CO2 Box Color
    heco2_box_fill = PatternFill(start_color="82CCEB", end_color="82CCEB", fill_type="solid")

    red_font = Font(color="FF0000")
    blue_font = Font(color="0000FF")
    orange_font = Font(color="E46C0A")
    green_font = Font(color="00B050")
    
    bold_red_font = Font(color="FF0000", bold=True)
    bold_blue_font = Font(color="0000FF", bold=True)
    bold_orange_font = Font(color="E46C0A", bold=True)
    bold_green_font = Font(color="00B050", bold=True)
    
    bold_blue_text_font = Font(color="0000FF", bold=True)
    bold_green_text_font = Font(color="00B050", bold=True)
    black_bold = Font(bold=True, color="000000")
    light_blue_bold = Font(bold=True, color="5DADE2")
    
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Border Styles
    thin_black_side = Side(style="thin", color="000000")
    medium_black_side = Side(style="medium", color="000000") # Thicker outer border
    double_black_side = Side(style="double", color="000000") # Double inner border
    
    # Number Formats
    FMT_3_DEC = "0.000"
    FMT_2_DEC = "0.00"

    # Settings
    stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    if stdev_threshold is None:
        print("⚠️ Warning: 'stdev_threshold' not found in settings. Conditional formatting will be skipped.")

    std_color_map = {
        "MRSI-STD-W1": red_font, "MRSI-STD-W2": blue_font,
        "USGS W-67400": orange_font, "USGS W-64444": green_font
    }
    std_bold_color_map = {
        "MRSI-STD-W1": bold_red_font, "MRSI-STD-W2": bold_blue_font,
        "USGS W-67400": bold_orange_font, "USGS W-64444": bold_green_font
    }

    # --- Border Helper with Thicker Outer & Double Inner capability ---
    def apply_box_border(ws, start_row, start_col, end_row, end_col, fill):
        """Applies fill and a MEDIUM outer border to a cell range."""
        # Fill
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
        
        # Outer Borders (Medium)
        for c in range(start_col, end_col + 1):
            # Top
            cell_top = ws.cell(row=start_row, column=c)
            b = copy(cell_top.border)
            cell_top.border = Border(top=medium_black_side, left=b.left, right=b.right, bottom=b.bottom)
            # Bottom
            cell_bot = ws.cell(row=end_row, column=c)
            b2 = copy(cell_bot.border)
            cell_bot.border = Border(bottom=medium_black_side, left=b2.left, right=b2.right, top=b2.top)
            
        for r in range(start_row, end_row + 1):
            # Left
            cell_l = ws.cell(row=r, column=start_col)
            b3 = copy(cell_l.border)
            cell_l.border = Border(left=medium_black_side, top=b3.top, right=b3.right, bottom=b3.bottom)
            # Right
            cell_r = ws.cell(row=r, column=end_col)
            b4 = copy(cell_r.border)
            cell_r.border = Border(right=medium_black_side, top=b4.top, left=b4.left, bottom=b4.bottom)

    def apply_vertical_divider(ws, col_idx, start_row, end_row):
        """Applies a DOUBLE line to the RIGHT of col_idx."""
        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            cur = cell.border
            cell.border = Border(left=cur.left, right=double_black_side, top=cur.top, bottom=cur.bottom)

    valid_heco2_src_rows = get_valid_heco2_indices(group_values_ws)

    # --- Setup Sheet ---
    # Apply Base Green Background (Rows 1-15)
    for r in range(1, 16):
        for c in range(1, 51):
            summary_ws.cell(row=r, column=c).fill = light_green_fill
            
    # Add Date to A1
    current_date = datetime.now().strftime("%Y-%m-%d")
    summary_ws["A1"] = current_date

    # === BOX 1: Blue Box (Row 2-8, C-O) ===
    apply_box_border(summary_ws, 2, 3, 8, 15, fill=box_blue_fill)
    apply_vertical_divider(summary_ws, 4, 2, 8) # Right of D
    
    # Horizontal Double Line Divider between Row 3 and 4 (C-O)
    for c in range(3, 16):
        cell = summary_ws.cell(row=3, column=c)
        cur = cell.border
        # Apply Double to bottom of Row 3
        cell.border = Border(left=cur.left, right=cur.right, top=cur.top, bottom=double_black_side)

    # === BOX 2: He/CO2 (Row 9, C-O) ===
    # Highlighting specific color #82cceb
    apply_box_border(summary_ws, 9, 3, 9, 15, fill=heco2_box_fill)
    summary_ws["C9"] = "He/CO2"
    summary_ws["C9"].font = black_bold
    summary_ws["C9"].alignment = center_align # Center Alignment added
    
    # 4. Make Row 9 taller
    summary_ws.row_dimensions[9].height = 20 

    # === BOX 3: Peach Box (Row 11-15, H-O) ===
    apply_box_border(summary_ws, 11, 8, 15, 15, fill=box_peach_fill)
    apply_vertical_divider(summary_ws, 9, 11, 15) # Right of I

    # === Text & Values ===
    summary_ws["C1"] = "Normalization"
    summary_ws["C1"].font = black_bold

    # Blue Box Content
    c2 = summary_ws["C2"]; c2.value = "EQ Time ="; c2.fill = yellow_fill; c2.font = black_bold
    summary_ws["C3"] = "Water Standards"; summary_ws["C3"].font = black_bold

    # Standards
    summary_ws["C5"] = "MRSI-STD-W1"; summary_ws["C5"].font = red_font
    summary_ws["C6"] = "MRSI-STD-W2"; summary_ws["C6"].font = blue_font
    summary_ws["C7"] = "USGS W-67400"; summary_ws["C7"].font = orange_font
    summary_ws["C8"] = "USGS W-64444"; summary_ws["C8"].font = green_font

    # F Column
    summary_ws["F2"] = "Published"; summary_ws["F2"].font = black_bold; summary_ws["F2"].alignment = center_align
    summary_ws["F3"] = "δ²H"; summary_ws["F3"].alignment = center_align; summary_ws["F3"].font = black_bold # Explicit bold request
    
    f_vals = [-3.52, -214.79, 1.250, -399.10]
    f_fonts = [red_font, blue_font, orange_font, green_font]
    for i, (val, ft) in enumerate(zip(f_vals, f_fonts), start=5):
        c = summary_ws[f"F{i}"]
        c.value = val; c.font = ft; c.alignment = center_align

    # G Column
    summary_ws["G3"] = "δ¹⁸O SMOW"; summary_ws["G3"].alignment = center_align; summary_ws["G3"].font = black_bold
    g_vals = [-0.580, -28.080, -1.97, -51.14]
    for i, (val, ft) in enumerate(zip(g_vals, f_fonts), start=5):
        c = summary_ws[f"G{i}"]
        c.value = val; c.font = ft; c.alignment = center_align

    # Headers K/N
    summary_ws["K2"] = "Measured Ave."; summary_ws["K2"].font = black_bold; summary_ws["K2"].alignment = center_align
    summary_ws["K3"] = "δ¹³C RAW"; summary_ws["K3"].alignment = center_align; summary_ws["K3"].font = black_bold
    summary_ws["N2"] = "Measured Ave."; summary_ws["N2"].font = black_bold; summary_ws["N2"].alignment = center_align
    summary_ws["N3"] = "δ¹⁸O RAW"; summary_ws["N3"].alignment = center_align; summary_ws["N3"].font = black_bold

    # O Column (Stretching Factor)
    summary_ws["O2"] = "Stretching"; summary_ws["O2"].font = black_bold
    summary_ws["O3"] = "Factor (λ)"; summary_ws["O3"].font = black_bold
    summary_ws["O5"].fill = yellow_fill; summary_ws["O7"].fill = yellow_fill
    
    # 1. & 2. Stretching Factor Formulas
    summary_ws["O5"] = "=(G5-G6)/(N5-N6)"
    summary_ws["O7"] = "=(G7-G8)/(N7-N8)" # Corrected from O8 to O7 per request
    
    # 3. Center Alignment for K5:O9
    for r in range(5, 10):
        for c in range(11, 16): # K(11) to O(15)
            summary_ws.cell(row=r, column=c).alignment = center_align

    # Peach Box Content (Row 11+)
    summary_ws["I11"] = "MRSI W1,W2"; summary_ws["I11"].font = bold_blue_text_font
    summary_ws["I14"] = "USGS"; summary_ws["I14"].font = bold_green_text_font
    summary_ws["J11"] = "slope"; summary_ws["J12"] = "intercept"
    summary_ws["J14"] = "slope"; summary_ws["J15"] = "intercept"

    # Formulas (shifted rows)
    summary_ws["N11"] = "=SLOPE(G5:G6,N5:N6)"; summary_ws["N11"].number_format = FMT_3_DEC
    summary_ws["N12"] = "=INTERCEPT(G5:G6,N5:N6)"; summary_ws["N12"].number_format = FMT_3_DEC
    summary_ws["N14"] = "=SLOPE(G7:G8,N7:N8)"; summary_ws["N14"].number_format = FMT_3_DEC
    summary_ws["N15"] = "=INTERCEPT(G7:G8,N7:N8)"; summary_ws["N15"].number_format = FMT_3_DEC

    # === Right Side Results (R-Y) ===
    col_R = 18; col_S = 19; col_T = 20; col_U = 21; col_V = 22; col_W = 23; col_X = 24; col_Y = 25
    summary_ws.cell(row=1, column=col_R, value="Water Standard Results").font = black_bold

    # 1. Apply Outer Medium Border to the entire R2:Y8 block (Fixes the corner/bottom/right borders)
    # Also applies the initial fill
    apply_box_border(summary_ws, 2, col_R, 8, col_Y, fill=gray_box_fill)

    # 2. Apply Double Line Dividers (Inner Lines)
    
    # Vertical divider: Right of S (Col 19) for rows 2-8 (S/T divider)
    apply_vertical_divider(summary_ws, col_S, 2, 8) 
    
    # Horizontal divider: Bottom of Row 3 for columns R-Y (Row 3/4 divider)
    for c in range(col_R, col_Y + 1):
        cell = summary_ws.cell(row=3, column=c)
        cur = cell.border
        # Apply Double to bottom of Row 3
        # Check if the bottom border is already the medium outer border (only at row 8, but R3 is safe)
        if cur.bottom.style != medium_black_side.style:
            cell.border = Border(left=cur.left, right=cur.right, top=cur.top, bottom=double_black_side)

    # Content for R-Y Box
    summary_ws.cell(row=3, column=col_R, value="Calibration").font = Font(bold=False, color="000000")
    summary_ws.cell(row=5, column=col_R, value="MRSI-STD-W1").font = red_font
    summary_ws.cell(row=6, column=col_R, value="MRSI-STD-W2").font = blue_font
    summary_ws.cell(row=7, column=col_R, value="USGS W-67400").font = orange_font
    summary_ws.cell(row=8, column=col_R, value="USGS W-64444").font = green_font

    delta18 = "δ¹⁸O SMOW"
    for c, val in [(col_V, delta18), (col_W, delta18), (col_X, "STDEV"), (col_Y, "N")]:
        cell = summary_ws.cell(row=2, column=c, value=val)
        cell.alignment = center_align; cell.font = black_bold

    summary_ws.cell(row=3, column=col_V, value="MRSI W1,W2").font = light_blue_bold
    summary_ws.cell(row=3, column=col_V).alignment = center_align
    summary_ws.cell(row=3, column=col_W, value="USGS").font = bold_green_text_font
    summary_ws.cell(row=3, column=col_W).alignment = center_align
    
    summary_ws.cell(row=16, column=col_V, value=delta18).font = black_bold
    summary_ws.cell(row=16, column=col_V).alignment = center_align
    summary_ws.cell(row=16, column=col_W, value="d18OVSMOW").font = Font(bold=True, color="000000")
    summary_ws.cell(row=16, column=col_W).alignment = center_align

    summary_ws.cell(row=17, column=col_V, value="MRSI W1/W2").font = light_blue_bold
    summary_ws.cell(row=17, column=col_V).alignment = center_align
    summary_ws.cell(row=17, column=col_W, value="USGS").font = Font(bold=True, color="00B050")
    summary_ws.cell(row=17, column=col_W).alignment = center_align

    # --- Copy Groups ---
    def create_group_key(identifier):
        if identifier is None: return None
        if not isinstance(identifier, str): identifier = str(identifier)
        return re.sub(r'\s+[rR]\d+(?:\.\d+)*(?:[a-zA-Z]*)?$', '', identifier).strip()

    def reference_base_key(identifier):
        if identifier is None: return None
        text = str(identifier).upper().strip()
        m_usgs = re.search(r'\bUSGS[- ]?W[- ]?(\d+)\b', text)
        if m_usgs: return f"USGS W-{m_usgs.group(1)}"
        m_mrsi = re.search(r'\b(MRSI)(?:[- ]?(?:STD))?(?:[- ]?W?[- ]?(\d+))?\b', text)
        if m_mrsi:
            base, num = m_mrsi.group(1), m_mrsi.group(2)
            if num:
                if re.search(r'\bSTD\b', text): return f"{base}-STD-W{num}"
                else: return f"{base}-W{num}"
            else: return base
        return None

    dest_start_row = 18 
    group_infos = []
    heco2_dst_ranges = []
    
    src_row = 3
    max_src_row = group_ws.max_row

    # Group Scan Loop
    while src_row <= max_src_row:
        id1_val = group_values_ws.cell(row=src_row, column=3).value
        if id1_val is None:
            src_row += 1
            continue
        data_start_src = src_row
        last_non_empty_data_row = data_start_src
        src_row += 1
        while src_row <= max_src_row:
            marker = group_values_ws.cell(row=src_row, column=10).value
            if isinstance(marker, str) and marker.strip().lower() == "--".lower():
                break
            current_id1_val = group_values_ws.cell(row=src_row, column=3).value
            if current_id1_val is not None:
                last_non_empty_data_row = src_row
            src_row += 1
        data_end_src = last_non_empty_data_row
        calc_mid_src = None
        if src_row <= max_src_row:
            look = src_row
            found = False
            while look <= min(max_src_row, src_row + 3):
                mval = group_values_ws.cell(row=look, column=10).value
                if isinstance(mval, str) and mval.strip().lower() == "--".lower():
                    calc_mid_src = look
                    found = True
                    break
                look += 1
            if not found: calc_mid_src = src_row
        else: calc_mid_src = src_row
        first_id = group_values_ws.cell(row=data_start_src, column=3).value
        base_key = create_group_key(first_id)
        ref_base = reference_base_key(base_key)
        group_infos.append({
            'src_data_start': data_start_src,
            'src_data_end': data_end_src,
            'src_calc_mid': calc_mid_src,
            'base_key': base_key,
            'ref_base': ref_base
        })
        src_row = (calc_mid_src or src_row) + 2

    src_to_dst_row = {}
    dst_row = dest_start_row
    
    # Copy Loop
    for src_r in range(1, group_ws.max_row + 1):
        is_valid_heco2 = (src_r in valid_heco2_src_rows)
        id_val = group_values_ws.cell(row=src_r, column=3).value
        id_color = None
        norm_ref = reference_base_key(id_val)
        if norm_ref and norm_ref in std_color_map:
            id_color = std_color_map[norm_ref]

        for col_idx in range(1, group_ws.max_column + 1):
            src_cell = group_ws.cell(row=src_r, column=col_idx)
            val_cell = group_values_ws.cell(row=src_r, column=col_idx)
            dest_cell = summary_ws.cell(row=dst_row, column=col_idx)
            dest_cell.value = val_cell.value
            try:
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.alignment = copy(src_cell.alignment)
            except Exception: pass
            
            if is_valid_heco2:
                dest_cell.fill = heco2_gray_fill
            if col_idx == 3 and id_color:
                dest_cell.font = id_color

        src_to_dst_row[src_r] = dst_row
        dst_row += 1

    # Map Dst Rows
    for gi in group_infos:
        gi['dst_data_start'] = src_to_dst_row.get(gi['src_data_start'])
        gi['dst_data_end'] = src_to_dst_row.get(gi['src_data_end'])
        gi['dst_calc_mid'] = src_to_dst_row.get(gi['src_calc_mid'])
        
        calc_mid = gi['dst_calc_mid']
        if calc_mid:
            norm_ref = gi['ref_base']
            if norm_ref and norm_ref in std_bold_color_map:
                bold_color = std_bold_color_map[norm_ref]
                for r_offset in [-1, 0]: 
                    target_r = calc_mid + r_offset
                    for c_idx in range(11, 17):
                        cell = summary_ws.cell(row=target_r, column=c_idx)
                        cell.font = bold_color
        
        bk = str(gi['base_key']).lower()
        if "heco2" in bk or "co2" in bk:
            heco2_dst_ranges.append(gi)

    # --- Group Summary Blocks ("All") ---
    ref_groups_by_base = {}
    for gi in group_infos:
        rb = gi['ref_base']
        if rb: ref_groups_by_base.setdefault(rb, []).append(gi)

    insert_actions = []
    for rb, gis in ref_groups_by_base.items():
        if len(gis) >= 1:
            last_gi = sorted(gis, key=lambda x: (x['dst_calc_mid'] or 0))[-1]
            insert_after = (last_gi['dst_calc_mid'] or last_gi['dst_data_end'])
            if insert_after is None: continue
            insert_actions.append({'insert_after': insert_after, 'ref_base': rb, 'groups': gis})

    insert_actions.sort(key=lambda x: x['insert_after'] or 0)

    def build_multi_range(col_letter, groups):
        rngs = []
        for g in groups:
            ds = g.get('dst_data_start')
            de = g.get('dst_data_end')
            if ds and de and de >= ds: rngs.append(f"{col_letter}{ds}:{col_letter}{de}")
        return ",".join(rngs) if rngs else None

    cumulative_offset = 0

    for action in insert_actions:
        groups = action['groups']
        base_insert_at = action['insert_after'] + 1 
        insert_at = base_insert_at + cumulative_offset
        
        summary_ws.insert_rows(insert_at, amount=3)
        cumulative_offset += 3 
        
        label_row = insert_at + 1
        value_row = insert_at + 2
        
        # --- Update: Color "All" header Yellow ---
        cell_all = summary_ws.cell(row=label_row, column=10, value="All")
        cell_all.font = black_bold
        cell_all.fill = yellow_fill # Added fill

        # Other Headers - Already have Yellow Fill
        labels = {11: "Average", 12: "Stdev", 13: "Count", 14: "Average", 15: "Stdev", 16: "Count"}
        for c, txt in labels.items():
            cell = summary_ws.cell(row=label_row, column=c, value=txt)
            cell.font = black_bold
            cell.fill = yellow_fill

        colK = get_column_letter(11)
        colN = get_column_letter(14)
        colQ = get_column_letter(17)

        k_ranges = build_multi_range(colK, groups)
        n_ranges = build_multi_range(colN, groups)
        q_ranges = build_multi_range(colQ, groups)
        
        ref_base = action['ref_base']
        norm_ref = None
        if ref_base:
            text = ref_base.upper().replace(" ", "").replace("-", "")
            if text.startswith("MRSI") and "1" in text: norm_ref = "MRSI-STD-W1"
            elif text.startswith("MRSI") and "2" in text: norm_ref = "MRSI-STD-W2"
            elif text.startswith("USGSW67400"): norm_ref = "USGS W-67400"
            elif text.startswith("USGSW64444"): norm_ref = "USGS W-64444"
            
        stats_font = std_bold_color_map.get(norm_ref, black_bold)

        # --- Update: Force values under headers to be Black (black_bold) ---
        if k_ranges:
            cell = summary_ws.cell(row=value_row, column=11, value=f"=AVERAGE({k_ranges})")
            cell.font = black_bold; cell.number_format = FMT_3_DEC
            cell = summary_ws.cell(row=value_row, column=12, value=f"=STDEV({k_ranges})")
            cell.font = black_bold; cell.number_format = FMT_3_DEC
            cell = summary_ws.cell(row=value_row, column=13, value=f"=SUM({','.join([f'COUNT({colK}{g['dst_data_start']}:{colK}{g['dst_data_end']})' for g in groups])})")
            cell.font = black_bold
        if n_ranges:
            cell = summary_ws.cell(row=value_row, column=14, value=f"=AVERAGE({n_ranges})")
            cell.font = black_bold; cell.number_format = FMT_3_DEC
            cell = summary_ws.cell(row=value_row, column=15, value=f"=STDEV({n_ranges})")
            cell.font = black_bold; cell.number_format = FMT_3_DEC
            cell = summary_ws.cell(row=value_row, column=16, value=f"=SUM({','.join([f'COUNT({colN}{g['dst_data_start']}:{colN}{g['dst_data_end']})' for g in groups])})")
            cell.font = black_bold
        if q_ranges:
            cell = summary_ws.cell(row=value_row, column=17, value=f"=AVERAGE({q_ranges})")
            cell.font = black_bold; cell.number_format = FMT_3_DEC

        for gi in group_infos:
            for key in ['dst_data_start', 'dst_data_end', 'dst_calc_mid']:
                if gi[key] is not None and gi[key] >= insert_at:
                    gi[key] += 3

        if norm_ref in ["MRSI-STD-W1", "MRSI-STD-W2", "USGS W-67400", "USGS W-64444"]:
            target_row_map = {
                "MRSI-STD-W1": 5, "MRSI-STD-W2": 6, 
                "USGS W-67400": 7, "USGS W-64444": 8
            }
            target_r = target_row_map[norm_ref]
            color_f = std_bold_color_map[norm_ref]
            c = summary_ws.cell(row=target_r, column=11, value=f"=K{value_row}")
            c.font = color_f; c.number_format = FMT_3_DEC
            c = summary_ws.cell(row=target_r, column=14, value=f"=N{value_row}")
            c.font = color_f; c.number_format = FMT_3_DEC
    
    # --- Fill in He/CO2 Box (Row 9) ---
    if heco2_dst_ranges:
        # Instead of recalculating, we reference the 'dst_calc_mid' row
        # which contains the pre-calculated values copied from the Group sheet.
        
        # Get the row numbers where the He/CO2 averages are located in this sheet
        calc_rows = [g['dst_calc_mid'] for g in heco2_dst_ranges if g.get('dst_calc_mid')]
        
        if calc_rows:
            colK = get_column_letter(11) # C Avg
            colL = get_column_letter(12) # C Stdev
            colN = get_column_letter(14) # O Avg
            colO = get_column_letter(15) # O Stdev

            # Helper to create the reference (e.g., "=K45" or "=AVERAGE(K45,K80)")
            def make_ref(col_let, rows):
                refs = [f"{col_let}{r}" for r in rows]
                if len(refs) == 1:
                    return f"={refs[0]}"
                return f"=AVERAGE({','.join(refs)})"

            # K9: C Avg
            c = summary_ws["K9"]
            c.value = make_ref(colK, calc_rows)
            c.number_format = FMT_3_DEC
            c.font = black_bold

            # L9: C Stdev
            c = summary_ws["L9"]
            c.value = make_ref(colL, calc_rows) 
            c.number_format = FMT_3_DEC
            c.font = black_bold

            # N9: O Avg
            c = summary_ws["N9"]
            c.value = make_ref(colN, calc_rows)
            c.number_format = FMT_3_DEC
            c.font = black_bold

            # O9: O Stdev
            c = summary_ws["O9"]
            c.value = make_ref(colO, calc_rows)
            c.number_format = FMT_3_DEC
            c.font = black_bold

    # --- Match widths ---
    for col in range(1, group_ws.max_column + 1):
        col_letter = get_column_letter(col)
        src_dim = group_ws.column_dimensions[col_letter]
        summary_ws.column_dimensions[col_letter].width = src_dim.width or 15
    summary_ws.column_dimensions['C'].width = 30
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['R'].width = 16
    summary_ws.freeze_panes = "B19"

    # --- Calculations Columns U..X ---
    standard_ranges = {
        "MRSI-STD-W1": {'col': col_V, 'ranges': []},
        "MRSI-STD-W2": {'col': col_V, 'ranges': []},
        "USGS W-67400": {'col': col_W, 'ranges': []},
        "USGS W-64444": {'col': col_W, 'ranges': []}
    }
    col_N_str = get_column_letter(14)
    
    for gi in group_infos:
        ds = gi.get('dst_data_start')
        de = gi.get('dst_data_end')
        if not ds or not de or de < ds: continue

        for r in range(ds, de + 1):
            id_val = summary_ws.cell(row=r, column=3).value
            skip_row = False
            if isinstance(id_val, str):
                normalized_val = id_val.strip().upper()
                if normalized_val.startswith("HECO2") or normalized_val.startswith("CO2"): skip_row = True
            
            if not skip_row: summary_ws.cell(row=r, column=col_S, value=id_val)
            else: summary_ws.cell(row=r, column=col_S, value=None)
            
            if skip_row:
                summary_ws.cell(row=r, column=col_V, value=None)
                summary_ws.cell(row=r, column=col_W, value=None)
                continue

            ref = reference_base_key(id_val)
            target_col = col_V 
            slope_cell = "$N$11"
            intercept_cell = "$N$12"
            norm_ref = None
            
            if ref:
                text = ref.upper().replace(" ", "").replace("-", "")
                if text.startswith("MRSI") and "1" in text:
                    norm_ref = "MRSI-STD-W1"; slope_cell = "$N$11"; intercept_cell = "$N$12"; target_col = col_V
                elif text.startswith("MRSI") and "2" in text:
                    norm_ref = "MRSI-STD-W2"; slope_cell = "$N$11"; intercept_cell = "$N$12"; target_col = col_V
                elif text.startswith("USGSW67400"):
                    norm_ref = "USGS W-67400"; slope_cell = "$N$14"; intercept_cell = "$N$15"; target_col = col_W
                elif text.startswith("USGSW64444"):
                    norm_ref = "USGS W-64444"; slope_cell = "$N$14"; intercept_cell = "$N$15"; target_col = col_W
            
            v_formula = f"=IF({col_N_str}{r}=\"\",\"\",{col_N_str}{r}*{slope_cell}+{intercept_cell})"
            c = summary_ws.cell(row=r, column=target_col, value=v_formula)
            c.number_format = FMT_2_DEC

            if target_col == col_V: summary_ws.cell(row=r, column=col_W, value=None)
            else: summary_ws.cell(row=r, column=col_V, value=None)
            
            if norm_ref: standard_ranges[norm_ref]['ranges'].append(r)

        calc_mid = gi.get('dst_calc_mid')
        if calc_mid and isinstance(calc_mid, int): avg_row = calc_mid + 1
        else: avg_row = de + 2

        if summary_ws.cell(row=avg_row, column=col_S).value is not None:
            summary_ws.insert_rows(avg_row, amount=1)
            for gi2 in group_infos:
                for key in ['dst_data_start', 'dst_data_end', 'dst_calc_mid']:
                    if gi2.get(key) is not None and gi2[key] >= avg_row: gi2[key] += 1
            if de >= avg_row: de += 1
            if calc_mid and calc_mid >= avg_row: calc_mid += 1

        summary_ws.cell(row=avg_row - 1, column=col_S, value="Average/STDEV").font = Font(bold=True)
        v_col_letter = get_column_letter(col_V)
        w_col_letter = get_column_letter(col_W)
        v_range = f"{v_col_letter}{ds}:{v_col_letter}{de}"
        w_range = f"{w_col_letter}{ds}:{w_col_letter}{de}"

        c = summary_ws.cell(row=avg_row - 1, column=col_V, value=f"=IF(COUNT({v_range})=0,\"\",AVERAGE({v_range}))")
        c.font = Font(bold=True); c.number_format = FMT_2_DEC
        c = summary_ws.cell(row=avg_row - 1, column=col_W, value=f"=IF(COUNT({w_range})=0,\"\",AVERAGE({w_range}))")
        c.font = Font(bold=True); c.number_format = FMT_2_DEC
        
        data_col = "V" if v_range else "W"
        data_range = v_range if v_range else w_range
        c = summary_ws.cell(row=avg_row - 1, column=col_X, value=f'=IF(COUNT({data_range})=0,"",STDEV({data_range}))')
        c.font = Font(bold=True); c.number_format = FMT_2_DEC
        
        # 5. Add COUNT() in Column Y
        c = summary_ws.cell(row=avg_row - 1, column=col_Y, value=f'=COUNT({data_range})')
        c.font = Font(bold=True); c.alignment = center_align
        
        # 5. Apply Blue Box Color to the calculation block (S to Y)
        for col_idx in range(col_S, col_Y + 1):
             summary_ws.cell(row=avg_row - 1, column=col_idx).fill = box_blue_fill
        
    # --- Conditional Formatting (Columns L and O) ---
    if stdev_threshold is not None:
        thresh_str = str(stdev_threshold)
        col_L = "L"
        col_O = "O"
        
        for gi in group_infos:
            ds = gi.get('dst_data_start')
            de = gi.get('dst_data_end')
            calc_row = gi.get('dst_calc_mid')
            
            # Ensure valid rows and respect "below row 19" (>= 19)
            if not ds or not de or not calc_row:
                continue
            
            # Only apply if data starts at row 19 or greater
            # (or if you want to include row 18, change 19 to 18)
            if ds < 19:
                continue

            # Data Ranges
            rng_L = f"{col_L}{ds}:{col_L}{de}"
            rng_O = f"{col_O}{ds}:{col_O}{de}"
            
            summary_ws.conditional_formatting.add(rng_L, CellIsRule(operator="greaterThan", formula=[thresh_str], fill=fill_error))
            summary_ws.conditional_formatting.add(rng_O, CellIsRule(operator="greaterThan", formula=[thresh_str], fill=fill_error))
            
            # Calc Row
            cell_L = f"{col_L}{calc_row}"
            cell_O = f"{col_O}{calc_row}"
            summary_ws.conditional_formatting.add(cell_L, CellIsRule(operator="greaterThan", formula=[thresh_str], fill=fill_error))
            summary_ws.conditional_formatting.add(cell_O, CellIsRule(operator="greaterThan", formula=[thresh_str], fill=fill_error))

    # --- Calibration Summary Boxes Logic ---
    standard_cells = {
        "MRSI-STD-W1": {'avg_col': col_V, 'stdev_col': col_X, 'count_col': col_Y, 'row': 5},
        "MRSI-STD-W2": {'avg_col': col_V, 'stdev_col': col_X, 'count_col': col_Y, 'row': 6},
        "USGS W-67400": {'avg_col': col_W, 'stdev_col': col_X, 'count_col': col_Y, 'row': 7},
        "USGS W-64444": {'avg_col': col_W, 'stdev_col': col_X, 'count_col': col_Y, 'row': 8},
    }

    row_text_styles = {5: "FF0000", 6: "0000FF", 7: "E46C0A", 8: "00B050"}

    for norm_ref, data in standard_ranges.items():
        cell_info = standard_cells.get(norm_ref)
        if not cell_info: continue
        target_col = get_column_letter(data['col'])
        target_row = cell_info['row']
        
        # Apply formulas to the summary boxes R5:Y8
        if data['ranges']:
            target_ranges = ",".join([f"{target_col}{r}" for r in data['ranges']])
            
            c = summary_ws.cell(row=target_row, column=cell_info['avg_col'], value=f"=IF(COUNT({target_ranges})=0,\"\",AVERAGE({target_ranges}))")
            c.number_format = FMT_2_DEC
            
            c = summary_ws.cell(row=target_row, column=cell_info['stdev_col'], value=f"=IF(COUNT({target_ranges})=0,\"\",STDEV({target_ranges}))")
            c.number_format = FMT_2_DEC
            
            summary_ws.cell(row=target_row, column=cell_info['count_col'], value=f"=COUNT({target_ranges})")

        # Apply specific font color and center alignment
        color = row_text_styles.get(target_row)
        if color:
            for c in range(col_R, col_Y + 1):
                cell = summary_ws.cell(row=target_row, column=c)
                cell.font = Font(color=color)
                cell.alignment = center_align
            
    for ws in wb.worksheets: ws.sheet_view.tabSelected = False
    summary_ws.sheet_view.tabSelected = True
    wb.active = wb.index(summary_ws)
    wb.save(file_path)
    print(f"✅ Step 6: Normalization completed on {file_path}")