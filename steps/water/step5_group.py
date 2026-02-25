import pandas as pd
import string
import statistics
import re
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.formatting.rule import CellIsRule 
import settings 
from utils import embed_settings_popup

# --- Constants & Styles ---
fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
heco2_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
header_fill = PatternFill(start_color="8ED973", end_color="8ED973", fill_type="solid")
bold_font = Font(bold=True)
THREE_DECIMAL_FORMAT = "0.000"

# --- Helper Functions ---

def extract_run_number(identifier):
    """Parses strings like 'r1', 'r2.1' into (major, minor)."""
    if not identifier: return None, None
    match = re.search(r'(?i)[rR](\d+)(?:\.(\d+))?', str(identifier))
    if match:
        major = int(match.group(1))
        minor = int(match.group(2)) if match.group(2) else 0
        return major, minor
    return None, None

def get_valid_heco2_source_rows(df):
    """Finds valid HeCO2 rows (lowest minor for each major, skipping r1)."""
    valid_source_rows = set()
    seen_majors = {} 
    
    for _, row in df.iterrows():
        ident = str(row['Identifier 1']).strip().lower()
        if 'heco2' in ident or 'co2' in ident:
            major, minor = extract_run_number(ident)
            if major is None or major == 1: continue
            
            if major not in seen_majors:
                seen_majors[major] = (minor, row['Source_Row'])
            else:
                prev_minor, prev_row = seen_majors[major]
                if minor < prev_minor:
                    seen_majors[major] = (minor, row['Source_Row'])
                    
    for _, s_row in seen_majors.values():
        valid_source_rows.add(s_row)
    return valid_source_rows

def get_stats_and_bounds(values, sigma=2):
    """Calculates Mean, Stdev, and Sigma Bounds (default 2-sigma)."""
    clean_vals = []
    for v in values:
        try:
            if v is not None:
                f = float(v)
                clean_vals.append(f)
        except (ValueError, TypeError):
            continue
            
    count = len(clean_vals)
    if count == 0:
        return None, None, None, None
    
    avg = statistics.mean(clean_vals)
    if count > 1:
        std = statistics.stdev(clean_vals)
    else:
        std = 0.0
        
    lower = avg - (sigma * std)
    upper = avg + (sigma * std)
    return avg, std, lower, upper

def copy_cell_style(src_cell, tgt_cell, strike=False):
    """Copies style. If strike=True, applies Red Bold Strikethrough."""
    if src_cell is None or tgt_cell is None: return
    if src_cell.has_style:
        try:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = src_cell.number_format 
            tgt_cell.alignment = copy(src_cell.alignment)
            tgt_cell.protection = copy(src_cell.protection)
            
            if strike:
                new_font = copy(src_cell.font)
                new_font.strike = True
                new_font.color = "FF0000"
                new_font.bold = True
                tgt_cell.font = new_font
        except: pass

def step5_group_water(file_path: str):
    """
    Step 5: Group Water
    - Groups HeCO2/Refs/Others
    - Validates HeCO2 rows (Gray Highlight)
    - 2-Sigma Outlier Detection (Red Strikethrough)
      - Supports "Individual" or "Exclude Row" modes from settings.
    - Calculates "All" and "Outlier Excl." Stats
    """
    
    # --- 1. Settings ---
    stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    if stdev_threshold is None: stdev_threshold = 0.08

    # NEW: Get Outlier Settings
    sigma_val = settings.get_setting("OUTLIER_SIGMA")
    if not isinstance(sigma_val, (int, float)): sigma_val = 2
    
    exclusion_mode = settings.get_setting("OUTLIER_EXCLUSION_MODE")
    if not exclusion_mode: exclusion_mode = "Individual" # Default to Individual if missing

    # --- 2. Load Workbook ---
    wb_values = load_workbook(file_path, data_only=True)
    wb = load_workbook(file_path)

    source_sheet_name = "Last 6_DNT"
    matched_source = next((s for s in wb.sheetnames if s.lower() == source_sheet_name.lower()), None)
    if matched_source is None:
        raise ValueError(f"Sheet '{source_sheet_name}' not found.")

    source_ws_vals = wb_values[matched_source]
    source_ws = wb[matched_source]

    new_sheet_name = "Group_DNT"
    if new_sheet_name in wb.sheetnames: del wb[new_sheet_name]

    # Column Mapping based on provided list
    # Col 11 (C) corresponds to Source Col 16 (Index 10)
    # Col 14 (O) corresponds to Source Col 19 (Index 13)
    source_cols_to_read = [1, 2, 3, 4, 5, 6, 7, 10, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    
    # Extract Data for sorting
    data = []
    for row_idx in range(2, source_ws.max_row + 1):
        if not source_ws.row_dimensions[row_idx].hidden:
            row_values = [source_ws_vals.cell(row=row_idx, column=c).value for c in source_cols_to_read]
            row_values.append(row_idx)
            data.append(row_values)

    # Clean headers
    raw_header_vals = [source_ws_vals.cell(row=1, column=c).value for c in source_cols_to_read]
    header_vals = [str(v).strip() if v is not None else "" for v in raw_header_vals]
    
    df = pd.DataFrame(data, columns=header_vals + ['Source_Row'])

    new_headers = [
        'Line', 'Time Code', 'Identifier 1', 'Comment', 'Identifier 2', 'Analysis', 'Preparation',
        'Ampl 44', '', '', 'C avg', 'C stdev', '', 'O avg', 'O stdev', '', 'Sum area all'
    ]

    # --- 3. Grouping Logic ---
    def create_group_key(ident):
        if pd.isna(ident) or not isinstance(ident, str): return ident
        return re.sub(r'\s+[rR]\d+(?:\.\d+)*(?:[a-zA-Z]*)?$', '', ident).strip()

    df['Group_Key'] = df['Identifier 1'].apply(create_group_key)
    
    if 'Line' in df.columns:
        df['Line_num'] = pd.to_numeric(df['Line'], errors='coerce')
    else:
        df['Line_num'] = 0

    group_min = df.groupby('Group_Key', sort=False)['Line_num'].min().reset_index(name='min_line')

    def is_heco2(k): return bool(re.search(r'(?i)\b(heco2|co2)\b', str(k))) if k else False
    def is_ref(k):
        pat = [r'\bMRSI\b', r'\bMRSI[- ]?\d+\b', r'\bMRSI[- ]?STD', r'\bUSGS']
        return any(re.search(p, str(k).upper()) for p in pat) if k else False

    group_min['is_heco2'] = group_min['Group_Key'].apply(is_heco2)
    group_min['is_ref'] = group_min['Group_Key'].apply(is_ref)

    heco2 = group_min[group_min['is_heco2']].sort_values(by=['min_line'])
    refs = group_min[~group_min['is_heco2'] & group_min['is_ref']].sort_values(by=['min_line'])
    others = group_min[~group_min['is_heco2'] & ~group_min['is_ref']].sort_values(by=['min_line'])

    ordered_groups = pd.concat([heco2, refs, others])['Group_Key'].tolist()
    df['Group_Key'] = pd.Categorical(df['Group_Key'], categories=ordered_groups, ordered=True)
    df_sorted = df.sort_values(by=['Group_Key', 'Line_num', 'Source_Row']).reset_index(drop=True)
    
    valid_heco2_rows = get_valid_heco2_source_rows(df)

    # --- 4. Write to Excel ---
    target_sheet = "Pre-Group_DNT"
    idx = wb.index(wb[target_sheet]) if target_sheet in wb.sheetnames else wb.index(source_ws)
    new_ws = wb.create_sheet(new_sheet_name, idx)

    for c, h in enumerate(new_headers, 1):
        cell = new_ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
    new_ws.freeze_panes = "A2"
    
    for t_idx, s_idx in enumerate(source_cols_to_read, 1):
        try:
            w = source_ws.column_dimensions[get_column_letter(s_idx)].width
            if w: new_ws.column_dimensions[get_column_letter(t_idx)].width = w
        except: pass

    cur_row = 3
    last_ref_key = refs['Group_Key'].iloc[-1] if not refs.empty else None

    # Group Iteration
    grouped = df_sorted.groupby('Group_Key', sort=False)
    
    for group_key, group in grouped:
        start_row = cur_row
        
        is_heco2_group = is_heco2(group_key)
        
        candidates_data = [] 
        
        # Pre-scan group
        for _, row_s in group.iterrows():
            s_idx = int(row_s['Source_Row']) if pd.notna(row_s['Source_Row']) else None
            is_valid_heco2 = (s_idx in valid_heco2_rows)
            
            is_candidate = True
            if is_heco2_group and not is_valid_heco2:
                is_candidate = False
            
            if is_candidate and s_idx is not None:
                c_val = source_ws_vals.cell(row=s_idx, column=16).value
                o_val = source_ws_vals.cell(row=s_idx, column=19).value
                candidates_data.append({'s_idx': s_idx, 'c': c_val, 'o': o_val})
        
        # --- B. Calculate Outliers ---
        c_list = [x['c'] for x in candidates_data]
        o_list = [x['o'] for x in candidates_data]
        
        # Use Dynamic Sigma from Settings
        mu_c, sig_c, low_c, high_c = get_stats_and_bounds(c_list, sigma=sigma_val)
        mu_o, sig_o, low_o, high_o = get_stats_and_bounds(o_list, sigma=sigma_val)
        
        # Identify outliers specifically for C and O
        outliers_c = set()
        outliers_o = set()
        
        for item in candidates_data:
            s_idx = item['s_idx']
            
            # Check C
            try:
                cv = float(item['c'])
                if low_c is not None and (cv < low_c or cv > high_c):
                    outliers_c.add(s_idx)
            except: pass
            
            # Check O
            try:
                ov = float(item['o'])
                if low_o is not None and (ov < low_o or ov > high_o):
                    outliers_o.add(s_idx)
            except: pass

        # Handle "Exclude Row" Mode
        # If "Exclude Row", if it's an outlier in ANY column, it's an outlier in ALL columns.
        if exclusion_mode == "Exclude Row":
            combined_outliers = outliers_c.union(outliers_o)
            outliers_c = combined_outliers
            outliers_o = combined_outliers

        # --- C. Write Rows ---
        all_candidate_rows = [] 
        valid_dest_rows_c = [] # Rows valid for C stats
        valid_dest_rows_o = [] # Rows valid for O stats
        
        for _, row_s in group.iterrows():
            s_idx = int(row_s['Source_Row']) if pd.notna(row_s['Source_Row']) else None
            is_valid_heco2 = (s_idx in valid_heco2_rows)
            
            # Determine Outlier Status for this specific row
            is_c_outlier = (s_idx in outliers_c)
            is_o_outlier = (s_idx in outliers_o)
            
            # Track Destination Rows for Formulas
            is_candidate = True
            if is_heco2_group and not is_valid_heco2: is_candidate = False

            if is_candidate:
                all_candidate_rows.append(cur_row)
                if not is_c_outlier: valid_dest_rows_c.append(cur_row)
                if not is_o_outlier: valid_dest_rows_o.append(cur_row)

            # Write Columns
            for col_off, header in enumerate(new_headers, 1):
                val = None
                if s_idx:
                    src_col = source_cols_to_read[col_off - 1]
                    val = source_ws_vals.cell(row=s_idx, column=src_col).value
                else:
                    val = row_s[header]
                
                cell = new_ws.cell(row=cur_row, column=col_off, value=val)
                
                if s_idx:
                    src_cell = source_ws.cell(row=s_idx, column=source_cols_to_read[col_off - 1])
                    
                    # Logic for Strikethrough
                    # If it's a C-related column (C avg=11, C stdev=12), strike if C outlier
                    # If it's an O-related column (O avg=14, O stdev=15), strike if O outlier
                    # If it's identifier/common, strike only if "Exclude Row" is active and it's an outlier? 
                    # Usually, we just strike the value itself.
                    
                    strike_this = False
                    if col_off in [11, 12]: # C Columns
                        if is_c_outlier: strike_this = True
                    elif col_off in [14, 15]: # O Columns
                        if is_o_outlier: strike_this = True
                    
                    copy_cell_style(src_cell, cell, strike=strike_this)
                
                if is_valid_heco2:
                    cell.fill = heco2_gray_fill

            # Set Row Height
            if s_idx:
                rh = source_ws.row_dimensions[s_idx].height
                if rh: new_ws.row_dimensions[cur_row].height = rh
            
            cur_row += 1

        # --- D. Write Stats Blocks ---
        
        # Helper to build range string
        def build_rng(rows, col_idx):
            let = get_column_letter(col_idx)
            return ",".join([f"{let}{r}" for r in rows])

        # 1. Block: "All" (Includes outliers)
        row_all = cur_row
        new_ws.cell(row=row_all, column=10, value="--").font = bold_font
        for i, t in enumerate(["Average", "Stdev", "Count"]):
            new_ws.cell(row=row_all, column=11+i, value=t).font = bold_font
            new_ws.cell(row=row_all, column=14+i, value=t).font = bold_font
            
        row_all_calc = row_all + 1
        
        if all_candidate_rows:
            rng_c = build_rng(all_candidate_rows, 11)
            rng_o = build_rng(all_candidate_rows, 14)
            rng_q = build_rng(all_candidate_rows, 17)
            
            new_ws.cell(row=row_all_calc, column=11, value=f"=AVERAGE({rng_c})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=row_all_calc, column=12, value=f"=STDEV({rng_c})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=row_all_calc, column=13, value=f"=COUNT({rng_c})").number_format = "0"
            
            new_ws.cell(row=row_all_calc, column=14, value=f"=AVERAGE({rng_o})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=row_all_calc, column=15, value=f"=STDEV({rng_o})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=row_all_calc, column=16, value=f"=COUNT({rng_o})").number_format = "0"
            
            new_ws.cell(row=row_all_calc, column=17, value=f"=AVERAGE({rng_q})").number_format = THREE_DECIMAL_FORMAT
        
        for c in range(11, 18): new_ws.cell(row=row_all_calc, column=c).font = bold_font
            
        # 2. Block: "Outlier Excl." (Excludes specific outliers per column)
        row_filt = row_all_calc + 2
        new_ws.cell(row=row_filt, column=10, value="Outlier Excl.").font = bold_font
        for i, t in enumerate(["Average", "Stdev", "Count"]):
            new_ws.cell(row=row_filt, column=11+i, value=t).font = bold_font
            new_ws.cell(row=row_filt, column=14+i, value=t).font = bold_font
            
        row_filt_calc = row_filt + 1
        
        # Calculate C Stats (using valid C rows)
        if valid_dest_rows_c:
            rng_c_filt = build_rng(valid_dest_rows_c, 11)
            new_ws.cell(row=row_filt_calc, column=11, value=f"=AVERAGE({rng_c_filt})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=row_filt_calc, column=12, value=f"=STDEV({rng_c_filt})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=row_filt_calc, column=13, value=f"=COUNT({rng_c_filt})").number_format = "0"
        else:
            new_ws.cell(row=row_filt_calc, column=11, value="--")

        # Calculate O Stats (using valid O rows)
        if valid_dest_rows_o:
            rng_o_filt = build_rng(valid_dest_rows_o, 14)
            new_ws.cell(row=row_filt_calc, column=14, value=f"=AVERAGE({rng_o_filt})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=row_filt_calc, column=15, value=f"=STDEV({rng_o_filt})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=row_filt_calc, column=16, value=f"=COUNT({rng_o_filt})").number_format = "0"
        else:
            new_ws.cell(row=row_filt_calc, column=14, value="--")

        # Calculate Q Stats (using valid C rows as default proxy, or union? Usually follows C or O. Let's use valid_c for now)
        if valid_dest_rows_c:
             rng_q_filt = build_rng(valid_dest_rows_c, 17)
             new_ws.cell(row=row_filt_calc, column=17, value=f"=AVERAGE({rng_q_filt})").number_format = THREE_DECIMAL_FORMAT

        for c in range(11, 18): new_ws.cell(row=row_filt_calc, column=c).font = bold_font

        # -------------------- CONDITIONAL FORMATTING --------------------
        if stdev_threshold:
            thresh = str(stdev_threshold)
            new_ws.conditional_formatting.add(f"L{row_all_calc}", CellIsRule(operator="greaterThan", formula=[thresh], fill=fill_error))
            new_ws.conditional_formatting.add(f"O{row_all_calc}", CellIsRule(operator="greaterThan", formula=[thresh], fill=fill_error))
            new_ws.conditional_formatting.add(f"L{row_filt_calc}", CellIsRule(operator="greaterThan", formula=[thresh], fill=fill_error))
            new_ws.conditional_formatting.add(f"O{row_filt_calc}", CellIsRule(operator="greaterThan", formula=[thresh], fill=fill_error))

        cur_row = row_filt_calc + 2

        if group_key == last_ref_key:
            cur_row += 1
            for _ in range(2):
                for c in range(1, 18):
                    new_ws.cell(row=cur_row, column=c).fill = gray_fill
                cur_row += 1
            cur_row += 1

    try:
        for r_idx in list(new_ws.row_dimensions.keys()):
            new_ws.row_dimensions[r_idx].outlineLevel = 0
    except: pass

    for s in wb.worksheets: s.sheet_view.tabSelected = False
    new_ws.sheet_view.tabSelected = True
    wb.active = wb.index(new_ws)
    new_ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Add Settings Popup Comment
    embed_settings_popup(new_ws, "R1")

    new_ws.column_dimensions["J"].width = 16 

    wb.save(file_path)
    print(f"✅ Step 5: Group sheet '{new_sheet_name}'")