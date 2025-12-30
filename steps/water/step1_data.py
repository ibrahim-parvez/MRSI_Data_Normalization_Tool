import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, NamedStyle # NamedStyle is still imported but not used for number formatting
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.views import Selection
import settings # <--- Using your required settings import

def step1_data_water(file_path, sheet_name='Default_Gas_Bench.wke', sparkline=False):
    """
    Step 1: DATA for Water
    Reads the Excel file, transforms it (padded rows, formulas, formatting),
    adds calculation rectangle and styling, and saves the file.
    
    FIXES:
    1. Removed reliance on NamedStyle for number formatting.
    2. Explicitly set cell.number_format = '0.000' on calculated cells (P, Q, S, T) and
       '0.00' on column V (Sum area all).
    3. Retained the explicit fill application to prevent the calc box from turning white.
    """
    new_sheet_name = 'Data_DNT'
    
    # --- Configuration & Styles ---
    # Use global threshold from settings module
    stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    
    # Read original data
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    
    # Headers for the main data
    headers = [
        'Line', 'Time Code', 'Identifier 1', 'Comment', 'Identifier 2', 'Analysis',
        'Preparation', 'Peak Nr', 'Rt', 'Ampl 44', 'Area All',
        'd 13C/12C', 'd 18O/16O'
    ]
    
    # Calculation rectangle headers (Flags have blank headers)
    calc_headers = [
        'C avg', 'C stdev', '', 'O avg', 'O stdev', '',
        'Sum area all', 'funny peaks', 'min intensity'
    ]
    
    # Load workbook and remove old sheet if exists
    wb = load_workbook(file_path)
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    
    # Insert new sheet to the left of original
    first_index = wb.index(wb[sheet_name])
    ws = wb.create_sheet(new_sheet_name, first_index)
    
    # Ensure only new sheet is selected
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    ws.sheet_view.tabSelected = True
    wb.active = wb.index(ws)
    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    
    # Fill colors
    header_fill_green = PatternFill(start_color="8ed973", end_color="8ed973", fill_type="solid")
    header_fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    calc_fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    flag_fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light Red
    
    # New red fill for Conditional Formatting error
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Define the explicit number format strings
    NUM_FORMAT_3_DECIMAL = '0.000'
    NUM_FORMAT_2_DECIMAL = '0.00'

    # --- Write full header row (extend across many columns) ---
    label_col = 15 # O = label column inside calc box
    start_col_calc = label_col
    total_cols = ws.max_column + 1000 # extend header fill far beyond data range
    
    # yellow columns target (H=8,I=9,K=11,L=12,M=13,W=23,X=24)
    yellow_cols = [8, 9, 11, 12, 13, 23, 24]
    
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill_yellow if col in yellow_cols else header_fill_green
    
    # Write headers A → M (actual header labels)
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    
    # Write calc metric headers P → X (shifted right by one so they sit in P..X)
    for i, h in enumerate(calc_headers, start=start_col_calc + 1):
        ws.cell(row=1, column=i, value=h)
        
    # Start writing data from row 3 (blank line at 2)
    cur_row = 3
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    
    # Normalize column names for matching
    def normalize_name(s):
        if s is None:
            return ''
        return ' '.join(str(s).split()).lower()
        
    df_cols_norm = {normalize_name(c): c for c in df.columns}
    header_to_dfcol = {h: df_cols_norm.get(normalize_name(h)) for h in headers}
    
    # Important columns (in the generated sheet A..M)
    col_area = col_map.get('Area All') # K (11)
    col_c = col_map.get('d 13C/12C') # L (12)
    col_o = col_map.get('d 18O/16O') # M (13)
    col_ampl = col_map.get('Ampl 44') # J (10)
    col_letter_area = get_column_letter(col_area) if col_area else None # e.g. 'K'
    col_letter_c = get_column_letter(col_c) if col_c else None # e.g. 'L'
    col_letter_o = get_column_letter(col_o) if col_o else None # e.g. 'M'
    col_letter_ampl = get_column_letter(col_ampl) if col_ampl else None # e.g. 'J'
    
    # Group by Line (preserve order)
    grouped = df.groupby('Line', sort=False)
    
    # For convenience: calc box start/end cols (now include X)
    calc_start_col = start_col_calc # 15 -> O (label column included)
    calc_end_col = calc_start_col + len(calc_headers) # 15 + 9 = 24 -> X included
    
    # --- Helper to write transparent separator row (no fill) across A..calc_end_col ---
    def write_transparent_separator(sheet, row_idx, last_col):
        for c in range(1, last_col + 1):
            cell = sheet.cell(row=row_idx, column=c)
            cell.value = None
        return
        
    # --- Write grouped data blocks with calc box formulas ---
    for line, group in grouped:
        first_data_row = cur_row
        num_data_rows = len(group) # Actual number of data rows
        expected_rows = 11
        num_rows_to_write = max(expected_rows, num_data_rows) # Pad up to 11 if less
        last_data_row = first_data_row + num_rows_to_write - 1
        
        # Row for the flag (r1)
        r_flag = first_data_row
        
        # Check if the number of data rows is 11 (the expected count)
        is_flagged = num_data_rows != expected_rows
        
        # Determine flag message
        if num_data_rows < expected_rows:
            flag_message = f"<{expected_rows} ({num_data_rows})"
        elif num_data_rows > expected_rows:
            flag_message = f">{expected_rows} ({num_data_rows})"
        else:
            flag_message = "" # No flag needed
        
        # positions used in formulas
        ref_rows = [first_data_row, first_data_row + 1, first_data_row + 3] # first, second, fourth
        first_after_ref = first_data_row + 4
        last6_start = max(first_data_row, last_data_row - 5)
        
        # Write each row of group (A→M) and fill pink box (O→X)
        for i in range(num_rows_to_write):
            row_idx = cur_row
            if i < num_data_rows:
                row_dict = group.iloc[i].to_dict()
            else:
                # pad with same Line/Time/Identifier1 for missing rows
                row_dict = {c: None for c in df.columns}
                for key in ['Line', 'Time Code', 'Identifier 1']:
                    row_dict[key] = group.iloc[0][key]
            
            # write data columns A..M
            for h in headers:
                col = col_map[h]
                val = row_dict.get(header_to_dfcol[h]) if header_to_dfcol[h] else None
                ws.cell(row=row_idx, column=col, value=val)
            
            # fill pink box O..X for this data row (include label col O)
            for c in range(calc_start_col, calc_end_col + 1):
                ws.cell(row=row_idx, column=c).fill = calc_fill
                
            cur_row += 1
            
        # Now fill the pink-box *content* (formulas / labels) for the 11-row block relative to first_data_row
        # Map relative positions 1..11 -> absolute rows
        r1 = first_data_row
        r2 = first_data_row + 1
        r3 = first_data_row + 2
        r4 = first_data_row + 3
        r5 = first_data_row + 4
        r6 = first_data_row + 5 # The 'last 6' row
        r7 = first_data_row + 6
        r8 = first_data_row + 7
        r9 = first_data_row + 8
        r10 = first_data_row + 9
        r11 = first_data_row + 10
        
        # Column O (label column)
        cell = ws.cell(row=r1, column=calc_start_col, value="Calculations")
        cell.font = Font(bold=True)
        cell.fill = calc_fill # Retain fill
        ws.cell(row=r3, column=calc_start_col, value="ref avg").fill = calc_fill # Retain fill
        ws.cell(row=r5, column=calc_start_col, value="all").fill = calc_fill # Retain fill
        ws.cell(row=r6, column=calc_start_col, value="last 6").fill = calc_fill # Retain fill
        ws.cell(row=r7, column=calc_start_col, value="sparkline").fill = calc_fill # Retain fill
        ws.cell(row=r8, column=calc_start_col, value="Amp 44").fill = calc_fill # Retain fill
        ws.cell(row=r9, column=calc_start_col, value="start 6").fill = calc_fill # Retain fill
        ws.cell(row=r10, column=calc_start_col, value="end 11").fill = calc_fill # Retain fill
        ws.cell(row=r11, column=calc_start_col, value="delta").fill = calc_fill # Retain fill
        
        # Column P (C avg) -> references col_letter_c (L)
        col_P = calc_start_col + 1 # P
        if col_letter_c:
            # ref avg (first, second, fourth)
            cell = ws.cell(row=r3, column=col_P,
                           value=f"=AVERAGE({col_letter_c}{ref_rows[0]},{col_letter_c}{ref_rows[1]},{col_letter_c}{ref_rows[2]})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # all: average of all values AFTER first 4
            cell = ws.cell(row=r5, column=col_P,
                           value=f"=AVERAGE({col_letter_c}{first_after_ref}:{col_letter_c}{last_data_row})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # last 6
            cell = ws.cell(row=r6, column=col_P,
                           value=f"=AVERAGE({col_letter_c}{last6_start}:{col_letter_c}{last_data_row})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # sparkline (optional)
            if sparkline:
                cell = ws.cell(row=r7, column=col_P,
                                value=f"=AVERAGE({col_letter_c}{first_after_ref}:{col_letter_c}{last_data_row})")
                cell.number_format = NUM_FORMAT_3_DECIMAL
                cell.fill = calc_fill # Retain fill
            # start 6 = first value of last 6
            cell = ws.cell(row=r9, column=col_P, value=f"={col_letter_c}{last6_start}")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # end 11 = last value of group
            cell = ws.cell(row=r10, column=col_P, value=f"={col_letter_c}{last_data_row}")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # delta = end11 - start6
            p_letter = get_column_letter(col_P)
            cell = ws.cell(row=r11, column=col_P, value=f"={p_letter}{r10}-{p_letter}{r9}")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
        
        # Column Q (C stdev) -> references col_letter_c (L)
        col_Q = calc_start_col + 2 # Q
        if col_letter_c:
            # ref stdev
            cell = ws.cell(row=r3, column=col_Q,
                           value=f"=STDEV({col_letter_c}{ref_rows[0]},{col_letter_c}{ref_rows[1]},{col_letter_c}{ref_rows[2]})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # all stdev
            cell = ws.cell(row=r5, column=col_Q,
                           value=f"=STDEV({col_letter_c}{first_after_ref}:{col_letter_c}{last_data_row})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # last 6 stdev
            cell = ws.cell(row=r6, column=col_Q,
                           value=f"=STDEV({col_letter_c}{last6_start}:{col_letter_c}{last_data_row})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
        
        # Column R (Flag C)
        col_R = calc_start_col + 3 
        if is_flagged:
            cell_R = ws.cell(row=r6, column=col_R, value=flag_message)
            cell_R.fill = flag_fill_red
        else:
            # Retain fill for blank cell
            ws.cell(row=r6, column=col_R, value="").fill = calc_fill

        # Columns S (O avg), T (O stdev), U (Flag O) referencing column M (col_letter_o)
        col_S = calc_start_col + 4 # S
        col_T = calc_start_col + 5 # T
        col_U = calc_start_col + 6 # U
        
        if col_letter_o:
            # S: ref avg
            cell = ws.cell(row=r3, column=col_S,
                           value=f"=AVERAGE({col_letter_o}{ref_rows[0]},{col_letter_o}{ref_rows[1]},{col_letter_o}{ref_rows[2]})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # S: all avg
            cell = ws.cell(row=r5, column=col_S,
                           value=f"=AVERAGE({col_letter_o}{first_after_ref}:{col_letter_o}{last_data_row})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # S: last 6 avg
            cell = ws.cell(row=r6, column=col_S,
                           value=f"=AVERAGE({col_letter_o}{last6_start}:{col_letter_o}{last_data_row})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # S: start 6 for O
            cell = ws.cell(row=r9, column=col_S, value=f"={col_letter_o}{last6_start}")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # S: end 11 for O
            cell = ws.cell(row=r10, column=col_S, value=f"={col_letter_o}{last_data_row}")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # S: delta for O
            s_letter = get_column_letter(col_S)
            cell = ws.cell(row=r11, column=col_S, value=f"={s_letter}{r10}-{s_letter}{r9}")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            
            # T: ref stdev
            cell = ws.cell(row=r3, column=col_T,
                           value=f"=STDEV({col_letter_o}{ref_rows[0]},{col_letter_o}{ref_rows[1]},{col_letter_o}{ref_rows[2]})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # T: all stdev
            cell = ws.cell(row=r5, column=col_T,
                           value=f"=STDEV({col_letter_o}{first_after_ref}:{col_letter_o}{last_data_row})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill
            # T: last 6 stdev
            cell = ws.cell(row=r6, column=col_T,
                           value=f"=STDEV({col_letter_o}{last6_start}:{col_letter_o}{last_data_row})")
            cell.number_format = NUM_FORMAT_3_DECIMAL
            cell.fill = calc_fill # Retain fill

        # Column U (Flag O)
        if is_flagged:
            cell_U = ws.cell(row=r6, column=col_U, value=flag_message)
            cell_U.fill = flag_fill_red
        else:
            # Retain fill for blank cell
            ws.cell(row=r6, column=col_U, value="").fill = calc_fill

        # Column V (Sum area all) -> sums on area column (col_letter_area, K)
        col_V = calc_start_col + 7 # V (22)
        if col_letter_area:
            # V: ref sum
            cell = ws.cell(row=r3, column=col_V,
                           value=f"=SUM({col_letter_area}{first_data_row}:{col_letter_area}{first_data_row+3})")
            cell.number_format = NUM_FORMAT_2_DECIMAL # CORRECTED TO 2 DECIMALS
            cell.fill = calc_fill # Retain fill
            # V: all sum
            cell = ws.cell(row=r5, column=col_V,
                           value=f"=SUM({col_letter_area}{first_after_ref}:{col_letter_area}{last_data_row})")
            cell.number_format = NUM_FORMAT_2_DECIMAL # CORRECTED TO 2 DECIMALS
            cell.fill = calc_fill # Retain fill
            # V: last 6 sum
            cell = ws.cell(row=r6, column=col_V,
                           value=f"=SUM({col_letter_area}{last6_start}:{col_letter_area}{last_data_row})")
            cell.number_format = NUM_FORMAT_2_DECIMAL # CORRECTED TO 2 DECIMALS
            cell.fill = calc_fill # Retain fill
        
        # Column W (funny peaks) -> first 4 rows "ref", then compare Ampl 44 (col_letter_ampl -> J)
        col_W = calc_start_col + 8 # W (23)
        if col_letter_ampl:
            # first 4 rows are "ref"
            for rr in (r1, r2, r3, r4):
                ws.cell(row=rr, column=col_W, value="ref").fill = calc_fill # Retain fill
            # from first_after_ref through last_data_row:
            for rr in range(first_after_ref, last_data_row + 1):
                ws.cell(row=rr, column=col_W,
                        value=f"=IF({col_letter_ampl}{rr}>{col_letter_ampl}{rr+1},IF({col_letter_ampl}{rr+1}<{col_letter_ampl}{rr},\"ok\",\"check\"),\"check\")").fill = calc_fill # Retain fill
        
        # Column X (min intensity) -> first cell "if 44<1000", next 3 empty, then formula =IF(Jr<1000,"check","ok")
        col_X = calc_start_col + 9 # X (24)
        if col_letter_ampl:
            ws.cell(row=r1, column=col_X, value="if 44<1000").fill = calc_fill # Retain fill
            for rr in range(first_after_ref, last_data_row + 1):
                ws.cell(row=rr, column=col_X, value=f"=IF({col_letter_ampl}{rr}<1000,\"check\",\"ok\")").fill = calc_fill # Retain fill
        
        # -------------------- CONDITIONAL FORMATTING FOR “LAST 6” (Q and T) --------------------
        if stdev_threshold is not None:
            # Column Q (C stdev) - Conditional Formatting for 'last 6' row (r6)
            ws.conditional_formatting.add(
                f"{get_column_letter(col_Q)}{r6}",
                CellIsRule(
                    operator="greaterThan",
                    formula=[str(stdev_threshold)],
                    fill=fill_error
                )
            )

            # Column T (O stdev) - Conditional Formatting for 'last 6' row (r6)
            ws.conditional_formatting.add(
                f"{get_column_letter(col_T)}{r6}",
                CellIsRule(
                    operator="greaterThan",
                    formula=[str(stdev_threshold)],
                    fill=fill_error
                )
            )
        # -------------------- END CF --------------------

        # --- Transparent separator row (A → calc_end_col) ---
        write_transparent_separator(ws, cur_row, calc_end_col)
        cur_row += 1
        
    wb.save(file_path)
    print(f"✅ Step 1: DATA completed on {file_path}")