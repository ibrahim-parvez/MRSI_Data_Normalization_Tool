import os
from copy import copy, deepcopy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
# Import CellIsRule, FormulaRule to ensure all imports are available for downstream use
from openpyxl.formatting.rule import CellIsRule, FormulaRule 

def _is_formula_cell(cell):
    """Return True if the cell is a formula."""
    try:
        if getattr(cell, "data_type", None) == "f":
            return True
        val = cell.value
        return isinstance(val, str) and val.startswith("=")
    except Exception:
        return False

def _try_refresh_with_xlwings(path):
    try:
        import xlwings as xw
    except Exception:
        return False

    app = None
    book = None
    try:
        app = xw.App(visible=False, add_book=False)
        book = app.books.open(os.path.abspath(path))
        try:
            book.app.api.Application.CalculateFull()
        except Exception:
            try:
                book.app.api.Application.Calculate()
            except Exception:
                try:
                    book.app.calculate()
                except Exception:
                    pass
        book.save()
        book.close()
        app.quit()
        return True
    except Exception:
        try:
            if book is not None:
                book.close()
        except Exception:
            pass
        try:
            if app is not None:
                app.quit()
        except Exception:
            pass
        return False

def step7_report_water(file_path):
    source_sheet = "Normalization_DNT"
    new_sheet_name = "Report_DNT"

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb_fmt = load_workbook(file_path, data_only=False)
    wb_val = load_workbook(file_path, data_only=True)

    if source_sheet not in wb_fmt.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found.")

    ws_fmt = wb_fmt[source_sheet]
    ws_val = wb_val[source_sheet]

    # --- Helper: Check for Silver/Gray #C0C0C0 ---
    def _cell_rgb_upper(cell):
        try:
            fg = getattr(cell.fill, "fgColor", None)
            if fg is not None:
                rgb = getattr(fg, "rgb", None)
                if rgb:
                    return str(rgb).upper()
            sc = getattr(cell.fill, "start_color", None)
            if sc is not None:
                rgb2 = getattr(sc, "rgb", None)
                if rgb2:
                    return str(rgb2).upper()
        except Exception:
            pass
        return None

    def _is_grayC0C0C0(cell):
        rgb = _cell_rgb_upper(cell)
        # Check specifically for C0C0C0 (often comes as FFC0C0C0 or 00C0C0C0)
        return bool(rgb and rgb.endswith("C0C0C0"))

    # --- 1. Define Columns (A-C and S-Y) ---
    # A=1, B=2, C=3
    # S=19, T=20, U=21, V=22, W=23, X=24, Y=25
    source_cols = [1, 2, 3] + list(range(19, 26)) 

    # --- 2. Check for Formula Updates ---
    # We check a small sample to see if xlwings refresh is needed
    needs_refresh = False
    # Check specifically in the header region (row 16-18) or just below
    check_rows = list(range(16, 19))
    for r in check_rows:
        for c in source_cols:
            src = ws_fmt.cell(row=r, column=c)
            valcell = ws_val.cell(row=r, column=c)
            if _is_formula_cell(src) and (valcell.value is None):
                needs_refresh = True
                break
        if needs_refresh:
            break

    if needs_refresh:
        refreshed = _try_refresh_with_xlwings(file_path)
        if refreshed:
            wb_fmt = load_workbook(file_path, data_only=False)
            wb_val = load_workbook(file_path, data_only=True)
            ws_fmt = wb_fmt[source_sheet]
            ws_val = wb_val[source_sheet]

    # --- 3. Prepare New Sheet ---
    if new_sheet_name in wb_fmt.sheetnames:
        del wb_fmt[new_sheet_name]

    ws_new = wb_fmt.create_sheet(new_sheet_name, index=wb_fmt.index(ws_fmt))

    # Column Mapping: Maps Source Column Index -> New Sequential Column Index
    # e.g. Col 1 -> 1, Col 19 -> 4
    mapping = {src_col: idx for idx, src_col in enumerate(source_cols, start=1)}

    # --- 4. Processing Function (Copy Logic) ---
    def copy_row(src_row_idx, dest_row_idx):
        """Copies specific columns from src_row to dest_row."""
        for src_col in source_cols:
            new_col = mapping[src_col]
            src_cell_fmt = ws_fmt.cell(row=src_row_idx, column=src_col)
            src_cell_val = ws_val.cell(row=src_row_idx, column=src_col)
            dst = ws_new.cell(row=dest_row_idx, column=new_col)

            # Value transfer
            value = None
            if src_cell_val.value is not None:
                value = src_cell_val.value
            elif not _is_formula_cell(src_cell_fmt):
                value = src_cell_fmt.value

            # Rich Text
            try:
                if hasattr(src_cell_fmt, "rich_text") and src_cell_fmt.rich_text:
                    rt = CellRichText()
                    for block in src_cell_fmt.rich_text:
                        if isinstance(block, TextBlock):
                            rt.append(deepcopy(block))
                    dst.rich_text = rt
                else:
                    dst.value = value
            except Exception:
                dst.value = value

            # Comments
            try:
                if getattr(src_cell_fmt, "comment", None) is not None:
                    dst.comment = deepcopy(src_cell_fmt.comment)
            except Exception:
                pass

            # Style (Font, Border, Fill, NumberFormat, Alignment, Protection)
            try:
                if src_cell_fmt.has_style:
                    dst.font = copy(src_cell_fmt.font)
                    dst.border = copy(src_cell_fmt.border)
                    dst.fill = copy(src_cell_fmt.fill)
                    dst.number_format = src_cell_fmt.number_format
                    dst.protection = copy(src_cell_fmt.protection)
                    dst.alignment = copy(src_cell_fmt.alignment)
            except Exception:
                pass

        # Row Height
        try:
            rd = ws_fmt.row_dimensions.get(src_row_idx)
            if rd is not None and getattr(rd, "height", None) is not None:
                ws_new.row_dimensions[dest_row_idx].height = rd.height
        except Exception:
            pass

    # --- 5. EXECUTION ---
    
    current_write_row = 1

    # STEP A: Copy Headers (Rows 16, 17, 18)
    for r in range(16, 19):
        copy_row(r, current_write_row)
        current_write_row += 1

    # STEP B: Find the 2-Row Gray Separator (#C0C0C0)
    # We search starting from row 19 downwards to find the silver separator
    separator_start_row = None
    search_limit = min(ws_fmt.max_row, 100) # Search reasonable range

    for r in range(19, search_limit):
        # Check if a significant number of columns in our range are C0C0C0
        # checking cols 1-3 is usually sufficient for a full band
        is_gray_r = _is_grayC0C0C0(ws_fmt.cell(row=r, column=1))
        is_gray_r1 = _is_grayC0C0C0(ws_fmt.cell(row=r+1, column=1))
        
        if is_gray_r and is_gray_r1:
            separator_start_row = r
            break
    
    if separator_start_row is None:
        print("Warning: Could not find the #C0C0C0 separator. Report might be incomplete.")
        # Fallback: If no separator found, maybe just continue from 19?
        data_start_row = 19 
    else:
        # User requested: "start from under the 2 rows"
        # Separator is at r and r+1. Data starts at r+2.
        data_start_row = separator_start_row + 2

    # STEP C: Copy Data Body
    # Copy from data_start_row to end of sheet
    for r in range(data_start_row, ws_fmt.max_row + 1):
        copy_row(r, current_write_row)
        current_write_row += 1

    # --- 6. Final Polish (Column Widths & View) ---
    for src_col, new_col in mapping.items():
        try:
            src_letter = get_column_letter(src_col)
            new_letter = get_column_letter(new_col)
            cd = ws_fmt.column_dimensions.get(src_letter)
            if cd is not None and getattr(cd, "width", None) is not None:
                ws_new.column_dimensions[new_letter].width = cd.width
        except Exception:
            pass

    # Set View to A1
    for s in wb_fmt.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    ws_new.sheet_view.tabSelected = True
    wb_fmt.active = wb_fmt.index(ws_new)
    ws_new.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    wb_fmt.save(file_path)
    print(f"Step 7: Water Report completed on {file_path}")