import string
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from copy import copy
import settings

def step3_last6_water(file_path: str):
    """
    Step 3: Last 6 (Water)
    - Specifically creates a sheet called 'Last 6_DNT'.
    - Copies from 'To Sort_DNT'.
    - Fixed filter: Only shows rows where Column O is 'last 6'.
    - Removes filter arrows for a clean, static report.
    - Applies Conditional Formatting for Stdev thresholds.
    """
    
    source_sheet_name = "To Sort_DNT"
    new_sheet_name = "Last 6_DNT"
    fixed_filter = "last 6"

    wb = load_workbook(file_path)

    if source_sheet_name not in wb.sheetnames:
        raise ValueError(f"Source sheet '{source_sheet_name}' not found. Please run Step 2 first.")
    
    # --- 1. Cleanup existing sheet ---
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    source_ws = wb[source_sheet_name]
    
    # Position it to the left of 'To Sort_DNT'
    idx = wb.index(source_ws)
    new_ws = wb.create_sheet(new_sheet_name, idx)
    
    # --- 2. Cell-by-Cell Copy ---
    for row in source_ws.iter_rows():
        for cell in row:
            cell_value = cell.value
            if cell.data_type == 'f' and hasattr(cell, 'formula'):
                cell_value = cell.formula
            
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell_value)
            
            if cell.has_style:
                try:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                except:
                    pass

    # Copy dimensions
    #for row_idx, row_dim in source_ws.row_dimensions.items():
    #    new_ws.row_dimensions[row_idx].height = row_dim.height
    #for col_letter, col_dim in source_ws.column_dimensions.items():
    #    new_ws.column_dimensions[col_letter].width = col_dim.width
    
    # --- 3. Fixed Filter Logic (Hide rows not matching 'last 6') ---
    last_row = new_ws.max_row
    col_o_idx = 15 # Column O
    
    for r in range(2, last_row + 1):
        cell_val = new_ws.cell(row=r, column=col_o_idx).value
        # Standardize string for comparison
        val_str = str(cell_val).strip().lower() if cell_val else ""
        
        if val_str != fixed_filter:
            new_ws.row_dimensions[r].hidden = True
            
    # Ensure AutoFilter arrows are removed
    new_ws.auto_filter.ref = None 

    # --- 4. Conditional Formatting (Red Stdev Highlights) ---
    stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Column Q (17) and T (20)
    COL_Q = 17
    COL_T = 20
    START_ROW_FOR_CF = 8 
    ROW_STEP = 12

    if stdev_threshold is not None:
        for r_idx in range(START_ROW_FOR_CF, last_row + 1, ROW_STEP):
            # Carbon Stdev (Q)
            q_ref = f"{get_column_letter(COL_Q)}{r_idx}"
            new_ws.conditional_formatting.add(
                q_ref,
                CellIsRule(operator="greaterThan", formula=[str(stdev_threshold)], fill=fill_error)
            )

            # Oxygen Stdev (T)
            t_ref = f"{get_column_letter(COL_T)}{r_idx}"
            new_ws.conditional_formatting.add(
                t_ref,
                CellIsRule(operator="greaterThan", formula=[str(stdev_threshold)], fill=fill_error)
            )

    # --- 5. Finalize View ---
    for s in wb.worksheets:
        s.sheet_view.tabSelected = False
    new_ws.sheet_view.tabSelected = True
    wb.active = wb.index(new_ws)
    new_ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    
    wb.save(file_path)
    print(f"✅ Step 3: '{new_sheet_name}' created.")