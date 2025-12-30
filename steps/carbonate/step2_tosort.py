import os
import time
import traceback
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter

def _try_force_excel_recalc(file_path, timeout=5.0):
    """
    Try to open the workbook in Excel via xlwings, calculate, save and close.
    Returns True on success, False on failure (e.g. xlwings not installed or Excel not available).
    """
    try:
        import xlwings as xw
    except Exception:
        return False

    app = None
    try:
        app = xw.App(visible=False)
        # Give Excel a moment to start
        time.sleep(0.2)
        book = app.books.open(os.path.abspath(file_path))
        # Force a full recalculation
        book.app.calculate()
        # Wait a little for Excel to finish (best-effort)
        time.sleep(min(1.0, timeout))
        book.save()
        book.close()
        app.quit()
        return True
    except Exception:
        try:
            if app:
                app.quit()
        except Exception:
            pass
        return False


def step2_tosort_carbonate(file_path, filter_choice="Last 6"):
    """
    Step 2: TO SORT (refactored + style copying)

    - Copies rows from 'Data' into 'To Sort' but converts formulas into raw values in To Sort.
    - Values are taken from a data_only workbook so formulas do not get copied.
    - Formatting (fonts, fills, borders, alignment, number formats, merged cells,
      row heights, column widths) is copied from the original 'Data' worksheet.
    - Attempts to force Excel recalc (via xlwings) so cached values exist; if recalculation fails,
      the code will still copy whatever cached values exist (may be None for some formula cells).
    - Finally: applies autofilter on column Q and hides rows not matching filter (unless "All").

    DISPLAY RULES APPLIED:
    - Columns that display averages/stdevs (R,S,U,V) will be formatted with "0.000".
    - Sum area column X (24) will be formatted with "0.00".
    """

    source_sheet = "Data_DNT"
    new_sheet_name = "To Sort_DNT"

    # Try to force recalculation in Excel (best-effort)
    recalc_ok = _try_force_excel_recalc(file_path)
    if not recalc_ok:
        print("Warning: unable to force Excel recalculation (xlwings missing or failed).")
        print("If 'Data' contains formulas without cached values, 'To Sort' may have empty cells for those formulas.")

    # Load two workbook views:
    # - wb (data_only=False) is used to write the result (we will write To Sort into it)
    # - wb_values (data_only=True) is used to read calculated values (no formulas)
    wb = load_workbook(file_path, data_only=False)
    wb_values = load_workbook(file_path, data_only=True)

    if source_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found in workbook. Run Step 1 first.")
    if source_sheet not in wb_values.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found in values workbook. Run Step 1 first.")

    # Remove old To Sort if present (from the formula workbook)
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    ws_source = wb[source_sheet]                 # formatting + formulas
    ws_source_values = wb_values[source_sheet]   # calculated values only

    # Create To Sort sheet to the LEFT of Data sheet
    ws_new = wb.create_sheet(new_sheet_name, index=wb.index(ws_source))

    # Columns we'd like to coerce to text so Excel marks them as text (green triangle)
    # (D,E,F) --> columns 4,5,6 (1-based)
    text_cols = {4, 5, 6}

    # Determine used range
    max_col_idx = ws_source.max_column or 0
    # Prefer max_row from values workbook (it represents cached values), but fallback to source
    max_row_from_values = ws_source_values.max_row if (hasattr(ws_source_values, "max_row") and ws_source_values.max_row) else 0
    max_row_idx = max(max_row_from_values, ws_source.max_row or 0)

    # Defensive: if worksheet is empty, ensure at least header is handled gracefully
    if max_row_idx == 0 or max_col_idx == 0:
        # Still set active sheet and save (creates an empty sheet)
        ws_new.sheet_view.tabSelected = True
        wb.active = wb.index(ws_new)
        wb.save(file_path)
        print(f"Step 2: TO SORT created empty sheet '{new_sheet_name}' in {file_path}")
        return

    # Copy row heights
    for r in range(1, max_row_idx + 1):
        rd = ws_source.row_dimensions.get(r)
        if rd is not None and rd.height is not None:
            ws_new.row_dimensions[r].height = rd.height

    # Copy column widths
    for c in range(1, max_col_idx + 1):
        col_letter = get_column_letter(c)
        cd = ws_source.column_dimensions.get(col_letter)
        if cd is not None and cd.width is not None:
            ws_new.column_dimensions[col_letter].width = cd.width

    # Copy merged cells (only those ranges that lie within our used range)
    try:
        for merged in list(ws_source.merged_cells.ranges):
            # merged is a MergedCellRange; string() gives like "A1:C1"
            try:
                ws_new.merge_cells(str(merged))
            except Exception:
                # Ignore any merge that cannot be applied (out-of-range etc.)
                continue
    except Exception:
        # If merged_cells structure is unexpected, just skip
        pass

    # --- COPY CONDITIONAL FORMATTING RULES ---
    try:
        for cf_range in ws_source.conditional_formatting._cf_rules:
            rules = ws_source.conditional_formatting._cf_rules[cf_range]

            for rule in rules:
                # Clone the rule
                new_rule = copy(rule)

                # Add the rule to the NEW sheet
                ws_new.conditional_formatting.add(cf_range, new_rule)

    except Exception as e:
        print("Warning: unable to copy conditional formatting rules:", e)


    # --- CORE: copy value from ws_source_values and style from ws_source ---
    # We'll iterate full grid 1..max_row_idx x 1..max_col_idx to ensure consistent layout,
    # explicitly setting None where no value exists.

    # Performance note: this is cell-by-cell and can be slower on very large sheets.
    for r in range(1, max_row_idx + 1):
        for c in range(1, max_col_idx + 1):
            new_cell = ws_new.cell(row=r, column=c)

            # Value from values-only workbook
            try:
                val = ws_source_values.cell(row=r, column=c).value
            except Exception:
                val = None

            # Convert certain columns to text explicitly to trigger Excel's green-triangle (if non-empty)
            if c in text_cols and val is not None:
                try:
                    val = str(val)
                except Exception:
                    # fallback: keep original
                    pass
            new_cell.value = val

            # Copy formatting from original worksheet cell if present
            try:
                src_cell = ws_source.cell(row=r, column=c)
                # Only copy if the source cell has style attributes set
                # Use copy() to duplicate style objects where appropriate
                if hasattr(src_cell, "has_style") and src_cell.has_style:
                    # Font, fill, border, alignment, protection are objects — use copy()
                    if src_cell.font is not None:
                        new_cell.font = copy(src_cell.font)
                    if src_cell.fill is not None:
                        new_cell.fill = copy(src_cell.fill)
                    if src_cell.border is not None:
                        new_cell.border = copy(src_cell.border)
                    if src_cell.alignment is not None:
                        new_cell.alignment = copy(src_cell.alignment)
                    if src_cell.protection is not None:
                        new_cell.protection = copy(src_cell.protection)
                    # Number format (string)
                    try:
                        new_cell.number_format = src_cell.number_format
                    except Exception:
                        # ignore issues copying number format
                        pass
            except Exception:
                # be defensive: failure to copy formatting for a cell should not abort the entire process
                continue

    # If, for any reason, the iteration above didn't reach the last row (very unlikely),
    # ensure last_row is computed for subsequent operations.
    last_row = max_row_idx

    # --- APPLY NUMBER FORMATTING OVERRIDES (display-only rules) ---
    # These are display formats you asked for; they override the copied number formats.
    fmt_three = "0.000"
    fmt_two = "0.00"

    # Only attempt format application if there are rows (skip header row if only one row)
    if last_row >= 2:
        for r in range(2, last_row + 1):
            try:
                # R (18), S (19)
                if 18 <= max_col_idx:
                    ws_new.cell(row=r, column=18).number_format = fmt_three
                if 19 <= max_col_idx:
                    ws_new.cell(row=r, column=19).number_format = fmt_three

                # U (21), V (22)
                if 21 <= max_col_idx:
                    ws_new.cell(row=r, column=21).number_format = fmt_three
                if 22 <= max_col_idx:
                    ws_new.cell(row=r, column=22).number_format = fmt_three

                # X (24): sum area (two decimals)
                if 24 <= max_col_idx:
                    ws_new.cell(row=r, column=24).number_format = fmt_two
            except Exception:
                # If a particular cell doesn't exist or cannot be formatted, skip it.
                continue

    # Apply autofilter across full used range (based on source's max row/col)
    last_col_letter = get_column_letter(max_col_idx)
    try:
        ws_new.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    except Exception:
        # If autofilter cannot be set (odd workbook), ignore
        pass

    # Prepare filter choice normalization
    filter_choice_norm = (filter_choice or "Last 6").strip().lower()

    # openpyxl expects zero-based index for add_filter_column (leftmost column = 0)
    # Column Q is 17th column 1-based -> zero-based index = 16
    target_filter_index = 16
    try:
        ws_new.auto_filter.add_filter_column(target_filter_index, [filter_choice_norm])
        ws_new.auto_filter.add_sort_condition(f"Q2:Q{last_row}")
    except Exception:
        # If add_filter_column or sort condition fails, ignore — we will hide rows below as fallback
        pass

    # Hide rows not matching filter (unless "all")
    if filter_choice_norm != "all":
        for r in range(2, last_row + 1):
            try:
                val = ws_new.cell(row=r, column=17).value  # column Q is 17 (1-based)
                if val is None:
                    ws_new.row_dimensions[r].hidden = True
                    continue
                try:
                    if str(val).strip().lower() != filter_choice_norm:
                        ws_new.row_dimensions[r].hidden = True
                    else:
                        ws_new.row_dimensions[r].hidden = False
                except Exception:
                    ws_new.row_dimensions[r].hidden = True
            except Exception:
                # If reading cell fails for some row, hide it (safe default)
                try:
                    ws_new.row_dimensions[r].hidden = True
                except Exception:
                    pass

    # Activate new sheet and set selection
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    try:
        ws_new.sheet_view.tabSelected = True
        wb.active = wb.index(ws_new)
        ws_new.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    except Exception:
        # ignore selection errors
        pass

    # Save the workbook (this writes To Sort into the same workbook that still has Data formulas)
    try:
        wb.save(file_path)
        print(f"Step 2: To Sort completed on {file_path}")
        if not recalc_ok:
            print("Note: xlwings recalculation was not run. If To Sort contains blanks in R–AA,")
            print("open the workbook in Excel and save once (or enable auto-calc), then re-run Step 2.")
    except Exception as e:
        # Provide some helpful debugging info if save fails
        print("Error: failed to save workbook after creating 'To Sort' sheet.")
        traceback.print_exc()
        raise e
