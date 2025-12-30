import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import re
import unicodedata
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy 
import settings

# 💡 UPDATED IMPORT for FormulaRule
from openpyxl.formatting.rule import CellIsRule, FormulaRule 

# --- Helper Functions (omitted for brevity, assume unchanged) ---
def _normalize_text(text):
# ... (rest of helper function code)
    if not text:
        return ""
    text = str(text)
    text = unicodedata.normalize("NFKD", text)
    text = re.sub(r"[^A-Za-z0-9]+", "", text)
    return text.lower().strip()

def create_rich_text(parts):
    rt = CellRichText()
    for font, text in parts:
        rt.append(TextBlock(font, text))
    return rt

def extract_sample_base(identifier):
    if not identifier or not isinstance(identifier, str):
        return ""
    identifier = identifier.strip()
    base = re.sub(r"\s*r\d+(\.\d+)?$", "", identifier, flags=re.IGNORECASE)
    return base.strip()

def extract_run_number(identifier):
    if not identifier or not isinstance(identifier, str):
        return (9999, 0)
    m = re.search(r"r(\d+)(?:\.(\d+))?", identifier, flags=re.IGNORECASE)
    if m:
        major = int(m.group(1))
        minor = int(m.group(2)) if m.group(2) else 0
        return (major, minor)
    return (9999, 0)

def _make_fill(hex_color):
    c = hex_color.replace("#", "").upper()
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def _get_valid_co2_rows(rows, col_identifier1):
    valid_indices = []
    seen = {}
    for i, r in enumerate(rows):
        ident_raw = str(r[col_identifier1 - 1] or "").strip()
        ident_lower = ident_raw.lower()
        # Normalize explicit HeCO2 prefix -> co2 (preserve rest after prefix)
        if ident_lower.startswith("heco2"):
            ident_clean = "co2" + ident_raw[len("heco2"):]
        # If it starts with co2 (any casing) normalize to 'co2' prefix too
        elif ident_lower.startswith("co2"):
            ident_clean = "co2" + ident_raw[len("co2"):]
        else:
            ident_clean = ident_raw
        major, minor = extract_run_number(ident_clean)
        # if extraction failed, skip
        if major is None or minor is None:
            continue
        # skip R1.*
        if major == 1:
            continue
        # keep the lowest minor for each major
        if major not in seen:
            seen[major] = (minor, i)
        else:
            prev_minor, prev_i = seen[major]
            if minor < prev_minor:
                seen[major] = (minor, i)
    return sorted(idx for _, idx in seen.values())

def get_summary_num_format(base_name):
    return '0.000'

def add_blue_box(ws):
    # Styles
    thick = Side(border_style="thick", color="000000")
    medium = Side(border_style="medium", color="000000")
    blue_fill = PatternFill(start_color="DAE9F8", end_color="DAE9F8", fill_type="solid")
    black_bold = Font(color="000000", bold=True)
    green_bold = Font(color="008000", bold=True)
    red_bold = Font(color="FF0000", bold=True)
    darkblue_bold = Font(color="000080", bold=True)
    lightblue_bold = Font(color="3399FF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    # A1: put date text "2025-0521"
    today_str = datetime.today().strftime("%Y-%m%d")
    ws.cell(row=1, column=1, value=today_str).alignment = Alignment(horizontal="left", vertical="center")
    
    # C1 heading: bold "Normalization"
    ws.cell(row=1, column=3, value="Normalization").font = black_bold
    ws.cell(row=1, column=3).alignment = center
    
    
    # --- Box around C2:C3 (medium border on outer edges) ---
    c_col = 3
    for r in range(2, 4): # rows 2-3 inclusive
        cell = ws.cell(row=r, column=c_col)
        cell.fill = blue_fill
        cell.alignment = center
        # apply outer border only
        for r in range(2, 4):
            top = medium if r == 2 else None
            bottom = medium if r == 3 else None
            left = medium
            right = medium
            ws.cell(row=r, column=c_col).border = Border(top=top, bottom=bottom, left=left, right=right)
    # C2: Reference Materials in Bold
    ws.cell(row=2, column=c_col, value="Reference Materials").font = black_bold
    ws.cell(row=2, column=c_col).alignment = center
    
    # --- Box C4:C8 (outer border) with entries ---
    for r in range(4, 9):
        ws.cell(row=r, column=c_col).fill = blue_fill
        ws.cell(row=r, column=c_col).alignment = center
    
    # outer border for C4:C8
    for r in range(4, 9):
        top = medium if r == 4 else None
        bottom = medium if r == 8 else None
        left = medium
        right = medium
        ws.cell(row=r, column=c_col).border = Border(top=top, bottom=bottom, left=left, right=right)
    
    # Fill C5-C8 values (centered, colored) — note: you specified non-bold here
    ws.cell(row=5, column=c_col, value="IAEA 603").font = Font(color="008000", bold=False)
    ws.cell(row=5, column=c_col).alignment = center
    ws.cell(row=6, column=c_col, value="LSVEC").font = Font(color="3399FF", bold=False)
    ws.cell(row=6, column=c_col).alignment = center
    ws.cell(row=7, column=c_col, value="NBS 18").font = Font(color="FF0000", bold=False)
    ws.cell(row=7, column=c_col).alignment = center
    ws.cell(row=8, column=c_col, value="NBS 19").font = Font(color="000080", bold=False)
    ws.cell(row=8, column=c_col).alignment = center
    
    # --- Box D2:H3 (D=4 .. H=8) outer border; also merge F2:G2 ---
    col_left = 4
    col_right = 8
    row_top = 2
    row_bot = 3
    
    # fill & center the area
    for r in range(row_top, row_bot + 1):
        for c in range(col_left, col_right + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_fill
            cell.alignment = center
            
    # draw outer border for D2:H3
    for r in range(row_top, row_bot + 1):
        for c in range(col_left, col_right + 1):
            top = thick if r == row_top else None
            bottom = thick if r == row_bot else None
            left = thick if c == col_left else None
            right = thick if c == col_right else None
            ws.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)
            
    # Merge F2:G2 (F=6, G=7) and write "Published (vs. VPDB)" centered bold
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    ws.cell(row=2, column=6, value="Published (vs. VPDB)").font = black_bold
    ws.cell(row=2, column=6).alignment = center
    # F3 and G3: isotope labels
    ws.cell(row=3, column=6, value="δ¹³C").alignment = center
    ws.cell(row=3, column=7, value="δ¹⁸O").alignment = center
    
    # --- Box D4:H8 (D=4..H=8 rows 4..8) outer border ---
    rtop = 4
    rbot = 8
    cleft = 4
    cright = 8
    for r in range(rtop, rbot + 1):
        for c in range(cleft, cright + 1):
            ws.cell(row=r, column=c).fill = blue_fill
            ws.cell(row=r, column=c).alignment = center
            
    # outer border
    for r in range(rtop, rbot + 1):
        for c in range(cleft, cright + 1):
            top = thick if r == rtop else None
            bottom = thick if r == rbot else None
            left = thick if c == cleft else None
            right = thick if c == cright else None
            ws.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)
            
    # Insert the requested numeric values (with colors & bold)
    # E6: -46.6 in light blue bold (E=5, row=6)
    ws.cell(row=6, column=5, value=-46.6).font = lightblue_bold
    ws.cell(row=6, column=5).alignment = center
    # F5: 2.46 in green bold. (F=6, row=5)
    ws.cell(row=5, column=6, value=2.46).font = green_bold
    ws.cell(row=5, column=6).alignment = center
    # G5: -2.37 green bold (G=7,row=5)
    ws.cell(row=5, column=7, value=-2.37).font = green_bold
    ws.cell(row=5, column=7).alignment = center
    # H6: -26.7 light blue bold (H=8,row=6)
    ws.cell(row=6, column=8, value=-26.7).font = lightblue_bold
    ws.cell(row=6, column=8).alignment = center
    # F7: red -5.01 (F=6,row=7)
    ws.cell(row=7, column=6, value=-5.01).font = red_bold
    ws.cell(row=7, column=6).alignment = center
    # G7: red -23.01
    ws.cell(row=7, column=7, value=-23.01).font = red_bold
    ws.cell(row=7, column=7).alignment = center
    # F8: dark blue bold 1.95
    ws.cell(row=8, column=6, value=1.95).font = darkblue_bold
    ws.cell(row=8, column=6).alignment = center
    # G8: dark blue bold -2.2
    ws.cell(row=8, column=7, value=-2.2).font = darkblue_bold
    ws.cell(row=8, column=7).alignment = center
    
    # --- NEW REQUESTED CELLS & FORMATTING OUTSIDE/ADJACENT TO THE ABOVE ---
    # InlineFont color hex format: '00RRGGBB' or similar used earlier
    red_if = InlineFont(color='00FF0000', b=True)
    blue_if = InlineFont(color='000000FF', b=True)
    darkblue_if = InlineFont(color='000080', b=True)
    green_if = InlineFont(color='008000', b=True)
    
    # 7) J10 & J13 should say "Slope" black bold; J11 & J14 "Intercept"
    # J = 10
    ws.cell(row=10, column=10, value="Slope").font = black_bold
    ws.cell(row=11, column=10, value="Intercept").font = black_bold
    ws.cell(row=13, column=10, value="Slope").font = black_bold
    ws.cell(row=14, column=10, value="Intercept").font = black_bold

    # 1. Add Rich Text Labels (I10 and I13)
    # I = 9
    # I10: "18 19" (NBS 18 and 19)
    ws.cell(row=10, column=9).value = create_rich_text([(red_if, "18 "), (blue_if, "19")])
    ws.cell(row=10, column=9).alignment = center
    # I13: "18 19 603" (NBS 18, 19, and IAEA 603)
    ws.cell(row=13, column=9).value = create_rich_text([(red_if, "18 "), (blue_if, "19 "), (green_if, "603")])
    ws.cell(row=13, column=9).alignment = center
    
    # 8) Add another box from J2 to N3, and a box from J4 to N8.
    # J=10, K=11, L=12, M=13, N=14
    
    # Fill & border for J2:N3
    for r in range(2, 4):
        for c in range(10, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_fill
            cell.alignment = center
    for r in range(2, 4):
        for c in range(10, 15):
            top = thick if r == 2 else None
            bottom = thick if r == 3 else None
            left = thick if c == 10 else None
            right = thick if c == 14 else None
            ws.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)
            
    # Merge J2:N2 and set text "Measured (vs. Working Standard)" centered black bold
    ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=14)
    ws.cell(row=2, column=10, value="Measured (vs. Working Standard)").font = black_bold
    ws.cell(row=2, column=10).alignment = center
    
    # K3 should have δ¹³C (K=11)
    ws.cell(row=3, column=11, value="δ¹³C").alignment = center
    # N3 should have δ¹⁸O (N=14)
    ws.cell(row=3, column=14, value="δ¹⁸O").alignment = center
    
    # Fill & border for J4:N8
    for r in range(4, 9):
        for c in range(10, 15):
            ws.cell(row=r, column=c).fill = blue_fill
            ws.cell(row=r, column=c).alignment = center
            
    # outer border
    for r in range(4, 9):
        for c in range(10, 15):
            top = thick if r == 4 else None
            bottom = thick if r == 8 else None
            left = thick if c == 10 else None
            right = thick if c == 14 else None
            ws.cell(row=r, column=c).border = Border(top=top, bottom=bottom, left=left, right=right)
            
    # --- Fill K5..K8 (C-avg) and N5..N8 (O-avg) from the precomputed "Average" rows in the sheet ---
    # mapping: reference name (lowercase) -> target row in the J/N box
    ref_to_target_row = {
        "iaea 603": 5,
        "lsvec": 6,
        "nbs 18": 7,
        "nbs 19": 8,
    }
    # initialize found flags to avoid overwriting if multiple average blocks exist
    found_ref = {k: False for k in ref_to_target_row}
    
    # Scan for the "Average" label that your write_group placed in column K (col 11).
    # The corresponding average values are written one row below that label (avg_row = label_row + 1).
    identifier_col = 3 # column C
    c_avg_col_ref = 11 # column K (C avg stored here)
    o_avg_col_ref = 14 # column N (O avg stored here)
    max_row = ws.max_row
    
    for r in range(1, max_row + 1):
        # We need to check columns K and N as they now hold the average labels
        val = ws.cell(row=r, column=c_avg_col_ref).value
        if val and str(val).strip().lower() == "average":
            avg_row = r + 1
            # find the identifier that belongs to this group: look at the last non-empty identifier above the label
            id_row = r - 1
            ident = ""
            # walk upwards until we find a non-empty identifier (limit search to 20 rows to be safe)
            scan_top = max(1, id_row - 20)
            for t in range(id_row, scan_top - 1, -1):
                cellv = ws.cell(row=t, column=identifier_col).value
                if cellv and str(cellv).strip():
                    ident = str(cellv).strip()
                    break
            
            ident_l = ident.lower()
            ref_key = None
            if "iaea" in ident_l or "603" in ident_l:
                ref_key = "iaea 603"
            elif "lsvec" in ident_l:
                ref_key = "lsvec"
            elif "nbs" in ident_l and "18" in ident_l:
                ref_key = "nbs 18"
            elif "nbs" in ident_l and "19" in ident_l:
                ref_key = "nbs 19"

            # if we detected a reference and haven't filled it yet, write formulas into K/N target cells
            if ref_key and ref_key in ref_to_target_row and not found_ref[ref_key]:
                target_row = ref_to_target_row[ref_key]
                num_format = get_summary_num_format(ref_key)
                
                # Write formulas that reference the avg cells (K{avg_row} and N{avg_row}) and round them
                ws.cell(row=target_row, column=11, value=f'=IFERROR({get_column_letter(c_avg_col_ref)}{avg_row},"")') # K = C-avg from K{avg_row}
                ws.cell(row=target_row, column=14, value=f'=IFERROR({get_column_letter(o_avg_col_ref)}{avg_row},"")') # N = O-avg from N{avg_row}
                ws.cell(row=target_row, column=11).number_format = num_format
                ws.cell(row=target_row, column=14).number_format = num_format
                
                # apply colors consistent with your sheet styling (not bold)
                if ref_key == "nbs 18":
                    ws.cell(row=target_row, column=11).font = Font(color="FF0000", bold=False)
                    ws.cell(row=target_row, column=14).font = Font(color="FF0000", bold=False)
                elif ref_key == "nbs 19":
                    ws.cell(row=target_row, column=11).font = Font(color="000080", bold=False)
                    ws.cell(row=target_row, column=14).font = Font(color="000080", bold=False)
                elif ref_key == "iaea 603":
                    ws.cell(row=target_row, column=11).font = Font(color="008000", bold=False)
                    ws.cell(row=target_row, column=14).font = Font(color="008000", bold=False)
                elif ref_key == "lsvec":
                    ws.cell(row=target_row, column=11).font = Font(color="3399FF", bold=False)
                    ws.cell(row=target_row, column=14).font = Font(color="3399FF", bold=False)
                found_ref[ref_key] = True

    # --- Determine numeric rows for columns K and N (5–8) ---
    def get_numeric_rows(ws, col, start=5, end=8):
        rows = []
        for r in range(start, end + 1):
            cell = ws.cell(row=r, column=col)
            val = cell.value
            # Handle numbers, numeric strings, or formulas with cached numeric results
            if isinstance(val, (int, float)):
                rows.append(r)
            elif isinstance(val, str):
                val_str = val.strip()
                # Include formula cells if not empty (i.e., starts with '=' but not just '=IF(...,"")')
                if val_str.startswith('=') and not val_str.upper().endswith('""'):
                    rows.append(r)
                else:
                    try:
                        float(val_str)
                        rows.append(r)
                    except ValueError:
                        continue
        return rows
        
    # K_rows will contain [5, 6, 7, 8] if all reference data is present
    k_rows = get_numeric_rows(ws, 11) # Column K (measured C)
    n_rows = get_numeric_rows(ws, 14) # Column N (measured O)

    # --- Helper to build Excel range strings ---
    def make_range(letter, rows):
        if not rows:
            return None
        # Use single row range if only one row, or the whole range
        if len(rows) == 1:
            return f"${letter}${rows[0]}:${letter}${rows[0]}"
        return f"${letter}${rows[0]}:${letter}${rows[-1]}"
        
    # --- Slope/Intercept 1: NBS 18/19 ONLY (Rows 7, 8) ---
    # Published columns: F (delta 13C), G (delta 18O)
    
    # K_18_19_rows: Filter k_rows for NBS 18 (row 7) and NBS 19 (row 8)
    k_18_19_rows = [r for r in k_rows if r in (7, 8)]
    n_18_19_rows = [r for r in n_rows if r in (7, 8)]

    # Carbon: Published F vs Measured K (Rows 7, 8) -> K10, K11
    if len(k_18_19_rows) >= 2:
        f_range_18_19 = make_range("F", k_18_19_rows)
        k_range_18_19 = make_range("K", k_18_19_rows)
        if f_range_18_19 and k_range_18_19:
            ws.cell(row=10, column=11).value = f"=IFERROR(SLOPE({f_range_18_19},{k_range_18_19}),\"\")"
            ws.cell(row=11, column=11).value = f"=IFERROR(INTERCEPT({f_range_18_19},{k_range_18_19}),\"\")"
        else:
            ws.cell(row=10, column=11).value = ""
            ws.cell(row=11, column=11).value = ""
    else:
        ws.cell(row=10, column=11).value = ""
        ws.cell(row=11, column=11).value = ""
        
    # Oxygen: Published G vs Measured N (Rows 7, 8) -> N10, N11
    if len(n_18_19_rows) >= 2:
        g_range_18_19 = make_range("G", n_18_19_rows)
        n_range_18_19 = make_range("N", n_18_19_rows)
        if g_range_18_19 and n_range_18_19:
            ws.cell(row=10, column=14).value = f"=IFERROR(SLOPE({g_range_18_19},{n_range_18_19}),\"\")"
            ws.cell(row=11, column=14).value = f"=IFERROR(INTERCEPT({g_range_18_19},{n_range_18_19}),\"\")"
        else:
            ws.cell(row=10, column=14).value = ""
            ws.cell(row=11, column=14).value = ""
    else:
        ws.cell(row=10, column=14).value = ""
        ws.cell(row=11, column=14).value = ""
        
    # --- Slope/Intercept 2: NBS 18/19 & IAEA 603 (Rows 5, 7, 8) ---
    
    # K_18_19_603_rows: Filter k_rows for NBS 18 (row 7), NBS 19 (row 8), and IAEA 603 (row 5)
    k_18_19_603_rows = [r for r in k_rows if r in (5, 7, 8)]
    n_18_19_603_rows = [r for r in n_rows if r in (5, 7, 8)]
    
    # Carbon: Published F vs Measured K (Rows 5, 7, 8) -> K13, K14
    if len(k_18_19_603_rows) >= 2 and 5 in k_18_19_603_rows: # Must have at least 2 points AND IAEA 603 (row 5) data
        f_range_18_19_603 = make_range("F", k_18_19_603_rows)
        k_range_18_19_603 = make_range("K", k_18_19_603_rows)
        if f_range_18_19_603 and k_range_18_19_603:
            ws.cell(row=13, column=11).value = f"=IFERROR(SLOPE({f_range_18_19_603},{k_range_18_19_603}),\"\")"
            ws.cell(row=14, column=11).value = f"=IFERROR(INTERCEPT({f_range_18_19_603},{k_range_18_19_603}),\"\")"
        else:
            ws.cell(row=13, column=11).value = ""
            ws.cell(row=14, column=11).value = ""
    else:
        ws.cell(row=13, column=11).value = ""
        ws.cell(row=14, column=11).value = ""
        
    # Oxygen: Published G vs Measured N (Rows 5, 7, 8) -> N13, N14
    if len(n_18_19_603_rows) >= 2 and 5 in n_18_19_603_rows: # Must have at least 2 points AND IAEA 603 (row 5) data
        g_range_18_19_603 = make_range("G", n_18_19_603_rows)
        n_range_18_19_603 = make_range("N", n_18_19_603_rows)
        if g_range_18_19_603 and n_range_18_19_603:
            ws.cell(row=13, column=14).value = f"=IFERROR(SLOPE({g_range_18_19_603},{n_range_18_19_603}),\"\")"
            ws.cell(row=14, column=14).value = f"=IFERROR(INTERCEPT({g_range_18_19_603},{n_range_18_19_603}),\"\")"
        else:
            ws.cell(row=13, column=14).value = ""
            ws.cell(row=14, column=14).value = ""
    else:
        ws.cell(row=13, column=14).value = ""
        ws.cell(row=14, column=14).value = ""
        
    # Format K10,K11,N10,N11, K13, K14, N13, N14 cells to be centered
    ws.cell(row=10, column=11).alignment = center
    ws.cell(row=11, column=11).alignment = center
    ws.cell(row=10, column=14).alignment = center
    ws.cell(row=11, column=14).alignment = center
    ws.cell(row=13, column=11).alignment = center
    ws.cell(row=14, column=11).alignment = center
    ws.cell(row=13, column=14).alignment = center
    ws.cell(row=14, column=14).alignment = center
    
    try:
        ws.column_dimensions["Z"].width = 11
    except Exception:
        pass

def draw_lower_boxes(ws, divider_top_row, blue_fill, black_bold, green_bold):
    # Shift amount (SHIFT = 2) is now incorporated into the numbers.
    
    # --- Box 1: Z:AE (Original 26-31) now maps to new columns 19-24 (Q->S, V->X) ---
    new_col_start1 = 19 # S
    new_col_end1 = 24 # X
    
    # New row range for Box 1 (5 cells, shifted down 1): divider_top_row - 3 to divider_top_row + 1
    for r in range(divider_top_row - 3, divider_top_row + 2):
        for c in range(new_col_start1, new_col_end1 + 1):
            cell = ws.cell(r, c)
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thick = Side(style="thick")
            
            # Top border is now at divider_top_row - 3
            top = thick if r == divider_top_row - 3 else None
            # Bottom border is at divider_top_row + 1
            bottom = thick if r == divider_top_row + 1 else None 
            left = thick if c == new_col_start1 else None
            right = thick if c == new_col_end1 else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Write headers 
    # Col 19 (S)
    ws.cell(divider_top_row - 3, 19, "Normalized").font = black_bold
    # Col 19 (S)
    ws.cell(divider_top_row - 2, 19, "VPDB").font = black_bold
    # Col 21 (U) and Col 22 (V) - Adjusted column indices based on Box 1 layout
    ws.cell(divider_top_row - 1, 22, "Calcite").font = black_bold
    ws.cell(divider_top_row - 1, 23, "Calcite").font = black_bold
    # Col 23 (W) - Aragonite
    ws.cell(divider_top_row - 1, 24, "Aragonite").font = green_bold
    
    # Row: divider_top_row
    # 19 (S)
    ws.cell(divider_top_row, 19, "δ¹³C")
    # 20 (T)
    ws.cell(divider_top_row, 20, "δ¹³C")
    # 21 (U)
    ws.cell(divider_top_row, 22, "δ¹⁸O")
    # 22 (V)
    ws.cell(divider_top_row, 23, "δ¹⁸O")
    # 23 (W)
    ws.cell(divider_top_row, 24, "δ¹⁸O")

    red_font = InlineFont(color='00FF0000', b=True)
    blue_font = InlineFont(color='000000FF', b=True)
    green_font = InlineFont(color='008000', b=True)
    
    # Write inline rich text
    # Row: divider_top_row + 1
    ws.cell(divider_top_row + 1, 19).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])
    ws.cell(divider_top_row + 1, 20).value = create_rich_text([(red_font, "18 "), (blue_font, "19 "), (green_font, "603")])
    ws.cell(divider_top_row + 1, 22).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])
    ws.cell(divider_top_row + 1, 23).value = create_rich_text([(red_font, "18 "), (blue_font, "19 "), (green_font, "603")])
    ws.cell(divider_top_row + 1, 24).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])
    
    # --- Box 2: AG:AH (Original 33–34) now maps to new columns 26-27 (Z->Z, AA->AA) ---
    new_col_start2 = 26 # Z
    new_col_end2 = 27 # AA
    
    # New row range for Box 2 (5 cells, shifted down 1): divider_top_row - 2 to divider_top_row + 2
    for r in range(divider_top_row - 2, divider_top_row + 2):
        for c in range(new_col_start2, new_col_end2 + 1):
            cell = ws.cell(r, c)
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Top border is now at divider_top_row - 2
            top = thick if r == divider_top_row - 2 else None
            # Bottom border is now at divider_top_row + 2
            bottom = thick if r == divider_top_row + 1 else None
            
            left = thick if c == new_col_start2 else None
            right = thick if c == new_col_end2 else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Write headers 
    # Col 26 (Z)
    ws.cell(divider_top_row - 2, 26, "VSMOW").font = black_bold
    # Col 26 (Z)
    ws.cell(divider_top_row - 1, 26, "Calcite").font = black_bold
    # Col 27 (AA)
    ws.cell(divider_top_row - 1, 27, "Aragonite").font = green_bold
    # Row: divider_top_row
    # Col 26 (Z)
    ws.cell(divider_top_row, 26, "δ¹⁸O")
    # Col 27 (AA)
    ws.cell(divider_top_row, 27, "δ¹⁸O")
    
    # Write inline rich text
    # Row: divider_top_row + 1
    ws.cell(divider_top_row + 1, 26).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])
    ws.cell(divider_top_row + 1, 27).value = create_rich_text([(red_font, "18 "), (blue_font, "19")])
    
    try:
        pass
    except Exception:
        pass

def _detect_decimal_places_from_format(fmt: str):
    if not fmt or not isinstance(fmt, str):
        return None
    first_section = fmt.split(';', 1)[0]
    if '.' not in first_section:
        return None
    after_dot = first_section.split('.', 1)[1]
    count = 0
    for ch in after_dot:
        if ch in ('0', '#'):
            count += 1
        else:
            break
    return count if count > 0 else None

# --- Main Function ---
def step6_normalization_carbonate(file_path):
    reference_names = ["CO2", "NBS 18", "NBS 19", "IAEA 603", "LSVEC", "HeCO2"]
    ref_set = {_normalize_text(r) for r in reference_names}
    
    # 🔴 Get the threshold setting
    try:
        threshold = settings.get_setting("STDEV_THRESHOLD")
    except AttributeError:
        # Fallback if settings.get_setting is not defined or fails
        threshold = 0.08
    
    # List of columns to EXCLUDE (1-based index)
    EXCLUDED_COLS = {8, 9, 11, 12, 13, 26, 27, 14, 15}
    MAX_SOURCE_COL = 24
    
    # Create a mapping of source column index -> destination column index
    col_map = {}
    dest_col = 1
    for src_col in range(1, MAX_SOURCE_COL + 1):
        if src_col not in EXCLUDED_COLS:
            col_map[src_col] = dest_col
            dest_col += 1
    
    # The identifier column (3, 'C') is preserved
    col_identifier1 = col_map.get(3, 0)
    if col_identifier1 != 3:
        raise Exception("Identifier column (3) was moved! Logic error in col_map.")
    
    # Use load_workbook from the import
    wb = openpyxl.load_workbook(file_path, data_only=False) 
    if "Last 6_DNT" not in wb.sheetnames:
        raise ValueError("Sheet 'Last 6' not found!")
    ws_last6 = wb["Last 6_DNT"]

    # Ensure Group sheet is recreated to the LEFT of "Group"
    if "Normalization_DNT" in wb.sheetnames:
        wb.remove(wb["Normalization_DNT"])
    pre_group_index = wb.sheetnames.index("Group_DNT")
    ws_group = wb.create_sheet("Normalization_DNT", pre_group_index)

    # make sure sheets are not grouped/selected together
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
            
    # mark Last 6 as the selected tab (prevents grouping with newly created sheet)
    try:
        ws_last6.sheet_view.tabSelected = True
        ws_group.sheet_view.tabSelected = False
    except Exception:
        pass
        
    blue_fill = _make_fill("DAE9F8")
    dark_fill = _make_fill("808080")
    gray_fill = _make_fill("E7E7E7")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    color_fonts = {
        "nbs18": Font(color="FF0000"),
        "nbs19": Font(color="000080"),
        "iaea603": Font(color="008000"),
        "lsvec": Font(color="3399FF"),
    }
    
    # --- Update MAX_GROUP_COL based on new exclusion list ---
    MAX_GROUP_COL = dest_col - 1
    
    # Fill top blue area (A1:W15) in the new Group sheet
    for row in ws_group.iter_rows(min_row=1, max_row=15, min_col=1, max_col=MAX_GROUP_COL):
        for cell in row:
            cell.fill = blue_fill
            
    # Copy headers, respecting the exclusion list AND copying styles (Row 18)
    headers = []
    first_row_cells = list(ws_last6[1]) if ws_last6.max_row >= 1 else []
    
    for src_col, dest_col_idx in col_map.items():
        src_cell = first_row_cells[src_col - 1] if src_col <= len(first_row_cells) else None
        
        # Copy header value
        header_val = src_cell.value if src_cell else None
        headers.append(header_val)
        dest_cell = ws_group.cell(row=18, column=dest_col_idx, value=header_val)
        
        # Copy style properties for header row
        if src_cell:
            dest_cell.number_format = copy(src_cell.number_format)
            dest_cell.font = copy(src_cell.font)
            dest_cell.alignment = copy(src_cell.alignment)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)

    # --- Collect original row index along with values and formats ---
    data_rows_with_index = []
    # Iterate over cell objects (not just values) to get the row index
    for row_idx, row_cells in enumerate(ws_last6.iter_rows(min_row=2, max_col=MAX_SOURCE_COL), start=2):
        row_values_filtered = []
        source_cell_formats = {}
        
        for src_col, cell in enumerate(row_cells, start=1):
            if src_col in col_map:
                row_values_filtered.append(cell.value)
                
                # --- Store a deep copy of all style attributes ---
                source_cell_formats[src_col] = {
                    'format': copy(cell.number_format),
                    'font': copy(cell.font), 
                    'fill': copy(cell.fill), 
                    'alignment': copy(cell.alignment),
                    'border': copy(cell.border)
                }
        
        if any(row_values_filtered):
            # Ensure the row has the correct number of columns
            row_values_filtered = row_values_filtered + [None] * (MAX_GROUP_COL - len(row_values_filtered))
            data_rows_with_index.append((row_idx, tuple(row_values_filtered), source_cell_formats))

    groups = {}

    # --- Grouping now stores (original_row_idx, row_values_filtered, source_cell_formats) ---
    for original_row_idx, r_values, r_formats in data_rows_with_index:
        # col_identifier1 is 3 for 'C' (Identifier) in the new, filtered list.
        ident = r_values[col_identifier1 - 1]
        base = extract_sample_base(ident)
        norm = _normalize_text(base)
        
        if norm not in groups:
            groups[norm] = {"base": base, "rows": []}
        groups[norm]["rows"].append((original_row_idx, r_values, r_formats)) # Store index, values, and formats
        
    for g in groups.values():
        # Sort based on the identifier in the 'row_values' tuple
        g["rows"].sort(key=lambda item: extract_run_number(item[1][col_identifier1 - 1]))
        
    ref_groups = []
    other_groups = []
    for norm, g in groups.items():
        if norm in ref_set:
            ref_groups.append((norm, g))
        else:
            other_groups.append((norm, g))

    # --- DETERMINE PRESENT REFERENCE MATERIALS ---
    present_refs_norm = {norm for norm, _ in ref_groups}
    has_iaea_603 = 'iaea603' in present_refs_norm
            
    current_row = 19
    # regex to detect "N Arag" or "N. Arag" (optional dot, optional spaces)
    n_arag_re = re.compile(r"\bn\.?\s*arag\b", flags=re.IGNORECASE)

    def _write_sample_output_cells(ws, r, is_arag, has_iaea_603, new_R_col, new_U_col, output_format):
        # Target Output Columns (Normalized Values)
        col_S = 19 # S: C VPDB (18 & 19 only)
        col_T = 20 # T: C VPDB (18 & 19 & 603)
        col_V = 22 # V: O VPDB (Calcite 18 & 19 only)
        col_W = 23 # W: O VPDB (Calcite 18 & 19 & 603)
        col_X = 24 # X: O VPDB (Aragonite 18 & 19)
        col_Z = 26 # Z: O VSMOW (Calcite 18 & 19)
        col_AA = 27 # AA: O VSMOW (Aragonite 18 & 19)
        
        # CALIBRATION FORMULA TEMPLATES (No 'r' here)
        c_18_19_formula = f'=IFERROR($K$10*{get_column_letter(new_R_col)}{r}+$K$11,"")'
        c_18_19_603_formula = f'=IFERROR($K$13*{get_column_letter(new_R_col)}{r}+$K$14,"")'
        o_18_19_formula = f'=IFERROR($N$10*{get_column_letter(new_U_col)}{r}+$N$11,"")'
        o_18_19_603_formula = f'=IFERROR($N$13*{get_column_letter(new_U_col)}{r}+$N$14,"")'

        # --- Delta 13C (S, T) ---
        # S (19): Normalized C (VPDB) (using only 18 and 19)
        cell_s = ws.cell(row=r, column=col_S, value=c_18_19_formula)
        cell_s.number_format = output_format
        
        # T (20): Normalized C (VPDB) (using 18, 19, and 603)
        if has_iaea_603:
            cell_t = ws.cell(row=r, column=col_T, value=c_18_19_603_formula)
            cell_t.number_format = output_format
        else:
            ws.cell(row=r, column=col_T, value=None)
        
        # --- Delta 18O (V, W, X, Z, AA) ---
        if is_arag:
            # --- ARAGONITE LOGIC ---
            ws.cell(row=r, column=col_V, value=None)
            ws.cell(row=r, column=col_W, value=None)
            
            # X (24): O VPDB (Aragonite 18 & 19 only)
            cell_x = ws.cell(row=r, column=col_X, value=o_18_19_formula)
            cell_x.font = Font(bold=True)
            cell_x.number_format = output_format
            
            if has_iaea_603:
                # If 603 is present, write Aragonite value into W (Calcite 18/19/603 column) as well
                cell_w_arag = ws.cell(row=r, column=col_W, value=o_18_19_603_formula)
                cell_w_arag.font = Font(bold=True)
                cell_w_arag.number_format = output_format
            
            # Z (26): O VSMOW (Calcite) - Empty for Aragonite
            ws.cell(row=r, column=col_Z, value=None) 
            
            # AA (27): O VSMOW (Aragonite)
            cell_aa = ws.cell(row=r, column=col_AA, value=f'=IFERROR((1.03092*{get_column_letter(col_X)}{r})+30.92,"")')
            cell_aa.font = Font(bold=True)
            cell_aa.number_format = output_format
            
            # Make entire row text green (Aragonite color)
            green_font = Font(color="008000")
            green_bold = Font(color="008000", bold=True)
            for c in range(1, col_AA + 1): 
                # Apply bold to the calculated columns S, T (if used), X, AA, and W (if used)
                if c in (col_S, col_T, col_X, col_AA) or (has_iaea_603 and c == col_W):
                    ws.cell(row=r, column=c).font = green_bold
                else:
                    # preserve original font if possible, but change color
                    original_font = copy(ws.cell(row=r, column=c).font)
                    original_font.color = green_font.color
                    ws.cell(row=r, column=c).font = original_font
        else:
            # --- CALCITE LOGIC (default) ---
            
            # V (22): O VPDB (Calcite 18 & 19 only)
            cell_v = ws.cell(row=r, column=col_V, value=o_18_19_formula)
            cell_v.number_format = output_format
            
            # W (23): O VPDB (Calcite 18 & 19 & 603)
            if has_iaea_603:
                cell_w = ws.cell(row=r, column=col_W, value=o_18_19_603_formula)
                cell_w.number_format = output_format
            else:
                ws.cell(row=r, column=col_W, value=None)
            
            # X (24): O VPDB (Aragonite) - Empty for Calcite
            ws.cell(row=r, column=col_X, value=None)
            
            # Z (26): O VSMOW (Calcite)
            cell_z = ws.cell(row=r, column=col_Z, value=f'=IFERROR((1.03092*{get_column_letter(col_V)}{r})+30.92,"")')
            cell_z.font = Font(bold=True)
            cell_z.number_format = output_format

            # AA (27): O VSMOW (Aragonite) - Empty for Calcite
            ws.cell(row=r, column=col_AA, value=None)

            # Apply bold to the calculated columns (S, V, Z) + (T, W if used)
            bold_cols = {col_S, col_V, col_Z}
            if has_iaea_603:
                bold_cols.add(col_T)
                bold_cols.add(col_W)
                
            for c in bold_cols:
                original_font = copy(ws.cell(row=r, column=c).font)
                original_font.bold = True
                ws.cell(row=r, column=c).font = original_font
        
        return col_AA # Return the max column used for data/formulas

    def write_group(norm, g, is_reference=True, has_iaea_603=False):
        nonlocal current_row
        base_name = _normalize_text(g["base"])
        rows_data = g["rows"] # This now contains (source_row_idx, row_values_filtered, source_cell_formats)
        start_row = current_row
        font_color = color_fonts.get(base_name)
        
        # Get row_values_filtered only for CO2 check
        row_values_list = [item[1] for item in rows_data]
        valid_indices = []
        if base_name in ("co2", "heco2"):
            valid_indices = _get_valid_co2_rows(row_values_list, col_identifier1)
            
        row_map = []
        max_col_written = 0

        # --- COPY DATA AND RETAIN STYLES ---
        for i, (source_row_idx, row_values_filtered, source_cell_formats) in enumerate(rows_data):
            excel_row = current_row
            
            # Map source_cell_formats (using original 1-24 keys) to the new dest_col (1-15 keys)
            for dest_col_idx in range(1, MAX_GROUP_COL + 1):
                # Find the source column index that maps to this dest_col_idx
                src_col = next(s for s, d in col_map.items() if d == dest_col_idx)

                val = row_values_filtered[dest_col_idx - 1]
                dest_cell = ws_group.cell(row=excel_row, column=dest_col_idx, value=val)
                
                # Get the original cell's formatting (already deep-copied)
                original_format = source_cell_formats.get(src_col, {})
                
                # Apply number_format, font, fill, alignment, and border
                if original_format.get('format') is not None:
                    dest_cell.number_format = original_format['format']
                if original_format.get('font') is not None:
                    dest_cell.font = original_format['font']
                if original_format.get('fill') is not None:
                    dest_cell.fill = original_format['fill']
                if original_format.get('alignment') is not None:
                    dest_cell.alignment = original_format['alignment']
                if original_format.get('border') is not None:
                    dest_cell.border = original_format['border']

                # Identifier font color override (applies to the new column 3)
                if font_color and dest_col_idx == col_identifier1:
                    # Merge the custom color with the copied font style
                    new_font = copy(dest_cell.font)
                    new_font.color = font_color.color
                    dest_cell.font = new_font
                    
            # =======================
            # CO₂ / HeCO₂ GRAY FILL OVERRIDE
            # =======================
            if base_name in ("co2", "heco2") and i in valid_indices:
                # Apply gray fill to the full width of the copied columns (1 to MAX_GROUP_COL)
                for col in range(1, MAX_GROUP_COL + 1):
                    ws_group.cell(row=excel_row, column=col).fill = gray_fill
                    
            if not is_reference:
                # Get the Identifier for aragonite check
                ident_val = str(ws_group.cell(row=excel_row, column=col_identifier1).value or "")
                is_arag = n_arag_re.search(ident_val)
                # Write individual sample data normalized values
                max_col_written = _write_sample_output_cells(
                    ws_group, excel_row, is_arag, has_iaea_603, 11, 14, '0.00'
                )

            row_map.append(excel_row)
            current_row += 1
            
        end_row = current_row - 1

        # NOTE: The summary calculation columns are now: 
        # C_Avg: K(11), C_Stdev: L(12), C_Count: M(13)
        # O_Avg: N(14), O_Stdev: O(15), O_Count: P(16)
        c_avg_col = 11
        c_stdev_col = 12
        c_count_col = 13
        o_avg_col = 14
        o_stdev_col = 15
        o_count_col = 16

        # --- SUMMARY CALCULATION (For Reference Materials AND Samples with > 1 run) ---
        # The logic is slightly different for CO2/HeCO2 groups.
        if is_reference or len(rows_data) > 1:
            # The row numbers (r) for the ranges are the same for all summary calculations
            if base_name in ("co2", "heco2") and valid_indices:
                # Use only the row numbers corresponding to the valid indices
                valid_rows = [row_map[i] for i in valid_indices]
                r_ranges = ",".join([f"{get_column_letter(c_avg_col)}{r}" for r in valid_rows])
                u_ranges = ",".join([f"{get_column_letter(o_avg_col)}{r}" for r in valid_rows])
            elif row_map: # Use full continuous range for other references or multi-run samples
                r_ranges = f"{get_column_letter(c_avg_col)}{start_row}:{get_column_letter(c_avg_col)}{end_row}"
                u_ranges = f"{get_column_letter(o_avg_col)}{start_row}:{get_column_letter(o_avg_col)}{end_row}"
            else:
                # Handle case where group is present but has no data rows (shouldn't happen, but safe)
                r_ranges = ""
                u_ranges = ""

            # Label rows
            for col_offset, label in enumerate(["Average", "Stdev", "Count"], start=0):
                cell = ws_group.cell(row=current_row, column=c_avg_col + col_offset, value=label)
                cell.alignment = Alignment(horizontal="right")
                cell2 = ws_group.cell(row=current_row, column=o_avg_col + col_offset, value=label)
                cell2.alignment = Alignment(horizontal="right")
                
            avg_row = current_row + 1
            
            data_format = get_summary_num_format(base_name)
            count_format = '0'
            
            # CARBON (K, L, M)
            ws_group.cell(row=avg_row, column=c_avg_col, value=f"=AVERAGE({r_ranges})").number_format = data_format
            ws_group.cell(row=avg_row, column=c_stdev_col, value=f"=STDEV({r_ranges})").number_format = data_format
            ws_group.cell(row=avg_row, column=c_count_col, value=f"=COUNT({r_ranges})").number_format = count_format
            
            # OXYGEN (N, O, P)
            ws_group.cell(row=avg_row, column=o_avg_col, value=f"=AVERAGE({u_ranges})").number_format = data_format
            ws_group.cell(row=avg_row, column=o_stdev_col, value=f"=STDEV({u_ranges})").number_format = data_format
            ws_group.cell(row=avg_row, column=o_count_col, value=f"=COUNT({u_ranges})").number_format = count_format

            # Apply font/color styles to labels and formulas
            # For non-reference samples, use bold black font for the summary.
            summary_font = font_color or Font(bold=True) 

            if is_reference and font_color:
                # 1. Create a copy of the colored font.
                ref_summary_font = copy(font_color)
                # 2. Force the bold attribute to True.
                ref_summary_font.bold = True
                summary_font = ref_summary_font
            
            for col in range(c_avg_col, o_count_col + 1):
                ws_group.cell(row=current_row, column=col).font = summary_font # Labels row
                ws_group.cell(row=avg_row, column=col).font = summary_font     # Formulas row
                
            # --- NON-REFERENCE SAMPLE SPECIFIC: CALCULATE NORMALIZED VALUES ON SUMMARY ROW ---
            if not is_reference:
                # Use the calculated Average column values (K and N) as input for the Normalized columns (S, T, V, W, X, Z, AA)
                
                # Check aragonite based on the BASE name of the group, not the individual identifier
                is_arag_group = n_arag_re.search(g["base"])
                
                # Write normalized values for the summary row (avg_row)
                max_col_written = _write_sample_output_cells(
                    ws_group, avg_row, is_arag_group, has_iaea_603, c_avg_col, o_avg_col, '0.00'
                )

                # Apply bold to all calculated summary cells (K-P, S, T, V, W, X, Z, AA)
                for c in range(c_avg_col, o_count_col + 1):
                    ws_group.cell(row=avg_row, column=c).font = Font(bold=True)
                # Also apply bold to the output columns written above, up to max_col_written
                for c in range(19, max_col_written + 1):
                    ws_group.cell(row=avg_row, column=c).font = Font(bold=True)
                
            current_row += 3
        else:
            # For non-reference groups with only 1 row, only advance the row
            current_row += 3
            
    # Write reference groups first
    for norm, g in ref_groups:
        write_group(norm, g, is_reference=True)
        
    # Divider
    if ref_groups:
        current_row += 8
        divider_top_row = current_row # store divider start row
        for _ in range(2):
            for col in range(1, 702):
                ws_group.cell(row=current_row, column=col).fill = dark_fill
            current_row += 1
            
        # Copy headers again below the divider, respecting the exclusion list AND copying styles (Row 18)
        current_header_col = 1
        for src_col, dest_col_idx in col_map.items():
            src_cell = first_row_cells[src_col - 1] if src_col <= len(first_row_cells) else None
            
            dest_cell = ws_group.cell(row=current_row, column=current_header_col, value=headers[current_header_col - 1])
            
            # Copy style properties for header row
            if src_cell:
                dest_cell.number_format = copy(src_cell.number_format)
                dest_cell.font = copy(src_cell.font)
                dest_cell.alignment = copy(src_cell.alignment)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)

            current_header_col += 1

        current_row += 1
        
        # Call the updated lower box drawing function with new column indices
        draw_lower_boxes(ws_group, divider_top_row, blue_fill, Font(bold=True, color="000000"), Font(bold=True, color="008000"))
        
    # Write non-reference groups
    # Pass the result of the reference check to the function call
    for norm, g in other_groups:
        write_group(norm, g, is_reference=False, has_iaea_603=has_iaea_603)
        
    # Fill measured columns with grey background (K and N)
    max_row = ws_group.max_row + 50
    for row in range(16, max_row + 1):
        for col in (11, 14): # New Measured C (K) and Measured O (N) columns
            ws_group.cell(row=row, column=col).fill = gray_fill
            
    # Set column widths
    ws_group.column_dimensions["C"].width = 22 
    ws_group.column_dimensions["H"].width = 15 
    
    # --- Call it after filling the groups ---
    add_blue_box(ws_group)
    
    # 🔴 APPLY DYNAMIC CONDITIONAL FORMATTING (Updated with ISNUMBER check)
    max_data_row = ws_group.max_row
    threshold_str = str(threshold)
    
    # Column L (12): Standard Deviation of Carbon
    # Formula: =AND(ISNUMBER(L19), L19 > 0.08)
    ws_group.conditional_formatting.add(
        f"L19:L{max_data_row}",
        FormulaRule(
            formula=[f'=AND(ISNUMBER(L19), L19 > {threshold_str})'],
            fill=red_fill
        )
    )

    # Column O (15): Standard Deviation of Oxygen
    # Formula: =AND(ISNUMBER(O19), O19 > 0.08)
    ws_group.conditional_formatting.add(
        f"O19:O{max_data_row}",
        FormulaRule(
            formula=[f'=AND(ISNUMBER(O19), O19 > {threshold_str})'],
            fill=red_fill
        )
    )
    # ------------------------------------------------------------------------

    ws_group.freeze_panes = 'B19'

    wb.save(file_path)
    print(f"✅ Step 6: Normalization completed on {file_path}")