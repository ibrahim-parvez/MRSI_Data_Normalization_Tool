import pandas as pd
import re
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule 
import settings 

# --- Helper Functions ---

def get_base_key(identifier):
    """
    Strips trailing run numbers (e.g., 'Sample r1', 'Sample r1.2') 
    to get the base name for grouping.
    """
    if identifier is None:
        return None
    s_id = str(identifier)
    # Remove trailing ' r1', ' r1.1', etc.
    return re.sub(r'\s+[rR]\d+(?:\.\d+)*(?:[a-zA-Z]*)?$', '', s_id).strip()

def copy_cell_style(src_cell, tgt_cell):
    """Copies Font, Border, Fill, Alignment, and NumberFormat."""
    if src_cell is None or tgt_cell is None: return
    if src_cell.has_style:
        try:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = src_cell.number_format 
            tgt_cell.alignment = copy(src_cell.alignment)
            tgt_cell.protection = copy(src_cell.protection)
        except: pass

def step5_group_carbonate(file_path: str):
    """
    Step 5: Group Carbonate (Non-Incremental)
    - Groups items by Base Key, collecting scattered rows into contiguous blocks.
    - Removes Blue Box/Normalization math.
    - Preserves Source Styling and Carbonate Column Mapping.
    - Adds Summary Stats (Avg/Stdev/Count) for C and O.
    - Places tab to the left of 'Pre-Group_DNT'.
    """

    new_sheet_name = "Group_DNT"
    source_sheet_name = "Last 6_DNT"
    target_sibling_name = "Pre-Group_DNT" # Tab to place the new sheet next to

    # --- 1. Settings & Styles ---
    # Get Threshold
    try:
        stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    except AttributeError:
        stdev_threshold = 0.08 # Fallback
    
    # Fills
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red for high Stdev
    gray_fill = PatternFill(start_color="E7E7E7", end_color="E7E7E7", fill_type="solid")  # Gray for data cols
    header_fill = PatternFill(start_color="8ED973", end_color="8ED973", fill_type="solid") # Green for headers
    bold_font = Font(bold=True)

    # Number Format
    THREE_DECIMAL_FORMAT = "0.000"

    # --- 2. Load Workbook ---
    wb_values = load_workbook(file_path, data_only=True) # To read calculated values
    wb = load_workbook(file_path) # To read styles

    # Find source sheet
    matched_source = next((s for s in wb.sheetnames if s.lower() == source_sheet_name.lower()), None)
    if matched_source is None:
        print(f"❌ Source sheet matching '{source_sheet_name}' not found.")
        return

    source_ws_vals = wb_values[matched_source]
    source_ws = wb[matched_source]

    # Delete existing Group sheet if exists
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    # --- POSITIONING LOGIC ---
    # Try to find Pre-Group_DNT to place it to the left
    if target_sibling_name in wb.sheetnames:
        idx = wb.sheetnames.index(target_sibling_name)
    else:
        # Fallback: Place it after the source sheet if Pre-Group doesn't exist
        idx = wb.index(source_ws) + 1
    
    new_ws = wb.create_sheet(new_sheet_name, idx)

    # --- 3. Define Column Mapping (Carbonate Specific) ---
    EXCLUDED_COLS = {8, 9, 11, 12, 13, 26, 27, 14, 15}
    MAX_SOURCE_COL = 24
    
    # source_col_index -> dest_col_index
    col_map = {}
    dest_col_counter = 1
    source_cols_ordered = [] 

    for src_c in range(1, MAX_SOURCE_COL + 1):
        if src_c not in EXCLUDED_COLS:
            col_map[src_c] = dest_col_counter
            source_cols_ordered.append(src_c)
            dest_col_counter += 1
            
    # Key Columns in Destination (for Summaries)
    col_C_meas = 11
    col_O_meas = 14
    
    # Identifier Column (Source col 3 usually maps to Dest col 3)
    col_id_dest = col_map.get(3, 3)

    # --- 4. Setup Headers ---
    src_header_row = list(source_ws[1])
    
    for src_c, dest_c in col_map.items():
        if src_c <= len(src_header_row):
            src_cell = source_ws.cell(row=1, column=src_c)
            dest_cell = new_ws.cell(row=1, column=dest_c, value=src_cell.value)
            copy_cell_style(src_cell, dest_cell)
            if not src_cell.fill or src_cell.fill.patternType is None:
                dest_cell.fill = header_fill

    # Set Column Widths
    new_ws.column_dimensions[get_column_letter(3)].width = 22 # Identifier
    new_ws.column_dimensions[get_column_letter(11)].width = 12 # C Meas
    new_ws.column_dimensions[get_column_letter(14)].width = 12 # O Meas
    new_ws.freeze_panes = "A2"

    # --- 5. Data Collection (Grouping Phase) ---
    # Instead of writing immediately, we collect rows into a dictionary based on Base Key.
    # Structure: { 'HeCo2': [row_idx_1, row_idx_15, row_idx_99], ... }
    
    grouped_rows = {}
    max_row = source_ws.max_row
    
    # Iterate source rows to group them
    for r in range(2, max_row + 1):
        if source_ws.row_dimensions[r].hidden:
            continue
            
        # Get Identifier
        id_val = source_ws_vals.cell(row=r, column=3).value
        base_key = get_base_key(id_val)
        
        # If ID is missing, treat as "Unknown" or skip? 
        # Usually valid data has an ID. We treat None as a group key "None"
        if base_key not in grouped_rows:
            grouped_rows[base_key] = []
        
        grouped_rows[base_key].append(r)

    # --- 6. Writing Output (Processing Phase) ---
    
    dest_row = 2 
    
    # Iterate through our groups
    for base_key, row_indices in grouped_rows.items():
        
        group_start_dest_row = dest_row
        
        # A. Write all rows for this group
        for src_row in row_indices:
            
            for src_c in source_cols_ordered:
                dest_c = col_map[src_c]
                
                # Value
                val = source_ws_vals.cell(row=src_row, column=src_c).value
                dest_cell = new_ws.cell(row=dest_row, column=dest_c, value=val)
                
                # Style
                src_cell = source_ws.cell(row=src_row, column=src_c)
                copy_cell_style(src_cell, dest_cell)
            
            # Apply Gray Fill to Measured Columns
            new_ws.cell(row=dest_row, column=col_C_meas).fill = gray_fill
            new_ws.cell(row=dest_row, column=col_O_meas).fill = gray_fill
            
            # Copy Height
            try:
                rh = source_ws.row_dimensions[src_row].height
                if rh is not None:
                    new_ws.row_dimensions[dest_row].height = rh
            except: pass
            
            dest_row += 1
            
        # B. Group Finished. Insert Summary Block.
        group_end_dest_row = dest_row - 1
        
        label_row = dest_row
        calc_row = dest_row + 1
        
        # Column Letters
        c_avg_let = get_column_letter(col_C_meas)     # K
        o_avg_let = get_column_letter(col_O_meas)     # N
        
        # 1. Write Labels
        labels = ["Average", "Stdev", "Count"]
        
        # Carbon Labels
        for i, txt in enumerate(labels):
            cell = new_ws.cell(row=label_row, column=col_C_meas + i, value=txt)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='right')
        
        # Oxygen Labels
        for i, txt in enumerate(labels):
            cell = new_ws.cell(row=label_row, column=col_O_meas + i, value=txt)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='right')

        # 2. Write Formulas
        # Carbon
        rng_c = f"{c_avg_let}{group_start_dest_row}:{c_avg_let}{group_end_dest_row}"
        new_ws.cell(row=calc_row, column=col_C_meas, value=f"=AVERAGE({rng_c})").number_format = THREE_DECIMAL_FORMAT
        new_ws.cell(row=calc_row, column=col_C_meas+1, value=f"=STDEV({rng_c})").number_format = THREE_DECIMAL_FORMAT
        new_ws.cell(row=calc_row, column=col_C_meas+2, value=f"=COUNT({rng_c})").number_format = "0"
        
        # Oxygen
        rng_o = f"{o_avg_let}{group_start_dest_row}:{o_avg_let}{group_end_dest_row}"
        new_ws.cell(row=calc_row, column=col_O_meas, value=f"=AVERAGE({rng_o})").number_format = THREE_DECIMAL_FORMAT
        new_ws.cell(row=calc_row, column=col_O_meas+1, value=f"=STDEV({rng_o})").number_format = THREE_DECIMAL_FORMAT
        new_ws.cell(row=calc_row, column=col_O_meas+2, value=f"=COUNT({rng_o})").number_format = "0"
        
        # Bold the calc row
        for c in range(col_C_meas, col_O_meas + 3):
            new_ws.cell(row=calc_row, column=c).font = bold_font

        # Update pointers (Space between groups)
        dest_row = calc_row + 2 

    # --- 7. Finalize & Conditional Formatting ---
    for s in wb.worksheets: s.sheet_view.tabSelected = False
    new_ws.sheet_view.tabSelected = True
    wb.active = wb.index(new_ws)
    new_ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    
    final_max_row = new_ws.max_row
    
    if stdev_threshold is not None:
        thresh_str = str(stdev_threshold)
        
        # Column L: Carbon Stdev
        rule_L = FormulaRule(
            formula=[f'AND(ISNUMBER(L2), L2 > {thresh_str})'],
            fill=fill_error
        )
        new_ws.conditional_formatting.add(f"L2:L{final_max_row}", rule_L)
        
        # Column O: Oxygen Stdev
        rule_O = FormulaRule(
            formula=[f'AND(ISNUMBER(O2), O2 > {thresh_str})'],
            fill=fill_error
        )
        new_ws.conditional_formatting.add(f"O2:O{final_max_row}", rule_O)

    wb.save(file_path)
    print(f"✅ Step 5: Group sheet '{new_sheet_name}' created.")