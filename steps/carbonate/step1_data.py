import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.views import Selection
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule

# Load global threshold from settings.py
import settings


def step1_data_carbonate(file_path, sheet_name='Default_Gas_Bench.wke'):
    """
    Step 1: DATA
    Reads the Excel file, generates padded data rows,
    summary metrics, formulas, and conditional formatting.

    - Formulas = full precision
    - Display formatting handles decimals
    - Conditional formatting highlights STDEV only for "last 6"
    """

    # Use global threshold
    stdev_threshold = settings.get_setting("STDEV_THRESHOLD")

    new_sheet_name = 'Data_DNT'

    # Read original data
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')

    # Data sheet headers:
    headers = [
        'Row', 'Time Code', 'Identifier 1', 'Comment', 'Identifier 2', 'Analysis',
        'Preparation', 'Peak Nr', 'Rt', 'Ampl 44', 'Area All',
        'd 13C/12C', 'd 18O/16O',
        '', '', '', '',  # spacer columns
        'C avg', 'C stdev', '', 'O avg', 'O stdev', '',
        'Sum area all', 'area peaks', 'funny peaks', 'min intensity'
    ]

    wb = load_workbook(file_path)
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    first_index = wb.index(wb[sheet_name])
    ws = wb.create_sheet(new_sheet_name, first_index)

    # Select tab
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except:
            pass
    ws.sheet_view.tabSelected = True
    wb.active = wb.index(ws)
    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Write headers
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

        # --- COLOR HEADER ROW ---
    green_fill = PatternFill(start_color="8ed973", end_color="8ed973", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Columns needing yellow fill
    yellow_cols = ["H", "I", "K", "L", "M", "Z", "AA"]

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx)

        if col_letter in yellow_cols:
            cell.fill = yellow_fill
        else:
            cell.fill = green_fill


    cur_row = 3

    # Header → column index
    col_map = {h: i + 1 for i, h in enumerate(headers) if h}

    # Normalization helper
    def normalize_name(s):
        if s is None:
            return ''
        return ' '.join(str(s).split()).lower()

    df_cols_norm = {normalize_name(c): c for c in df.columns}

    # Map headers to df columns
    header_to_dfcol = {}
    for h in headers:
        if not h:
            header_to_dfcol[h] = None
            continue
        nh = normalize_name(h)
        if nh in df_cols_norm:
            header_to_dfcol[h] = df_cols_norm[nh]
            continue
        nh_join = nh.replace(' ', '')
        match = None
        for dc_norm, dc in df_cols_norm.items():
            if dc_norm.replace(' ', '') == nh_join:
                match = dc
                break
        if match:
            header_to_dfcol[h] = match
            continue
        match = None
        for dc_norm, dc in df_cols_norm.items():
            if nh in dc_norm or dc_norm in nh or nh_join in dc_norm.replace(' ', ''):
                match = dc
                break
        header_to_dfcol[h] = match

    # Key column indices
    col_area = col_map.get('Area All')
    col_c = col_map.get('d 13C/12C')
    col_o = col_map.get('d 18O/16O')
    col_ampl = col_map.get('Ampl 44')
    col_funny = col_map.get('funny peaks')
    col_minint = col_map.get('min intensity')

    check_headers = ['Rt', 'Ampl 44', 'Area All', 'd 13C/12C', 'd 18O/16O']
    check_df_cols = [header_to_dfcol.get(h) for h in check_headers if header_to_dfcol.get(h)]

    col_letter_area = get_column_letter(col_area)
    col_letter_c = get_column_letter(col_c)
    col_letter_o = get_column_letter(col_o)
    col_letter_ampl = get_column_letter(col_ampl)

    # Summary row layout
    summary_layout = [
        ("ref avg", 0),
        ("all", 3),
        ("last 6", 0),
        ("start", 2),
        ("end", 0),
        ("delta", 0),
    ]

    # Summary column offsets
    col_label = 17  # Q
    col_c_avg = col_label + 1  # R
    col_c_stdev = col_label + 2  # S
    col_o_avg = col_label + 4  # U
    col_o_stdev = col_label + 5  # V
    col_sum_area = col_label + 7  # X

    # Formatting
    fill_label = PatternFill(start_color="cdffcc", end_color="cdffcc", fill_type="solid")
    fill_funny_min = PatternFill(start_color="cdfeff", end_color="cdfeff", fill_type="solid")
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    group_highlight_ranges = []
    group_spacer_rows = []
    all_delta_rows = []

    grouped = df.groupby('Row', sort=False)

    # PROCESS GROUPS
    for Row, group in grouped:

        # Blank line between groups
        if cur_row != 3:
            cur_row += 1

        first_data_row = cur_row
        row_count = max(11, len(group))

        # Pad group rows to 11
        padded_rows = []
        for i in range(row_count):
            if i < len(group):
                padded_rows.append(group.iloc[i].to_dict())
            else:
                base_vals = group.iloc[0].to_dict() if len(group) else {}
                new_blank = {col: None for col in df.columns}
                for nm in ["Row", "Time Code", "Identifier 1"]:
                    new_blank[nm] = base_vals.get(nm)
                new_blank["Peak Nr"] = i + 1
                padded_rows.append(new_blank)

        # Write data rows
        for row_dict in padded_rows:
            for h in headers:
                if not h or h == "Sum area all":
                    continue
                excel_col = col_map[h]
                source_col = header_to_dfcol.get(h)
                val = row_dict.get(source_col) if source_col else None
                cell = ws.cell(row=cur_row, column=excel_col, value=val)
                if h in ("Identifier 2", "Analysis") and val is not None:
                    cell.number_format = '@'
            cur_row += 1

        last_data_row = cur_row - 1
        last7_start = max(first_data_row, last_data_row - 6)
        last6_start = max(first_data_row, last_data_row - 5)
        start_last6 = first_data_row + 5 if (first_data_row + 5) <= last_data_row else first_data_row

        summary_row = first_data_row
        row_positions = {}

        data_count = group[check_df_cols].dropna(how='all').shape[0]

        # SUMMARY FORMULAS
        for label, spacing in summary_layout:
            summary_row += spacing
            row_positions[label] = summary_row

            ws.cell(summary_row, col_label, value=label)

            if label == "last 6":
                target_col = col_c_stdev + 1
                if data_count < 11:
                    c = ws.cell(summary_row, target_col, value=f"< 11: {data_count}")
                    c.fill = fill_error
                    c.font = Font(bold=True)
                elif data_count > 11:
                    c = ws.cell(summary_row, target_col, value=f"> 11: {data_count}")
                    c.fill = fill_error
                    c.font = Font(bold=True)

            # --- ref avg ---
            if label == "ref avg":
                idx1, idx2, idx4 = first_data_row, first_data_row + 1, first_data_row + 3

                ws.cell(summary_row, col_c_avg,
                        value=f"=AVERAGE({col_letter_c}{idx1},{col_letter_c}{idx2},{col_letter_c}{idx4})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_c_stdev,
                        value=f"=STDEV({col_letter_c}{idx1},{col_letter_c}{idx2},{col_letter_c}{idx4})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_o_avg,
                        value=f"=AVERAGE({col_letter_o}{idx1},{col_letter_o}{idx2},{col_letter_o}{idx4})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_o_stdev,
                        value=f"=STDEV({col_letter_o}{idx1},{col_letter_o}{idx2},{col_letter_o}{idx4})"
                        ).number_format = "0.000"

            # --- all ---
            elif label == "all":
                ws.cell(summary_row, col_c_avg,
                        value=f"=AVERAGE({col_letter_c}{last7_start}:{col_letter_c}{last_data_row})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_c_stdev,
                        value=f"=STDEV({col_letter_c}{last7_start}:{col_letter_c}{last_data_row})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_o_avg,
                        value=f"=AVERAGE({col_letter_o}{last7_start}:{col_letter_o}{last_data_row})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_o_stdev,
                        value=f"=STDEV({col_letter_o}{last7_start}:{col_letter_o}{last_data_row})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_sum_area,
                        value=f"=SUM({col_letter_area}{last7_start}:{col_letter_area}{last_data_row})"
                        ).number_format = "0.00"

            # --- last 6 ---
            elif label == "last 6":
                ws.cell(summary_row, col_c_avg,
                        value=f"=AVERAGE({col_letter_c}{last6_start}:{col_letter_c}{last_data_row})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_c_stdev,
                        value=f"=STDEV({col_letter_c}{last6_start}:{col_letter_c}{last_data_row})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_o_avg,
                        value=f"=AVERAGE({col_letter_o}{last6_start}:{col_letter_o}{last_data_row})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_o_stdev,
                        value=f"=STDEV({col_letter_o}{last6_start}:{col_letter_o}{last_data_row})"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_sum_area,
                        value=f"=SUM({col_letter_area}{last6_start}:{col_letter_area}{last_data_row})"
                        ).number_format = "0.00"

            # --- start ---
            elif label == "start":
                ws.cell(summary_row, col_c_avg,
                        value=f"={col_letter_c}{start_last6}"
                        ).number_format = "0.000"
                ws.cell(summary_row, col_o_avg,
                        value=f"={col_letter_o}{start_last6}"
                        ).number_format = "0.000"

            # --- end ---
            elif label == "end":
                ws.cell(summary_row, col_c_avg,
                        value=f"={col_letter_c}{last_data_row}"
                        ).number_format = "0.000"
                ws.cell(summary_row, col_o_avg,
                        value=f"={col_letter_o}{last_data_row - 1}"
                        ).number_format = "0.000"

            # --- delta ---
            elif label == "delta":
                sr = row_positions["start"]
                er = row_positions["end"]

                ws.cell(summary_row, col_c_avg,
                        value=f"={get_column_letter(col_c_avg)}{er}-{get_column_letter(col_c_avg)}{sr}"
                        ).number_format = "0.000"

                ws.cell(summary_row, col_o_avg,
                        value=f"={get_column_letter(col_o_avg)}{er}-{get_column_letter(col_o_avg)}{sr}"
                        ).number_format = "0.000"

            summary_row += 1

        # Track areas for highlighting
        delta_row = row_positions["delta"]
        all_delta_rows.append(delta_row)
        highlight_end = max(last_data_row, delta_row)
        group_highlight_ranges.append((first_data_row, highlight_end))
        group_spacer_rows.append(last_data_row + 1)

        # Funny/min-intensity calculations
        if col_letter_ampl and col_funny and col_minint:
            for i in range(row_count):
                r = first_data_row + i

                # Funny Peaks logic (unchanged)
                if i < 4:
                    ws.cell(r, col_funny, value="ref")
                else:
                    ws.cell(
                        r, col_funny,
                        value=f'=IF({col_letter_ampl}{r}>{col_letter_ampl}{r+1},'
                            f'IF({col_letter_ampl}{r+1}<{col_letter_ampl}{r},"ok","check"),"check")'
                    )

                # --- MIN INTENSITY LOGIC (AA COLUMN) ---
                if i == 0:
                    # First row of group → text
                    ws.cell(r, col_minint, value="if 44<1000")
                elif i < 4:
                    # Rows 2–4 → blank
                    ws.cell(r, col_minint, value="")
                else:
                    # Row 5 onward → formula
                    ws.cell(
                        r, col_minint,
                        value=f'=IF({col_letter_ampl}{r}<400,"check","ok")'
                    )


        # -------------------- CONDITIONAL FORMATTING FOR “LAST 6” --------------------

        last6_row = row_positions.get("last 6")
        if last6_row:
            ws.conditional_formatting.add(
                f"{get_column_letter(col_c_stdev)}{last6_row}",
                CellIsRule(
                    operator="greaterThan",
                    formula=[str(stdev_threshold)],
                    fill=fill_error
                )
            )

            ws.conditional_formatting.add(
                f"{get_column_letter(col_o_stdev)}{last6_row}",
                CellIsRule(
                    operator="greaterThan",
                    formula=[str(stdev_threshold)],
                    fill=fill_error
                )
            )

    # Apply background fills
    max_row = ws.max_row
    for s, e in group_highlight_ranges:
        s = max(2, s)
        e = min(max_row, e)
        for r in range(s, e + 1):
            ws.cell(r, col_label).fill = fill_label
            ws.cell(r, col_funny).fill = fill_funny_min
            ws.cell(r, col_minint).fill = fill_funny_min

    # Clear spacer rows
    for spacer in group_spacer_rows:
        for c in (col_label, col_funny, col_minint):
            ws.cell(spacer, c).fill = PatternFill(fill_type=None)

    wb.save(file_path)
    print(f"Step 1: Data completed on {file_path}")
