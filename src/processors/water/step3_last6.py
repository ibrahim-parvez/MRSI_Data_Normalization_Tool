import string
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from copy import copy
import utils.settings as settings
from utils.common_utils import embed_settings_popup

def step3_last6_water(file_path: str):
    """
    Step 3: Last 6 (Water)
    - Specifically creates a sheet called 'Last 6_DNT'.
    - Copies from 'To Sort_DNT'.
    - Dynamic Filter: Shows 'last 6' OR 'outliers excl.' based on Settings.
    - Applies Conditional Formatting for Stdev thresholds.
    """
    
    source_sheet_name = "To Sort_DNT"
    new_sheet_name = "Last 6_DNT"
    
    # --- Check Calculation Mode ---
    # Options: "Last 6" or "Last 6 Outliers Excluded"
    calc_mode = settings.get_setting("CALC_MODE_STEP3")
    
    if calc_mode == "Last 6 Outliers Excl.":
        # Matches the label we wrote in Step 1 (row 7 of calc box)
        fixed_filter = "last 6 outliers excl." 
    else:
        fixed_filter = "last 6"

    print(f"   ℹ️ Step 3 Calculation Mode: {calc_mode} -> Filtering for '{fixed_filter}'")

    wb = load_workbook(file_path)

    if source_sheet_name not in wb.sheetnames:
        raise ValueError(f"Source sheet '{source_sheet_name}' not found. Please run Step 2 first.")
    
    # --- 1. Cleanup existing sheet ---
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    source_ws = wb[source_sheet_name]
    
    # Position it to the left of 'To Sort_DNT'
    idx = wb.index(source_ws)
    new_ws = wb.create_sheet(new_sheet_name, idx)
    
    # --- 2. Cell-by-Cell Copy ---
    for row in source_ws.iter_rows():
        for cell in row:
            cell_value = cell.value
            if cell.data_type == 'f' and hasattr(cell, 'formula'):
                cell_value = cell.formula
            
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell_value)
            
            if cell.has_style:
                try:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                except:
                    pass

    # --- 3. Dynamic Filter Logic ---
    last_row = new_ws.max_row
    col_o_idx = 15 # Column O (Labels)
    
    # Note: If filtering for "outliers excl.", we need to adjust row heights or ensure
    # the target row exists. In Step 1, "outliers excl." is Row 7 relative to data start.
    
    for r in range(2, last_row + 1):
        cell_val = new_ws.cell(row=r, column=col_o_idx).value
        val_str = str(cell_val).strip().lower() if cell_val else ""
        
        # We assume the header row (1) is always shown.
        # Hide any row that is NOT the header AND does NOT match our filter.
        if r > 1 and val_str != fixed_filter:
            new_ws.row_dimensions[r].hidden = True
            
    # Ensure AutoFilter arrows are removed
    #new_ws.auto_filter.ref = None 

    # --- 4. Conditional Formatting (Red Stdev Highlights) ---
    # --- Configuration from Settings ---
    stdev_is_enabled = settings.get_setting("STDEV_THRESHOLD_ENABLED")
    
    # If disabled, set the variable to None so it bypasses conditional formatting
    if stdev_is_enabled:
        stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    else:
        stdev_threshold = None
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Column Q (17) and T (20)
    COL_Q = 17
    COL_T = 20
    
    # We must apply formatting to the visible rows. Since we can't easily know *which* # specific rows are visible without iterating, we'll apply it generally to the whole column range
    # or iterate. Iterating is safer to avoid huge rules.
    
    # Note: If filtering 'outliers excl.', the row index changes from Step 1 logic.
    # We apply to ALL rows, but since hidden ones aren't seen, it's fine.
    
    if stdev_threshold is not None:
        # Applying a rule to the entire column range (e.g. Q2:Q1000) is more efficient 
        # than adding 100 separate rules.
        
        range_q = f"{get_column_letter(COL_Q)}2:{get_column_letter(COL_Q)}{last_row}"
        range_t = f"{get_column_letter(COL_T)}2:{get_column_letter(COL_T)}{last_row}"
        
        new_ws.conditional_formatting.add(
            range_q,
            CellIsRule(operator="greaterThan", formula=[str(stdev_threshold)], fill=fill_error)
        )
        new_ws.conditional_formatting.add(
            range_t,
            CellIsRule(operator="greaterThan", formula=[str(stdev_threshold)], fill=fill_error)
        )

    # --- 5. Finalize View ---
    for s in wb.worksheets:
        s.sheet_view.tabSelected = False
    new_ws.sheet_view.tabSelected = True
    wb.active = wb.index(new_ws)
    new_ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Add Settings Popup Comment
    embed_settings_popup(new_ws, "Y1")

    # Set column widths
    new_ws.column_dimensions["O"].width = 16 
    
    wb.save(file_path)
    print(f"✅ Step 3: '{new_sheet_name}' created (Filter: {fixed_filter}).")