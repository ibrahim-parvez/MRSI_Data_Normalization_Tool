from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font 
from openpyxl.formatting.rule import CellIsRule 
from copy import copy
import utils.settings as settings 
from utils.common_utils import embed_settings_popup

def step2_tosort_water(file_path: str, filter_choice: str = "Last 6"):
    """
    Step 2: To Sort
    - Creates 'To Sort_DNT' as a copy of 'Data_DNT'.
    - Hides rows in 'To Sort_DNT' based on the filter_choice applied to Column O.
    - RE-ADDED: Keeps filter dropdown arrows on Row 1.
    - Applies Conditional Formatting (Red Stdev) ONLY if filter_choice is 'Last 6'.
    """
    
    source_sheet_name = "Data_DNT"
    new_sheet_name = "To Sort_DNT"
    
    # Clean the filter choice
    filter_choice = filter_choice.strip()
    
    wb = load_workbook(file_path)

    if source_sheet_name not in wb.sheetnames:
        raise ValueError(f"Source sheet '{source_sheet_name}' not found.")

    # --- 1. CLEANUP: Delete existing 'To Sort_DNT' ---
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    # --- 2. CREATE SHEET ---
    source_ws = wb[source_sheet_name]
    idx = wb.index(source_ws)
    new_ws = wb.create_sheet(new_sheet_name, idx)

    # --- 3. COPY DATA ---
    for row in source_ws.iter_rows():
        for cell in row:
            val = cell.value
            if cell.data_type == 'f' and hasattr(cell, 'formula'):
                val = cell.formula
            
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=val)
            
            if cell.has_style:
                try:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                except: pass

    # Copy Dimensions
    for r_idx, r_dim in source_ws.row_dimensions.items():
        new_ws.row_dimensions[r_idx].height = r_dim.height
    for c_let, c_dim in source_ws.column_dimensions.items():
        new_ws.column_dimensions[c_let].width = c_dim.width

    # --- 4. APPLY FILTER LOGIC (HIDE ROWS & KEEP ARROWS) ---
    col_o_idx = 15 # Column O
    last_row = new_ws.max_row
    last_col_letter = get_column_letter(new_ws.max_column)
    filter_choice_lower = filter_choice.lower()

    # Define the filter range (e.g., A1:X100)
    new_ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # If filter is NOT 'all', hide rows and set the filter criteria
    if filter_choice_lower != 'all':
        # Add the filter icon state to Column O (0-indexed for add_filter_column, so O is 14)
        new_ws.auto_filter.add_filter_column(14, [filter_choice])
        
        for r in range(2, last_row + 1):
            cell_val = new_ws.cell(row=r, column=col_o_idx).value
            cell_val_str = str(cell_val).strip().lower() if cell_val else ""
            
            if cell_val_str != filter_choice_lower:
                new_ws.row_dimensions[r].hidden = True
    
    # --- 5. CONDITIONAL FORMATTING (Only for "Last 6") ---
    if filter_choice_lower == "last 6":
        stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
        fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        COL_Q = 17
        COL_T = 20
        START_ROW_FOR_CF = 8 
        ROW_STEP = 12

        if stdev_threshold is not None:
            for r_idx in range(START_ROW_FOR_CF, last_row + 1, ROW_STEP):
                q_ref = f"{get_column_letter(COL_Q)}{r_idx}"
                new_ws.conditional_formatting.add(
                    q_ref,
                    CellIsRule(operator="greaterThan", formula=[str(stdev_threshold)], fill=fill_error)
                )

                t_ref = f"{get_column_letter(COL_T)}{r_idx}"
                new_ws.conditional_formatting.add(
                    t_ref,
                    CellIsRule(operator="greaterThan", formula=[str(stdev_threshold)], fill=fill_error)
                )

    # --- 6. FINALIZE ---
    for s in wb.worksheets:
        s.sheet_view.tabSelected = False
    new_ws.sheet_view.tabSelected = True
    wb.active = wb.index(new_ws)
    new_ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Add Settings Popup Comment
    embed_settings_popup(new_ws, "Y1")

    # Set column widths
    new_ws.column_dimensions["O"].width = 16 
    
    wb.save(file_path)
    print(f"Step 2: To Sort completed on {file_path} using filter '{filter_choice}'")