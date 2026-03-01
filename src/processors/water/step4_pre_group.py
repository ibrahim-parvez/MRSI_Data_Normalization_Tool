import pandas as pd
import string
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border
from copy import copy
import numpy as np
import re
from openpyxl.formatting.rule import CellIsRule 
import utils.settings as settings 
from utils.common_utils import embed_settings_popup

# Define the error fill for Conditional Formatting
fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def step4_pre_group_water(file_path: str):
    """
    Fixed Step 4: Linear Grouping by Consecutive Identifier
    - Row 2 is left blank (Data starts at Row 3).
    - Sheet placed to the left of "Last 6".
    - Scans 'Last 6' linearly.
    - Groups consecutive rows with the same base identifier.
    - Inserts Summary Block (Avg/Stdev/Count) when identifier changes.
    - Applies Conditional Formatting based on settings.
    - Preserves column widths and styles.
    """

    new_sheet_name = "Pre-Group_DNT"
    source_sheet_name = "Last 6_DNT"
    
    # Check for the required setting
    stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    if stdev_threshold is None:
        print("⚠️ Warning: 'stdev_threshold' not found in settings. Conditional formatting will be skipped.")

    # Define the number format for 3 decimal places
    THREE_DECIMAL_FORMAT = "0.000"

    # --- 1. Load Workbook ---
    wb_values = load_workbook(file_path, data_only=True)
    wb = load_workbook(file_path)

    # Find the source sheet (case-insensitive match)
    matched_source = next((s for s in wb.sheetnames if s.lower() == source_sheet_name.lower()), None)
    if matched_source is None:
        print(f"❌ Source sheet matching '{source_sheet_name}' not found.")
        return

    source_ws_vals = wb_values[matched_source]
    source_ws = wb[matched_source]

    # Clean up existing destination sheet
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    # Create new sheet to the left of source
    # wb.index(sheet) returns the index. Inserting AT that index pushes the existing sheet to the right.
    idx = wb.index(source_ws)
    new_ws = wb.create_sheet(new_sheet_name, idx)

    # --- 2. Configuration ---
    # Specific columns to read from Source
    source_cols_to_read = [1, 2, 3, 4, 5, 6, 7, 10, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    
    # Headers for the New Sheet
    new_headers = [
        'Line', 'Time Code', 'Identifier 1', 'Comment', 'Identifier 2', 'Analysis', 'Preparation',
        'Ampl 44', '', '', 'C avg', 'C stdev', '', 'O avg', 'O stdev', '', 'Sum area all'
    ]

    # Style Definitions
    header_fill = PatternFill(start_color="8ED973", end_color="8ED973", fill_type="solid")
    bold_font = Font(bold=True)

    # --- 3. Setup Headers & Column Widths ---
    # Write Headers to Row 1
    for col_idx, header in enumerate(new_headers, start=1):
        cell = new_ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font()
        cell.fill = header_fill

    new_ws.freeze_panes = "A2"

    # Copy column widths
    for t_col_idx, src_col_idx in enumerate(source_cols_to_read, start=1):
        src_letter = get_column_letter(src_col_idx)
        tgt_letter = get_column_letter(t_col_idx)
        try:
            width = source_ws.column_dimensions[src_letter].width
            if width is not None:
                new_ws.column_dimensions[tgt_letter].width = width
        except Exception:
            pass

    # --- 4. Helper Functions ---
    def get_base_key(identifier):
        if identifier is None:
            return None
        s_id = str(identifier)
        # Remove trailing ' r1', ' r1.1', etc. to find the base group name
        return re.sub(r'\s+[rR]\d+(?:\.\d+)*(?:[a-zA-Z]*)?$', '', s_id).strip()

    def copy_cell_style(src_cell, tgt_cell):
        if src_cell is None or tgt_cell is None: return
        if src_cell.has_style:
            try:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.number_format = src_cell.number_format 
                tgt_cell.alignment = copy(src_cell.alignment)
                tgt_cell.protection = copy(src_cell.protection)
            except: pass

    # --- 5. Linear Processing Loop ---
    
    src_row = 2
    
    # CHANGED: Start writing at Row 3 to leave Row 2 blank
    dest_row = 3
    
    max_row = source_ws.max_row
    
    # Track the start of the current group in the DESTINATION sheet
    group_start_dest_row = dest_row 

    while src_row <= max_row:
        # Check if source row is hidden; if so, skip logic but keep loops moving
        if source_ws.row_dimensions[src_row].hidden:
            src_row += 1
            continue

        # A. Read Current Row Data
        # We need Identifier 1 (Index 2 in our list, which corresponds to Source Col 3)
        id_val = source_ws_vals.cell(row=src_row, column=3).value
        current_base = get_base_key(id_val)

        # B. Copy Row to Destination
        for col_offset, src_col_idx in enumerate(source_cols_to_read, start=1):
            # Value from values-wb
            val = source_ws_vals.cell(row=src_row, column=src_col_idx).value
            tgt_cell = new_ws.cell(row=dest_row, column=col_offset, value=val)
            
            # Style from style-wb
            src_cell = source_ws.cell(row=src_row, column=src_col_idx)
            copy_cell_style(src_cell, tgt_cell)

        # Copy Row Height
        try:
            rh = source_ws.row_dimensions[src_row].height
            if rh is not None:
                new_ws.row_dimensions[dest_row].height = rh
        except: pass

        # C. Look Ahead Logic
        next_src_row = src_row + 1
        next_base = None
        
        # Find next visible row to compare
        while next_src_row <= max_row:
            if not source_ws.row_dimensions[next_src_row].hidden:
                next_val = source_ws_vals.cell(row=next_src_row, column=3).value
                next_base = get_base_key(next_val)
                break
            next_src_row += 1

        # D. Check if Group Ends (Difference in base key OR end of file)
        if current_base != next_base:
            # Group has finished at 'dest_row'.
            group_end_dest_row = dest_row
            
            # --- Insert Calculation Block ---
            calc_row_mid = dest_row + 2 # Leave 1 blank row (dest_row + 1), then calcs (dest_row + 2)
            
            # Marker for Step 5
            new_ws.cell(row=calc_row_mid, column=10, value="--").font = bold_font
            
            # Write Labels at Row N+1 (The blank row immediately following data)
            label_row = dest_row + 1
            new_ws.cell(row=label_row, column=11, value="Average").font = bold_font
            new_ws.cell(row=label_row, column=12, value="Stdev").font = bold_font
            new_ws.cell(row=label_row, column=13, value="Count").font = bold_font
            new_ws.cell(row=label_row, column=14, value="Average").font = bold_font
            new_ws.cell(row=label_row, column=15, value="Stdev").font = bold_font
            new_ws.cell(row=label_row, column=16, value="Count").font = bold_font

            # Formulas at Row N+2
            col_K = get_column_letter(11)
            col_L = get_column_letter(12) # C stdev
            col_N = get_column_letter(14)
            col_O = get_column_letter(15) # O stdev
            col_Q = get_column_letter(17)

            formulas = {
                11: f"=AVERAGE({col_K}{group_start_dest_row}:{col_K}{group_end_dest_row})",
                12: f"=STDEV({col_K}{group_start_dest_row}:{col_K}{group_end_dest_row})",
                13: f"=COUNT({col_K}{group_start_dest_row}:{col_K}{group_end_dest_row})",
                14: f"=AVERAGE({col_N}{group_start_dest_row}:{col_N}{group_end_dest_row})",
                15: f"=STDEV({col_N}{group_start_dest_row}:{col_N}{group_end_dest_row})",
                16: f"=COUNT({col_N}{group_start_dest_row}:{col_N}{group_end_dest_row})",
                17: f"=AVERAGE({col_Q}{group_start_dest_row}:{col_Q}{group_end_dest_row})"
            }

            for c, f in formulas.items():
                cell = new_ws.cell(row=calc_row_mid, column=c, value=f)
                cell.font = bold_font
                if c in [11, 12, 14, 15, 17]:
                    cell.number_format = THREE_DECIMAL_FORMAT

            # --- Conditional Formatting ---
            if stdev_threshold is not None:
                # 1. CF on Data Rows
                data_rng_L = f"{col_L}{group_start_dest_row}:{col_L}{group_end_dest_row}"
                data_rng_O = f"{col_O}{group_start_dest_row}:{col_O}{group_end_dest_row}"
                thresh_str = str(stdev_threshold)

                new_ws.conditional_formatting.add(data_rng_L, CellIsRule(operator="greaterThan", formula=[thresh_str], fill=fill_error))
                new_ws.conditional_formatting.add(data_rng_O, CellIsRule(operator="greaterThan", formula=[thresh_str], fill=fill_error))

                # 2. CF on Formula Rows
                new_ws.conditional_formatting.add(f"{col_L}{calc_row_mid}", CellIsRule(operator="greaterThan", formula=[thresh_str], fill=fill_error))
                new_ws.conditional_formatting.add(f"{col_O}{calc_row_mid}", CellIsRule(operator="greaterThan", formula=[thresh_str], fill=fill_error))

            # Update dest_row for the next group
            # Layout: Data (ends at dest_row) -> Label Row -> Calc Row -> Blank Row -> Next Data
            # Next Data starts at Calc Row + 2
            dest_row = calc_row_mid + 2 
            group_start_dest_row = dest_row
        else:
            # Same group, just advance row
            dest_row += 1

        src_row += 1

    # --- Finalize ---
    for s in wb.worksheets: s.sheet_view.tabSelected = False
    new_ws.sheet_view.tabSelected = True
    wb.active = wb.index(new_ws)
    new_ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Add Settings Popup Comment
    embed_settings_popup(new_ws, "R1")

    # Set column widths
    new_ws.column_dimensions["J"].width = 16 

    wb.save(file_path)
    print(f"✅ Step 4: Pre-Grouping sheet '{new_sheet_name}' created.")