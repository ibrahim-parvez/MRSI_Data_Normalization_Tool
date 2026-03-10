import pandas as pd
import statistics
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, NamedStyle, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.views import Selection
import utils.settings as settings
from utils.common_utils import embed_settings_popup

def step1_data_water(file_path, sheet_name='ExportGB2.wke', sparkline=False):
    """
    Step 1: DATA for Water
    Reads the Excel file, transforms it, adds calculation rectangle and styling.
    
    UPDATES:
    1. Replaced Sparkline with "Outliers Excl."
    2. Formulas are built dynamically using explicit cell references (e.g. =AVERAGE(L9,L10,L12)).
    3. Conditional Formatting visualizes exactly which cells are excluded.
    4. DUPLICATES 'Sum Area' and 'Flags' to the "Outliers Excl." row so they persist after filtering.
    """
    new_sheet_name = 'Data_DNT'
    
    # --- Configuration from Settings ---
    stdev_is_enabled = settings.get_setting("STDEV_THRESHOLD_ENABLED")
    
    # If disabled, set the variable to None so it bypasses conditional formatting
    if stdev_is_enabled:
        stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    else:
        stdev_threshold = None
    sigma_mult = settings.get_setting("OUTLIER_SIGMA") or 2
    exclusion_mode = settings.get_setting("OUTLIER_EXCLUSION_MODE") or "Individual"
    
    # Read original data
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
    
    # Headers
    headers = [
        'Line', 'Time Code', 'Identifier 1', 'Comment', 'Identifier 2', 'Analysis',
        'Preparation', 'Peak Nr', 'Rt', 'Ampl 44', 'Area All',
        'd 13C/12C', 'd 18O/16O'
    ]
    
    # Calculation rectangle headers
    calc_headers = [
        'C avg', 'C stdev', '', 'O avg', 'O stdev', '',
        'Sum area all', 'funny peaks', 'min intensity'
    ]
    
    # Load workbook and reset sheet
    wb = load_workbook(file_path)
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]
    
    first_index = wb.index(wb[sheet_name])
    ws = wb.create_sheet(new_sheet_name, first_index)
    
    # View settings
    for s in wb.worksheets:
        try: s.sheet_view.tabSelected = False
        except: pass
    ws.sheet_view.tabSelected = True
    wb.active = wb.index(ws)
    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    
    # Styles
    header_fill_green = PatternFill(start_color="8ed973", end_color="8ed973", fill_type="solid")
    header_fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    calc_fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    flag_fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Font for Conditional Formatting (Red + Strikethrough)
    font_outlier = Font(color="FF0000", strike=True)
    
    font_bold = Font(bold=True)
    NUM_FORMAT_3_DECIMAL = '0.000'
    NUM_FORMAT_2_DECIMAL = '0.00'

    # --- Write Headers ---
    label_col = 15 # O
    start_col_calc = label_col
    total_cols = ws.max_column + 1000
    yellow_cols = [8, 9, 11, 12, 13, 23, 24]
    
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill_yellow if col in yellow_cols else header_fill_green
    
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    
    for i, h in enumerate(calc_headers, start=start_col_calc + 1):
        ws.cell(row=1, column=i, value=h)
        
    cur_row = 3
    col_map = {h: i + 1 for i, h in enumerate(headers)}
    
    # Normalize headers
    def normalize_name(s): return ' '.join(str(s).split()).lower() if s else ''
    df_cols_norm = {normalize_name(c): c for c in df.columns}
    header_to_dfcol = {h: df_cols_norm.get(normalize_name(h)) for h in headers}
    
    # Column Letters
    col_area = col_map.get('Area All')
    col_c = col_map.get('d 13C/12C') # L
    col_o = col_map.get('d 18O/16O') # M
    col_ampl = col_map.get('Ampl 44')
    
    col_letter_area = get_column_letter(col_area) if col_area else None
    col_letter_c = get_column_letter(col_c) if col_c else None
    col_letter_o = get_column_letter(col_o) if col_o else None
    col_letter_ampl = get_column_letter(col_ampl) if col_ampl else None
    
    # Group Processing
    grouped = df.groupby('Line', sort=False)
    
    calc_start_col = start_col_calc
    calc_end_col = calc_start_col + len(calc_headers)
    
    def write_transparent_separator(sheet, row_idx, last_col):
        for c in range(1, last_col + 1):
            sheet.cell(row=row_idx, column=c, value=None)

    for line, group in grouped:
        first_data_row = cur_row
        num_data_rows = len(group)
        expected_rows = 11
        num_rows_to_write = max(expected_rows, num_data_rows)
        last_data_row = first_data_row + num_rows_to_write - 1
        
        # --- WRITE DATA ROWS ---
        padded_rows = [] # Store row data for analysis
        
        for i in range(num_rows_to_write):
            row_idx = cur_row
            if i < num_data_rows:
                row_dict = group.iloc[i].to_dict()
                padded_rows.append(row_dict)
            else:
                row_dict = {c: None for c in df.columns}
                for key in ['Line', 'Time Code', 'Identifier 1']:
                    row_dict[key] = group.iloc[0][key]
                padded_rows.append(row_dict)
            
            for h in headers:
                col = col_map[h]
                val = row_dict.get(header_to_dfcol[h]) if header_to_dfcol[h] else None
                ws.cell(row=row_idx, column=col, value=val)

            for c in range(calc_start_col, calc_end_col + 1):
                ws.cell(row=row_idx, column=c).fill = calc_fill
                
            cur_row += 1
            
        # --- CALCULATION BLOCK FORMULAS ---
        r1 = first_data_row
        r2 = first_data_row + 1
        r3 = first_data_row + 2
        r4 = first_data_row + 3
        r5 = first_data_row + 4
        r6 = first_data_row + 5 # Last 6
        r7 = first_data_row + 6 # Outliers Excl
        r8 = first_data_row + 7
        r9 = first_data_row + 8
        r10 = first_data_row + 9
        r11 = first_data_row + 10
        
        ref_rows = [first_data_row, first_data_row + 1, first_data_row + 3]
        first_after_ref = first_data_row + 4
        last6_start = max(first_data_row, last_data_row - 5)
        
        # Ranges for Formulas
        range_c_last6 = f"{col_letter_c}{last6_start}:{col_letter_c}{last_data_row}"
        range_o_last6 = f"{col_letter_o}{last6_start}:{col_letter_o}{last_data_row}"
        
        # Col O Labels
        ws.cell(row=r1, column=calc_start_col, value="Calculations").font = font_bold
        ws.cell(row=r3, column=calc_start_col, value="ref avg")
        ws.cell(row=r5, column=calc_start_col, value="all")
        ws.cell(row=r6, column=calc_start_col, value="last 6")
        ws.cell(row=r7, column=calc_start_col, value="last 6 outliers excl.")
        ws.cell(row=r8, column=calc_start_col, value="Amp 44")
        ws.cell(row=r9, column=calc_start_col, value="start 6")
        ws.cell(row=r10, column=calc_start_col, value="end 11")
        ws.cell(row=r11, column=calc_start_col, value="delta")
        
        # --- Column P (Carbon Stats) ---
        col_P = calc_start_col + 1
        col_Q = calc_start_col + 2 
        
        # Absolute references for Conditional Formatting logic
        c_mean_cell_abs  = f"${get_column_letter(col_P)}${r6}"
        c_stdev_cell_abs = f"${get_column_letter(col_Q)}${r6}"
        o_mean_cell_abs  = f"${get_column_letter(calc_start_col + 4)}${r6}"
        o_stdev_cell_abs = f"${get_column_letter(calc_start_col + 5)}${r6}"

        if col_letter_c:
            ws.cell(row=r3, column=col_P, value=f"=AVERAGE({col_letter_c}{ref_rows[0]},{col_letter_c}{ref_rows[1]},{col_letter_c}{ref_rows[2]})").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r5, column=col_P, value=f"=AVERAGE({col_letter_c}{first_after_ref}:{col_letter_c}{last_data_row})").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r6, column=col_P, value=f"=AVERAGE({range_c_last6})").number_format = NUM_FORMAT_3_DECIMAL
            
            ws.cell(row=r3, column=col_Q, value=f"=STDEV({col_letter_c}{ref_rows[0]},{col_letter_c}{ref_rows[1]},{col_letter_c}{ref_rows[2]})").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r5, column=col_Q, value=f"=STDEV({col_letter_c}{first_after_ref}:{col_letter_c}{last_data_row})").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r6, column=col_Q, value=f"=STDEV({range_c_last6})").number_format = NUM_FORMAT_3_DECIMAL

            ws.cell(row=r9, column=col_P, value=f"={col_letter_c}{last6_start}").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r10, column=col_P, value=f"={col_letter_c}{last_data_row}").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r11, column=col_P, value=f"={get_column_letter(col_P)}{r10}-{get_column_letter(col_P)}{r9}").number_format = NUM_FORMAT_3_DECIMAL

        # --- Flag Messages ---
        flag_message = ""
        # Fix logic: water usually uses fixed headers logic, but here we count actual rows
        expected_rows = 11
        if num_data_rows < expected_rows: flag_message = f"<{expected_rows} ({num_data_rows})"
        elif num_data_rows > expected_rows: flag_message = f">{expected_rows} ({num_data_rows})"

        col_R = calc_start_col + 3
        if flag_message:
            ws.cell(row=r6, column=col_R, value=flag_message).fill = flag_fill_red
            # DUPLICATE FLAG TO OUTLIER ROW (r7)
            ws.cell(row=r7, column=col_R, value=flag_message).fill = flag_fill_red
        
        # --- Column S (Oxygen Stats) ---
        col_S = calc_start_col + 4
        col_T = calc_start_col + 5
        
        if col_letter_o:
            ws.cell(row=r3, column=col_S, value=f"=AVERAGE({col_letter_o}{ref_rows[0]},{col_letter_o}{ref_rows[1]},{col_letter_o}{ref_rows[2]})").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r5, column=col_S, value=f"=AVERAGE({col_letter_o}{first_after_ref}:{col_letter_o}{last_data_row})").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r6, column=col_S, value=f"=AVERAGE({range_o_last6})").number_format = NUM_FORMAT_3_DECIMAL
            
            ws.cell(row=r3, column=col_T, value=f"=STDEV({col_letter_o}{ref_rows[0]},{col_letter_o}{ref_rows[1]},{col_letter_o}{ref_rows[2]})").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r5, column=col_T, value=f"=STDEV({col_letter_o}{first_after_ref}:{col_letter_o}{last_data_row})").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r6, column=col_T, value=f"=STDEV({range_o_last6})").number_format = NUM_FORMAT_3_DECIMAL

            ws.cell(row=r9, column=col_S, value=f"={col_letter_o}{last6_start}").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r10, column=col_S, value=f"={col_letter_o}{last_data_row}").number_format = NUM_FORMAT_3_DECIMAL
            ws.cell(row=r11, column=col_S, value=f"={get_column_letter(col_S)}{r10}-{get_column_letter(col_S)}{r9}").number_format = NUM_FORMAT_3_DECIMAL

        col_U = calc_start_col + 6
        if flag_message:
            ws.cell(row=r6, column=col_U, value=flag_message).fill = flag_fill_red
            # DUPLICATE FLAG TO OUTLIER ROW (r7)
            ws.cell(row=r7, column=col_U, value=flag_message).fill = flag_fill_red

        # -------------------------------------------------------------------------
        # --- DYNAMIC FORMULA CONSTRUCTION FOR "OUTLIERS EXCL." (ROW 7) ---
        # -------------------------------------------------------------------------
        # 1. Gather Data for analysis
        c_values = []
        o_values = []
        row_indices = []
        
        # Wait, simple logic: absolute rows are last6_start to last_data_row
        for r_idx in range(last6_start, last_data_row + 1):
            # Map absolute row back to index in padded_rows
            rel_idx = r_idx - first_data_row
            row_data = padded_rows[rel_idx]
            
            try: vc = float(row_data.get(header_to_dfcol['d 13C/12C']))
            except: vc = None
            try: vo = float(row_data.get(header_to_dfcol['d 18O/16O']))
            except: vo = None
            
            c_values.append(vc)
            o_values.append(vo)
            row_indices.append(r_idx)

        # 2. Calculate Stats for Bounds
        valid_c = [x for x in c_values if x is not None]
        valid_o = [x for x in o_values if x is not None]
        
        mean_c = statistics.mean(valid_c) if len(valid_c) > 0 else 0
        std_c = statistics.stdev(valid_c) if len(valid_c) > 1 else 0
        mean_o = statistics.mean(valid_o) if len(valid_o) > 0 else 0
        std_o = statistics.stdev(valid_o) if len(valid_o) > 1 else 0
        
        c_up = mean_c + (sigma_mult * std_c)
        c_low = mean_c - (sigma_mult * std_c)
        o_up = mean_o + (sigma_mult * std_o)
        o_low = mean_o - (sigma_mult * std_o)
        
        # 3. Build List of Valid Cells
        valid_c_cells = []
        valid_o_cells = []
        
        for idx, r_num in enumerate(row_indices):
            vc = c_values[idx]
            vo = o_values[idx]
            
            is_c_out = False
            is_o_out = False
            
            if vc is not None:
                if vc > c_up or vc < c_low: is_c_out = True
            if vo is not None:
                if vo > o_up or vo < o_low: is_o_out = True
                
            if exclusion_mode == "Exclude Row":
                # If EITHER is bad, EXCLUDE BOTH
                if not is_c_out and not is_o_out:
                    if vc is not None: valid_c_cells.append(f"{col_letter_c}{r_num}")
                    if vo is not None: valid_o_cells.append(f"{col_letter_o}{r_num}")
            else:
                # Individual Logic
                if not is_c_out and vc is not None: valid_c_cells.append(f"{col_letter_c}{r_num}")
                if not is_o_out and vo is not None: valid_o_cells.append(f"{col_letter_o}{r_num}")

        # 4. Write Explicit Formulas (e.g. =AVERAGE(L20,L21,L23))
        if valid_c_cells:
            c_range_str = ",".join(valid_c_cells)
            f_c_avg = f"=AVERAGE({c_range_str})"
            f_c_std = f"=STDEV({c_range_str})" if len(valid_c_cells) > 1 else "0"
        else:
            f_c_avg, f_c_std = "No Data", "0"
            
        if valid_o_cells:
            o_range_str = ",".join(valid_o_cells)
            f_o_avg = f"=AVERAGE({o_range_str})"
            f_o_std = f"=STDEV({o_range_str})" if len(valid_o_cells) > 1 else "0"
        else:
            f_o_avg, f_o_std = "No Data", "0"

        ws.cell(row=r7, column=col_P, value=f_c_avg).number_format = NUM_FORMAT_3_DECIMAL
        ws.cell(row=r7, column=col_Q, value=f_c_std).number_format = NUM_FORMAT_3_DECIMAL
        ws.cell(row=r7, column=col_S, value=f_o_avg).number_format = NUM_FORMAT_3_DECIMAL
        ws.cell(row=r7, column=col_T, value=f_o_std).number_format = NUM_FORMAT_3_DECIMAL

        # -------------------------------------------------------------------------
        # --- CONDITIONAL FORMATTING (Red Strikethrough) ---
        # -------------------------------------------------------------------------
        
        # Logic: OR(Value > Mean+Sigma*Stdev, Value < Mean-Sigma*Stdev)
        # Using ABSOLUTE references to the 'Last 6' stats cells in row r6
        
        c_logic = f"OR({col_letter_c}{last6_start}>({c_mean_cell_abs}+({sigma_mult}*{c_stdev_cell_abs})), {col_letter_c}{last6_start}<({c_mean_cell_abs}-({sigma_mult}*{c_stdev_cell_abs})))"
        o_logic = f"OR({col_letter_o}{last6_start}>({o_mean_cell_abs}+({sigma_mult}*{o_stdev_cell_abs})), {col_letter_o}{last6_start}<({o_mean_cell_abs}-({sigma_mult}*{o_stdev_cell_abs})))"

        if exclusion_mode == "Exclude Row":
            final_c_logic = f"OR({c_logic}, {o_logic})"
            final_o_logic = f"OR({c_logic}, {o_logic})"
        else:
            final_c_logic = c_logic
            final_o_logic = o_logic

        # Apply Rules
        if col_letter_c:
            ws.conditional_formatting.add(range_c_last6, FormulaRule(formula=[final_c_logic], font=font_outlier))
        if col_letter_o:
            ws.conditional_formatting.add(range_o_last6, FormulaRule(formula=[final_o_logic], font=font_outlier))

        # Col V (Sum area)
        col_V = calc_start_col + 7
        if col_letter_area:
            ws.cell(row=r3, column=col_V, value=f"=SUM({col_letter_area}{first_data_row}:{col_letter_area}{first_data_row+3})").number_format = NUM_FORMAT_2_DECIMAL
            ws.cell(row=r5, column=col_V, value=f"=SUM({col_letter_area}{first_after_ref}:{col_letter_area}{last_data_row})").number_format = NUM_FORMAT_2_DECIMAL
            
            # Row 6: Last 6 Sum
            ws.cell(row=r6, column=col_V, value=f"=SUM({col_letter_area}{last6_start}:{col_letter_area}{last_data_row})").number_format = NUM_FORMAT_2_DECIMAL
            
            # Row 7: DUPLICATE SUM (so it persists after filtering)
            ws.cell(row=r7, column=col_V, value=f"=SUM({col_letter_area}{last6_start}:{col_letter_area}{last_data_row})").number_format = NUM_FORMAT_2_DECIMAL
            
        # Col W (Funny peaks)
        col_W = calc_start_col + 8
        if col_letter_ampl:
            for rr in (r1, r2, r3, r4): ws.cell(row=rr, column=col_W, value="ref")
            for rr in range(first_after_ref, last_data_row + 1):
                ws.cell(row=rr, column=col_W, value=f"=IF({col_letter_ampl}{rr}>{col_letter_ampl}{rr+1},IF({col_letter_ampl}{rr+1}<{col_letter_ampl}{rr},\"ok\",\"check\"),\"check\")")
        
        # Col X (Min Intensity)
        col_X = calc_start_col + 9
        if col_letter_ampl:
            ws.cell(row=r1, column=col_X, value="if 44<1000")
            for rr in range(first_after_ref, last_data_row + 1):
                ws.cell(row=rr, column=col_X, value=f"=IF({col_letter_ampl}{rr}<1000,\"check\",\"ok\")")

        # CF for Stdev Threshold (Last 6 row)
        if stdev_threshold:
            for col_letter in [get_column_letter(col_Q), get_column_letter(col_T)]:
                ws.conditional_formatting.add(f"{col_letter}{r6}", CellIsRule(operator="greaterThan", formula=[str(stdev_threshold)], fill=fill_error))

        write_transparent_separator(ws, cur_row, calc_end_col)
        cur_row += 1
    
    # Add Settings Popup Comment
    embed_settings_popup(ws, "Y1")
    
    # Set column widths
    ws.column_dimensions["O"].width = 16 
        
    wb.save(file_path)
    print(f"✅ Step 1: DATA completed on {file_path}")