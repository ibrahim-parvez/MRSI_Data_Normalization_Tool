import os
import re
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from utils.common_utils import embed_settings_popup

# --- Helper: Force Excel to Calculate Formulas ---
def _try_refresh_with_xlwings(path):
    """
    Opens the workbook in the background using Excel, 
    calculates all formulas, and saves it. 
    """
    try:
        import xlwings as xw
    except ImportError:
        return False

    app = None
    book = None
    try:
        app = xw.App(visible=False, add_book=False)
        book = app.books.open(os.path.abspath(path))
        book.app.api.Application.CalculateFull()
        book.save()
        book.close()
        app.quit()
        return True
    except Exception as e:
        print(f"⚠️ xlwings refresh failed: {e}")
        try:
            if book: book.close()
            if app: app.quit()
        except: pass
        return False

def step7_report_water(file_path):
    source_sheet = "Normalization_DNT"
    new_sheet_name = "Report_DNT"

    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return

    # 1. Force Calculation Redundant Step
    #print("⏳ Recalculating formulas with Excel...")
    #_try_refresh_with_xlwings(file_path)

    # 2. Load Workbooks
    print("Loading workbook (this may take a moment)...")
    wb_fmt = load_workbook(file_path, data_only=False)
    wb_val = load_workbook(file_path, data_only=True)

    if source_sheet not in wb_fmt.sheetnames:
        print(f"❌ Sheet '{source_sheet}' not found.")
        return

    ws_fmt = wb_fmt[source_sheet]
    ws_val = wb_val[source_sheet]

    # --- 3. Determine Header Location ---
    # We look for the "main" header row (usually contains "δ¹⁸O SMOW" or via freeze pane)
    header_main_row = None
    
    # Strategy A: Freeze Pane (often points to the first data row, so headers are above)
    freeze_loc = ws_fmt.freeze_panes
    if freeze_loc:
        match = re.search(r"(\d+)", str(freeze_loc))
        if match:
            # Usually freeze is at the first data row. 
            # If freeze is row 22, headers are usually 20, 21.
            header_main_row = int(match.group(1)) - 1
    
    # Strategy B: Fallback Search
    if not header_main_row or header_main_row < 2:
        for r in range(1, 100):
            val_v = ws_val.cell(row=r, column=22).value 
            if val_v == "δ¹⁸O SMOW":
                header_main_row = r 
                break
    
    if not header_main_row: 
        header_main_row = 20 # Fallback default

    # The user wants the row ABOVE the main header + the main header
    header_rows = [header_main_row - 1, header_main_row]
    print(f"ℹ️ Headers identified at rows {header_rows}")

    # --- 4. Locate Data Start (Grey Row Separation) ---
    # We look for 2 consecutive rows with #C0C0C0 to denote the end of standards
    data_start_row = None
    consecutive_grey = 0
    
    # Start searching below headers
    search_start = header_rows[-1] + 1
    
    for r in range(search_start, ws_fmt.max_row + 1):
        # Check Column B (2) for the grey color. 
        # Standard grey is often "FFC0C0C0" or Theme index with tint.
        # We check simply if "C0C0C0" is in the hex string.
        cell = ws_fmt.cell(row=r, column=2)
        color_hex = str(cell.fill.start_color.rgb) if cell.fill.start_color else ""
        
        if "C0C0C0" in color_hex.upper():
            consecutive_grey += 1
        else:
            consecutive_grey = 0
        
        if consecutive_grey == 2:
            # Found the separator! The data starts immediately after these 2 rows.
            data_start_row = r + 1
            break
            
    if not data_start_row:
        print("⚠️ Could not find the grey separator rows (#C0C0C0). Defaulting to copying all rows after headers.")
        data_start_row = header_rows[-1] + 1
    else:
        print(f"ℹ️ Found grey separator. Data starts at row {data_start_row}")

    # --- 5. Prepare New Sheet ---
    if new_sheet_name in wb_fmt.sheetnames:
        del wb_fmt[new_sheet_name]

    ws_new = wb_fmt.create_sheet(new_sheet_name, index=wb_fmt.index(ws_fmt))

    # --- 6. Define Column Mapping (With Blank D) ---
    # Source: A(1), B(2), C(3) ... then S(19) onwards
    # Dest:   A(1), B(2), C(3), [D-BLANK], E(5), F(6)...
    
    cols_left = [1, 2, 3] # A, B, C
    cols_right = list(range(19, ws_fmt.max_column + 1)) # S onwards
    
    mapping = {}
    
    # Map Left Columns (Direct Copy)
    for c in cols_left:
        mapping[c] = c 
        
    # Map Right Columns (Shifted to start at E/5)
    dest_col_idx = 5
    for c in cols_right:
        mapping[c] = dest_col_idx
        dest_col_idx += 1

    # --- 7. Copy Execution ---
    current_dest_row = 1
    
    # Helper to copy a row
    # Helper to copy a row
    def copy_row(src_row_idx):
        nonlocal current_dest_row
        
        # Check if source row is empty (skip empty rows in data section)
        if src_row_idx >= data_start_row:
             # Just check the columns we care about
             is_empty = True
             for c_chk in (cols_left + cols_right):
                 if ws_val.cell(row=src_row_idx, column=c_chk).value is not None:
                     is_empty = False
                     break
             # If it's empty, tell the main loop we skipped it
             if is_empty: return False

        # Perform Copy
        for src_col, dst_col in mapping.items():
            src_cell_fmt = ws_fmt.cell(row=src_row_idx, column=src_col)
            src_cell_val = ws_val.cell(row=src_row_idx, column=src_col)
            dst_cell = ws_new.cell(row=current_dest_row, column=dst_col)

            # Copy Value
            val = src_cell_val.value
            if val is None: val = src_cell_fmt.value
            dst_cell.value = val

            # Copy Styles
            if src_cell_fmt.has_style:
                dst_cell.font = copy(src_cell_fmt.font)
                dst_cell.border = copy(src_cell_fmt.border)
                dst_cell.fill = copy(src_cell_fmt.fill)
                dst_cell.number_format = copy(src_cell_fmt.number_format)
                dst_cell.alignment = copy(src_cell_fmt.alignment)
                dst_cell.protection = copy(src_cell_fmt.protection)

        # Copy Row Height
        rd = ws_fmt.row_dimensions.get(src_row_idx)
        if rd and rd.height is not None:
            ws_new.row_dimensions[current_dest_row].height = rd.height
        
        current_dest_row += 1
        return True  # Tell the main loop we successfully copied data

    # A. Copy Headers
    for h_row in header_rows:
        copy_row(h_row)

    current_dest_row += 2

    # B. Copy Data (After Grey Rows)
    in_group = False

    for d_row in range(data_start_row, ws_fmt.max_row + 1):
        row_was_copied = copy_row(d_row)
        
        if row_was_copied:
            # We are actively reading a block of data
            in_group = True
        else:
            # We hit an empty row. Did we just finish a group?
            if in_group:
                current_dest_row += 2  # Add exactly 2 blank rows
                in_group = False       # Reset until we hit data again

    # --- 8. Final Formatting ---
    # Adjust Column Widths
    for src_col, dst_col in mapping.items():
        src_letter = get_column_letter(src_col)
        dst_letter = get_column_letter(dst_col)
        cd = ws_fmt.column_dimensions.get(src_letter)
        if cd and cd.width:
            ws_new.column_dimensions[dst_letter].width = cd.width
        else:
            ws_new.column_dimensions[dst_letter].width = 13
    
    # Set Blank Column D width to be narrow (optional aesthetic tweak)
    ws_new.column_dimensions['D'].width = 10

    # Freeze Panes: D3 (Rows 1-2 Frozen, Columns A-C Frozen)
    ws_new.freeze_panes = "D3"
    
    ws_new.sheet_view.tabSelected = True
    ws_fmt.sheet_view.tabSelected = False
    wb_fmt.active = wb_fmt.index(ws_new)
    ws_new.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Add Settings Popup Comment
    embed_settings_popup(ws_new, "A1")

    wb_fmt.save(file_path)
    print(f"✅ Step 7: Water Report completed on {file_path}")