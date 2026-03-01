from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from copy import copy
import re
from datetime import datetime
import utils.settings as settings
from utils.common_utils import embed_settings_popup

# --- Helper Functions for He/CO2 Logic ---

def extract_run_number(identifier):
    if not identifier:
        return None, None
    match = re.search(r'(?i)[rR](\d+)(?:\.(\d+))?', str(identifier))
    if match:
        major = int(match.group(1))
        minor = int(match.group(2)) if match.group(2) else 0
        return major, minor
    return None, None

def get_valid_heco2_indices(ws):
    valid_rows = set()
    seen_majors = {} 
    max_row = ws.max_row
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=3).value
        if not val:
            continue
        s_val = str(val).strip().lower()
        if "heco2" in s_val or "co2" in s_val:
            major, minor = extract_run_number(s_val)
            if major is None or major == 1: 
                continue
            if major not in seen_majors:
                seen_majors[major] = (minor, r)
            else:
                prev_minor, prev_r = seen_majors[major]
                if minor < prev_minor:
                    seen_majors[major] = (minor, r)
    for _, row_idx in seen_majors.values():
        valid_rows.add(row_idx)
    return valid_rows

def step6_normalization_water(file_path: str):
    wb = load_workbook(file_path)
    # Load again to get calculated values for logic
    wb_values = load_workbook(file_path, data_only=True)

    if "Group_DNT" not in wb.sheetnames:
        print("❌ Sheet 'Group_DNT' not found.")
        return 0

    group_ws = wb["Group_DNT"]
    group_values_ws = wb_values["Group_DNT"]

    if "Normalization_DNT" in wb.sheetnames:
        del wb["Normalization_DNT"]

    summary_ws = wb.create_sheet("Normalization_DNT", wb.index(group_ws))

    # --- 1. Load Settings & Determine Offsets ---
    stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    calc_mode = settings.get_setting("CALC_MODE_STEP7")
    
    use_outliers = (calc_mode == "Outliers Excluded")
    
    # Value Offset: 
    # Row 0: -- (Average Label)
    # Row 1: All Values
    # Row 2: (Blank)
    # Row 3: Outlier Excl (Average Label)
    # Row 4: Outlier Values
    calc_value_offset = 4 if use_outliers else 1
    
    print(f"ℹ️ Step 6 Mode: {calc_mode} (Value Offset: {calc_value_offset})")

    # --- 2. Styles Definition ---
    light_green_fill = PatternFill(start_color="DBF2D0", end_color="DBF2D0", fill_type="solid")
    box_blue_fill = PatternFill(start_color="DAE8F9", end_color="DAE8F9", fill_type="solid")
    box_peach_fill = PatternFill(start_color="FBE2D5", end_color="FBE2D5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    gray_box_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    heco2_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    heco2_box_fill = PatternFill(start_color="82CCEB", end_color="82CCEB", fill_type="solid")

    red_font = Font(color="FF0000")
    blue_font = Font(color="0000FF")
    orange_font = Font(color="E46C0A")
    green_font = Font(color="00B050")
    
    bold_red_font = Font(color="FF0000", bold=True)
    bold_blue_font = Font(color="0000FF", bold=True)
    bold_orange_font = Font(color="E46C0A", bold=True)
    bold_green_font = Font(color="00B050", bold=True)
    
    bold_blue_text_font = Font(color="0000FF", bold=True)
    bold_green_text_font = Font(color="00B050", bold=True)
    black_bold = Font(bold=True, color="000000")
    light_blue_bold = Font(bold=True, color="5DADE2")
    
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Border Styles
    thin_black_side = Side(style="thin", color="000000")
    medium_black_side = Side(style="medium", color="000000")
    double_black_side = Side(style="double", color="000000")
    
    FMT_3_DEC = "0.000"
    FMT_2_DEC = "0.00"

    # --- Dynamic Color & Font Mapping from Settings ---
    # Map text color names to Hex codes used by openpyxl
    color_hex_map = {
        "red": "FF0000", "blue": "0000FF", "darkblue": "00008B",
        "orange": "E46C0A", "green": "00B050", "lightblue": "5DADE2",
        "black": "000000"
    }

    water_stds_settings = settings.get_setting("REFERENCE_MATERIALS", "Water")
    slope_groups_settings = settings.get_setting("SLOPE_INTERCEPT_GROUPS", "Water")

    # Rebuild the maps dynamically
    std_color_map = {}
    std_bold_color_map = {}
    
    # Store Row locations for reference later
    std_row_index_map = {} 

    for std in water_stds_settings:
        name = std["col_c"]
        c_name = std.get("color", "black").lower()
        hex_code = color_hex_map.get(c_name, "000000")
        
        std_color_map[name] = Font(color=hex_code)
        std_bold_color_map[name] = Font(color=hex_code, bold=True)

    # --- Border Helper ---
    def apply_box_border(ws, start_row, start_col, end_row, end_col, fill):
        """Applies fill and a MEDIUM outer border to a cell range."""
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
        
        for c in range(start_col, end_col + 1):
            cell_top = ws.cell(row=start_row, column=c)
            b = copy(cell_top.border)
            cell_top.border = Border(top=medium_black_side, left=b.left, right=b.right, bottom=b.bottom)
            cell_bot = ws.cell(row=end_row, column=c)
            b2 = copy(cell_bot.border)
            cell_bot.border = Border(bottom=medium_black_side, left=b2.left, right=b2.right, top=b2.top)
            
        for r in range(start_row, end_row + 1):
            cell_l = ws.cell(row=r, column=start_col)
            b3 = copy(cell_l.border)
            cell_l.border = Border(left=medium_black_side, top=b3.top, right=b3.right, bottom=b3.bottom)
            cell_r = ws.cell(row=r, column=end_col)
            b4 = copy(cell_r.border)
            cell_r.border = Border(right=medium_black_side, top=b4.top, left=b4.left, bottom=b4.bottom)

    def apply_vertical_divider(ws, col_idx, start_row, end_row):
        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            cur = cell.border
            cell.border = Border(left=cur.left, right=double_black_side, top=cur.top, bottom=cur.bottom)

    valid_heco2_src_rows = get_valid_heco2_indices(group_values_ws)

    # --- Setup Sheet Layout (Dynamic) ---
    
    """
    # 1. Background Fill (Initial wipe) - Estimate height based on dynamic lists
    est_header_height = 15 # Safe buffer
    for r in range(1, est_header_height):
        for c in range(1, 51):
            summary_ws.cell(row=r, column=c).fill = light_green_fill
            
    summary_ws["A1"] = datetime.now().strftime("%Y-%m-%d")
    """
    # ==========================================
    # === BOX 1: Water Standards (Dynamic) ===
    # ==========================================
    std_start_row = 5
    std_end_row = std_start_row + len(water_stds_settings) - 1
    
    # Apply Blue Box Border
    apply_box_border(summary_ws, 2, 3, std_end_row, 15, fill=box_blue_fill)
    apply_vertical_divider(summary_ws, 4, 2, std_end_row) 
    
    # Header Line
    for c in range(3, 16):
        cell = summary_ws.cell(row=3, column=c)
        cur = cell.border
        cell.border = Border(left=cur.left, right=cur.right, top=cur.top, bottom=double_black_side)

    # Static Headers
    summary_ws["C1"] = "Normalization"; summary_ws["C1"].font = black_bold
    c2 = summary_ws["C2"]; c2.value = "EQ Time ="; c2.fill = yellow_fill; c2.font = black_bold
    summary_ws["C3"] = "Water Standards"; summary_ws["C3"].font = black_bold
    
    summary_ws["F2"] = "Published"; summary_ws["F2"].font = black_bold; summary_ws["F2"].alignment = center_align
    summary_ws["F3"] = "δ²H"; summary_ws["F3"].alignment = center_align; summary_ws["F3"].font = black_bold
    summary_ws["G3"] = "δ¹⁸O SMOW"; summary_ws["G3"].alignment = center_align; summary_ws["G3"].font = black_bold

    summary_ws["K2"] = "Measured Ave."; summary_ws["K2"].font = black_bold; summary_ws["K2"].alignment = center_align
    summary_ws["K3"] = "δ¹³C RAW"; summary_ws["K3"].alignment = center_align; summary_ws["K3"].font = black_bold
    summary_ws["N2"] = "Measured Ave."; summary_ws["N2"].font = black_bold; summary_ws["N2"].alignment = center_align
    summary_ws["N3"] = "δ¹⁸O RAW"; summary_ws["N3"].alignment = center_align; summary_ws["N3"].font = black_bold

    summary_ws["O2"] = "Stretching"; summary_ws["O2"].font = black_bold
    summary_ws["O3"] = "Factor (λ)"; summary_ws["O3"].font = black_bold

    # Fill Dynamic Standards
    current_r = std_start_row
    for std in water_stds_settings:
        name = std["col_c"]
        std_row_index_map[name] = current_r # Store for formulas later
        
        # Name
        c = summary_ws.cell(row=current_r, column=3, value=name)
        c.font = std_color_map.get(name, black_bold)
        
        # Published Values
        try: d2h = float(std["col_f"])
        except: d2h = None
        try: d18o = float(std["col_g"])
        except: d18o = None
        
        c_f = summary_ws.cell(row=current_r, column=6, value=d2h)
        c_f.font = std_color_map.get(name, black_bold); c_f.alignment = center_align
        
        c_g = summary_ws.cell(row=current_r, column=7, value=d18o)
        c_g.font = std_color_map.get(name, black_bold); c_g.alignment = center_align
        
        # Stretching Formula: (G_this - G_next) / (N_this - N_next)
        # Note: Last row cannot calculate stretching against "next"
        if current_r < std_end_row:
             # Apply yellow fill to stretching cells
            summary_ws.cell(row=current_r, column=15).fill = yellow_fill
            summary_ws.cell(row=current_r, column=15, value=f"=(G{current_r}-G{current_r+1})/(N{current_r}-N{current_r+1})")

        current_r += 1
    
    # Center alignment for measured cols
    for r in range(std_start_row, std_end_row + 1):
        for c in range(11, 16):
            summary_ws.cell(row=r, column=c).alignment = center_align

    # ==========================================
    # === BOX 2: He/CO2 (Dynamic Location) ===
    # ==========================================
    heco2_row = std_end_row + 1
    apply_box_border(summary_ws, heco2_row, 3, heco2_row, 15, fill=heco2_box_fill)
    summary_ws.cell(row=heco2_row, column=3, value="He/CO2").font = black_bold
    summary_ws.cell(row=heco2_row, column=3).alignment = center_align
    summary_ws.row_dimensions[heco2_row].height = 20 

    # ==========================================
    # === BOX 3: Slope Intercept (Dynamic) ===
    # ==========================================
    # Calculate height needed: 1 row header + (2 rows * number of groups)
    slope_box_start = heco2_row + 2
    
    # Mapping to store where the Slope/Int values land for each group
    # Structure: {"GroupString": {"slope_row": X, "int_row": Y}}
    group_slope_map = {} 
    
    current_slope_row = slope_box_start
    
    for grp_idx, group_list in enumerate(slope_groups_settings):
        # We need at least 2 items to make a slope
        if len(group_list) < 2: continue
        
        # Determine rows of the standards involved
        rows_involved = []
        for std_name in group_list:
            if std_name in std_row_index_map:
                rows_involved.append(std_row_index_map[std_name])
        
        # Sort rows to ensure order (top to bottom)
        rows_involved.sort()
        
        if len(rows_involved) < 2: continue
        
        # Check if rows are contiguous (e.g., 5, 6, 7 is contiguous; 5, 8 is not)
        is_contiguous = (rows_involved[-1] - rows_involved[0] + 1) == len(rows_involved)

        # Render Label
        lbl_cell = summary_ws.cell(row=current_slope_row, column=9)
        short_names = [n.replace("MRSI-STD-", "").replace("USGS ", "") for n in group_list]
        lbl_cell.value = ", ".join(short_names)
        lbl_cell.font = bold_blue_text_font if grp_idx == 0 else bold_green_text_font 
        
        # Render Slope/Int Text
        summary_ws.cell(row=current_slope_row, column=10, value="slope")
        summary_ws.cell(row=current_slope_row+1, column=10, value="intercept")
        
        # Build Formulas
        slope_cell = summary_ws.cell(row=current_slope_row, column=14)
        int_cell = summary_ws.cell(row=current_slope_row+1, column=14)
        
        slope_formula = ""
        int_formula = ""

        if is_contiguous:
            # Simple Range Formula: =SLOPE(G5:G7, N5:N7)
            r_min, r_max = rows_involved[0], rows_involved[-1]
            range_g = f"G{r_min}:G{r_max}"
            range_n = f"N{r_min}:N{r_max}"
            
            slope_formula = f"=SLOPE({range_g},{range_n})"
            int_formula = f"=INTERCEPT({range_g},{range_n})"
            
        else:
            # Non-Contiguous: Use CHOOSE({1,2...}, Cell1, Cell2...) to create dynamic array
            # Example: =SLOPE(CHOOSE({1,2}, G5, G8), CHOOSE({1,2}, N5, N8))
            
            # 1. Create index array string: "{1,2,3}"
            indices = list(range(1, len(rows_involved) + 1))
            idx_str = "{" + ",".join(map(str, indices)) + "}"
            
            # 2. List individual cells
            g_cells = ",".join([f"G{r}" for r in rows_involved])
            n_cells = ",".join([f"N{r}" for r in rows_involved])
            
            slope_formula = f"=SLOPE(CHOOSE({idx_str},{g_cells}),CHOOSE({idx_str},{n_cells}))"
            int_formula = f"=INTERCEPT(CHOOSE({idx_str},{g_cells}),CHOOSE({idx_str},{n_cells}))"

        slope_cell.value = slope_formula
        slope_cell.number_format = FMT_3_DEC
        int_cell.value = int_formula
        int_cell.number_format = FMT_3_DEC
        
        # Store location for Data Calculation Loop
        group_key = tuple(sorted(group_list))
        group_slope_map[group_key] = {
            "slope_addr": f"$N${current_slope_row}",
            "int_addr": f"$N${current_slope_row+1}"
        }

        current_slope_row += 2
        
    slope_box_end = current_slope_row - 1
    apply_box_border(summary_ws, slope_box_start, 8, slope_box_end, 15, fill=box_peach_fill)
    apply_vertical_divider(summary_ws, 9, slope_box_start, slope_box_end)

    # --- Dynamic Green Background Fill ---
    # We do this AFTER drawing the boxes so we know the exact height (slope_box_end).
    # We check 'fill_type is None' so we don't overwrite the Blue/Peach boxes.
    
    header_fill_bottom = slope_box_end + 1 # Add a small buffer row below the last box
    
    for r in range(1, header_fill_bottom + 1):
        for c in range(1, 51):
            cell = summary_ws.cell(row=r, column=c)
            # Only fill if the cell is currently empty/white (no existing fill)
            if cell.fill.fill_type is None:
                cell.fill = light_green_fill
    
    summary_ws["A1"] = datetime.now().strftime("%Y-%m-%d")

    # ==========================================
    # === Right Side Results (Dynamic R-Y) ===
    # ==========================================
    col_R = 18
    col_S = 19
    # T and U are spacers/unused in this logic usually, Calculation starts at V (22)
    start_calc_col = 22 
    
    # --- Map Groups to Columns Dynamically ---
    # Structure: { tuple(group_members): column_index }
    group_col_map = {}
    
    # 1. Assign Columns to Groups
    for idx, grp_list in enumerate(slope_groups_settings):
        # Sort key to ensure matching works later
        group_key = tuple(sorted(grp_list))
        # V=22, W=23, X=24, etc.
        col_idx = start_calc_col + idx 
        group_col_map[group_key] = col_idx

    # Determine the last column used
    last_calc_col = start_calc_col + len(slope_groups_settings) - 1 if slope_groups_settings else start_calc_col
    
    # Define summary columns relative to the calculation columns
    # We leave a gap after the last calculation column for STDEV and N
    col_Stdev_Res = last_calc_col + 1
    col_Count_Res = last_calc_col + 2
    
    summary_ws.cell(row=1, column=col_R, value="Water Standard Results").font = black_bold

    # Results box should align with Standards box (Row 2 to std_end_row)
    # Apply border from R to the Count Column
    apply_box_border(summary_ws, 2, col_R, std_end_row, col_Count_Res, fill=gray_box_fill)
    apply_vertical_divider(summary_ws, col_S, 2, std_end_row) 
    
    # Headers for Results Box
    
    # Row 2: "δ¹⁸O SMOW" label above all group columns
    # We merge cells across all the group columns for a cleaner look
    if group_col_map:
        first_col_idx = start_calc_col
        last_col_idx = start_calc_col + len(group_col_map) - 1
        
        # Merge and Write "δ¹⁸O SMOW"
        summary_ws.merge_cells(start_row=2, start_column=first_col_idx, end_row=2, end_column=last_col_idx)
        cell = summary_ws.cell(row=2, column=first_col_idx, value="δ¹⁸O SMOW")
        cell.alignment = center_align
        cell.font = black_bold
    
    # Row 3: Write Group Names (Moved down from Row 2)
    for group_key, col_idx in group_col_map.items():
        # LOGIC CHANGE: Check if all items in this group are USGS
        if all("USGS" in n for n in group_key):
            header_name = "USGS"
        else:
            short_names = [n.replace("MRSI-STD-", "").replace("USGS ", "") for n in group_key]
            header_name = "/".join(short_names)
        
        cell = summary_ws.cell(row=3, column=col_idx, value=header_name)
        cell.alignment = center_align
        cell.font = Font(bold=True, size=9)
    
    # 2. Write STDEV and N headers (Moved to Row 2 to align with top label or keep at 2?)
    # Usually these align with the specific column headers, so let's put them on Row 3 too
    summary_ws.cell(row=3, column=col_Stdev_Res, value="STDEV").font = black_bold
    summary_ws.cell(row=3, column=col_Count_Res, value="N").font = black_bold
    
    # Underline Header (Row 3 is now the bottom of the header)
    for c in range(col_R, col_Count_Res + 1):
        cell = summary_ws.cell(row=3, column=c)
        # Apply bottom double border
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=double_black_side)

    # Auto-adjust column widths for the calculation columns
    for col_idx in range(start_calc_col, col_Count_Res + 1):
        col_letter = get_column_letter(col_idx)
        # Set a reasonable default or estimate based on content length
        # Simple estimation: length of header string + buffer
        header_val = summary_ws.cell(row=3, column=col_idx).value or ""
        width = max(12, len(str(header_val)) + 2) 
        summary_ws.column_dimensions[col_letter].width = width
    
    # Underline Header
    for c in range(col_R, col_Count_Res + 1):
        cell = summary_ws.cell(row=3, column=c)
        if cell.border.bottom.style != medium_black_side.style:
            cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=double_black_side)

    summary_ws.cell(row=3, column=col_R, value="Calibration").font = Font(bold=False, color="000000")
    
    # Write Dynamic Rows in Result Box
    current_r_res = std_start_row
    standard_cells_map = {} 

    for std in water_stds_settings:
        name = std["col_c"]
        c_font = std_color_map.get(name, black_bold)
        
        # Name
        summary_ws.cell(row=current_r_res, column=col_R, value=name).font = c_font
        
        # Determine which column this standard belongs to
        # Logic: Find the group this standard is in, and get that group's column
        target_res_col = None
        
        for grp_list in slope_groups_settings:
            # Check if name is in this group
            if any(name.replace(" ","").upper() in m.replace(" ","").upper() for m in grp_list):
                group_key = tuple(sorted(grp_list))
                target_res_col = group_col_map.get(group_key)
                break
        
        # If not found (shouldn't happen if settings are correct), default to first available
        if target_res_col is None:
            target_res_col = start_calc_col

        standard_cells_map[name] = {
            "avg_col": target_res_col, 
            "stdev_col": col_Stdev_Res, 
            "count_col": col_Count_Res, 
            "row": current_r_res
        }
        current_r_res += 1

    # --- Group Identification Logic ---
    # Define Start Row dynamically based on slope box
    dest_start_row = slope_box_end + 3

    # --- Write Column Headers for Data Table ---
    
    super_header_row = dest_start_row - 1
    group_header_row = dest_start_row
    
    # 1. Write "δ¹⁸O SMOW" Super Header (Merged)
    if group_col_map:
        first_col = start_calc_col
        last_col = start_calc_col + len(group_col_map) - 1
        
        summary_ws.merge_cells(start_row=super_header_row, start_column=first_col, end_row=super_header_row, end_column=last_col)
        cell = summary_ws.cell(row=super_header_row, column=first_col, value="δ¹⁸O SMOW")
        cell.font = black_bold
        cell.alignment = center_align

    # 2. Write Group Names (with Light Blue Fill & Border)
    for group_key, col_idx in group_col_map.items():
        # LOGIC CHANGE: Check if all items in this group are USGS
        if all("USGS" in n for n in group_key):
            header_text = "USGS"
        else:
            short_names = [n.replace("MRSI-STD-", "").replace("USGS ", "") for n in group_key]
            header_text = "/".join(short_names)
        
        cell = summary_ws.cell(row=group_header_row, column=col_idx, value=header_text)
        cell.font = black_bold
        cell.alignment = center_align
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # Light Blue
        
        # Add a Full Box Border around this cell
        cell.border = Border(
            left=medium_black_side, 
            right=medium_black_side, 
            top=medium_black_side, 
            bottom=medium_black_side
        )

    # Important: Push the data start row down by 2 so it doesn't overwrite the headers
    dest_start_row += 1

    # --- Group Identification Logic ---
    def create_group_key(identifier):
        if identifier is None: return None
        if not isinstance(identifier, str): identifier = str(identifier)
        return re.sub(r'\s+[rR]\d+(?:\.\d+)*(?:[a-zA-Z]*)?$', '', identifier).strip()

    def reference_base_key(identifier):
        """
        Dynamic: Checks if the identifier contains the name of any standard 
        defined in settings, with fuzzy matching to ignore 'STD' and dashes.
        """
        if identifier is None: return None
        
        # 1. Clean the incoming cell identifier
        raw_text = str(identifier).upper().strip()
        text_clean = re.sub(r'[\s\-_]+', '', raw_text) # Removes spaces, dashes, underscores
        text_no_std = text_clean.replace("STD", "")
        
        # 2. Iterate through the dynamic standards loaded from settings
        for std in water_stds_settings:
            std_name = std.get("col_c")
            if not std_name: 
                continue
            
            # Clean the standard name from settings
            std_clean = re.sub(r'[\s\-_]+', '', std_name.upper())
            std_no_std = std_clean.replace("STD", "")
            
            # Match Condition 1: Exact clean match (e.g., MRSISTDW1 in MRSISTDW1RUN1)
            if std_clean in text_clean:
                return std_name
                
            # Match Condition 2: Lenient match ignoring "STD" (e.g., MRSIW1 in MRSIW1RUN1)
            # Enforce a minimum length of 4 so it doesn't accidentally match tiny strings
            if len(std_no_std) >= 4 and std_no_std in text_no_std:
                return std_name
                
        return None

    dest_start_row = slope_box_end + 3
    group_infos = []
    heco2_dst_ranges = []
    
    src_row = 3
    max_src_row = group_ws.max_row

    # Scan Groups
    while src_row <= max_src_row:
        id1_val = group_values_ws.cell(row=src_row, column=3).value
        if id1_val is None:
            src_row += 1
            continue
        data_start_src = src_row
        last_non_empty_data_row = data_start_src
        src_row += 1
        
        # Scan data rows
        while src_row <= max_src_row:
            # Check if this row is the Start of the Calculation Block (Look for "Average" in Col K/11)
            val_k = group_values_ws.cell(row=src_row, column=11).value
            if isinstance(val_k, str) and "Average" in val_k:
                break
            
            # Also check if we hit a new group ID (backup check)
            current_id1_val = group_values_ws.cell(row=src_row, column=3).value
            if current_id1_val is not None:
                last_non_empty_data_row = src_row
            src_row += 1
            
        data_end_src = last_non_empty_data_row
        calc_mid_src = None
        
        # Now verify if we are at the calculation block
        if src_row <= max_src_row:
            val_k = group_values_ws.cell(row=src_row, column=11).value
            if isinstance(val_k, str) and "Average" in val_k:
                calc_mid_src = src_row # This is the Label row (Average)
            else:
                for offset in range(1, 4):
                    if src_row + offset <= max_src_row:
                        v = group_values_ws.cell(row=src_row + offset, column=11).value
                        if isinstance(v, str) and "Average" in v:
                            calc_mid_src = src_row + offset
                            break
        
        # If still not found, fallback
        if calc_mid_src is None:
            calc_mid_src = data_end_src + 1 

        first_id = group_values_ws.cell(row=data_start_src, column=3).value
        base_key = create_group_key(first_id)
        ref_base = reference_base_key(base_key)
        group_infos.append({
            'src_data_start': data_start_src,
            'src_data_end': data_end_src,
            'src_calc_mid': calc_mid_src,
            'base_key': base_key,
            'ref_base': ref_base
        })
        # Move past the calculation block to find next group.
        # Spacing in Step 5 is at least 6 rows for full block
        src_row = calc_mid_src + 6 

    src_to_dst_row = {}
    dst_row = dest_start_row
    
    # Copy Data
    for src_r in range(1, group_ws.max_row + 1):
        is_valid_heco2 = (src_r in valid_heco2_src_rows)
        id_val = group_values_ws.cell(row=src_r, column=3).value
        id_color = None
        norm_ref = reference_base_key(id_val)
        if norm_ref and norm_ref in std_color_map:
            id_color = std_color_map[norm_ref]

        # --- CHANGED: Only copy columns 1 to 17 (A to Q) ---
        max_col_to_copy = min(group_ws.max_column, 17)
        for col_idx in range(1, max_col_to_copy + 1):
            src_cell = group_ws.cell(row=src_r, column=col_idx)
            val_cell = group_values_ws.cell(row=src_r, column=col_idx)
            dest_cell = summary_ws.cell(row=dst_row, column=col_idx)
            dest_cell.value = val_cell.value
            try:
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.alignment = copy(src_cell.alignment)
            except Exception: pass
            
            if is_valid_heco2:
                dest_cell.fill = heco2_gray_fill
            if col_idx == 3 and id_color:
                dest_cell.font = id_color

        src_to_dst_row[src_r] = dst_row
        dst_row += 1

    # Map Dst Rows
    for gi in group_infos:
        gi['dst_data_start'] = src_to_dst_row.get(gi['src_data_start'])
        gi['dst_data_end'] = src_to_dst_row.get(gi['src_data_end'])
        gi['dst_calc_mid'] = src_to_dst_row.get(gi['src_calc_mid'])
        
        # Colorize summary lines in the main body
        calc_mid = gi['dst_calc_mid']
        if calc_mid:
            norm_ref = gi['ref_base']
            if norm_ref and norm_ref in std_bold_color_map:
                bold_color = std_bold_color_map[norm_ref]
                # Highlighting covers Average (0), Values (1), Blank (2), Outlier Label(3), Outlier Val(4)
                # Apply to 5 rows to catch everything
                for r_offset in range(5): 
                    target_r = calc_mid + r_offset
                    if target_r < summary_ws.max_row:
                        for c_idx in range(11, 17):
                            cell = summary_ws.cell(row=target_r, column=c_idx)
                            if cell.value:
                                cell.font = bold_color
        
        bk = str(gi['base_key']).lower()
        if "heco2" in bk or "co2" in bk:
            heco2_dst_ranges.append(gi)

    # --- 3. Construct "All" Summary Blocks ---
    ref_groups_by_base = {}
    for gi in group_infos:
        rb = gi['ref_base']
        if rb: ref_groups_by_base.setdefault(rb, []).append(gi)

    insert_actions = []
    for rb, gis in ref_groups_by_base.items():
        if len(gis) >= 1:
            last_gi = sorted(gis, key=lambda x: (x['dst_calc_mid'] or 0))[-1]
            
            # --- INSERTION FIX ---
            # Check for "Outlier Excl." block (located 3 rows down from mid because of gap)
            current_calc_mid = last_gi['dst_calc_mid']
            physical_block_height = 1 # Default (0..1)
            
            if current_calc_mid:
                # Look at Row 3 (Mid+3) for Label
                check_row = current_calc_mid + 3
                if check_row <= summary_ws.max_row:
                    check_val = summary_ws.cell(row=check_row, column=11).value
                    if check_val and "Average" in str(check_val):
                        # Block is: Label(0), Val(1), Gap(2), Label(3), Val(4)
                        physical_block_height = 5 
                
                # If we didn't find Outlier block, just insert after All Values (Mid+1)
                if physical_block_height == 1:
                    insert_after = current_calc_mid + 1
                else:
                    insert_after = current_calc_mid + 4

            else:
                insert_after = last_gi['dst_data_end']
                
            if insert_after is None: continue
            insert_actions.append({'insert_after': insert_after, 'ref_base': rb, 'groups': gis})

    insert_actions.sort(key=lambda x: x['insert_after'] or 0)

    cumulative_offset = 0

    for action in insert_actions:
        groups = action['groups']
        base_insert_at = action['insert_after'] + 1 
        insert_at = base_insert_at + cumulative_offset
        
        summary_ws.insert_rows(insert_at, amount=3)
        cumulative_offset += 3 
        
        label_row = insert_at + 1
        value_row = insert_at + 2
        
        # Headers
        title = "All" if not use_outliers else "All (Outliers Excl)"
        cell_all = summary_ws.cell(row=label_row, column=10, value=title)
        cell_all.font = black_bold
        cell_all.fill = yellow_fill

        labels = {11: "Average", 12: "Stdev", 13: "Count", 14: "Average", 15: "Stdev", 16: "Count"}
        for c, txt in labels.items():
            cell = summary_ws.cell(row=label_row, column=c, value=txt)
            cell.font = black_bold
            cell.fill = yellow_fill

        # --- NEW LOGIC: Collect INDIVIDUAL raw data cells instead of the sub-group averages ---
        k_refs = [] # For Carbon (Col 11)
        n_refs = [] # For Oxygen (Col 14)
        
        for g in groups:
            ds = g.get('dst_data_start')
            de = g.get('dst_data_end')
            if not ds or not de:
                continue
                
            for r in range(ds, de + 1):
                # Check if it's a valid data row (Skip HeCO2 inner headers)
                id_val = summary_ws.cell(row=r, column=3).value
                if isinstance(id_val, str):
                    norm_val = id_val.strip().upper()
                    if norm_val.startswith("HECO2") or norm_val.startswith("CO2"):
                        continue 
                
                # Target Cells
                cell_k = summary_ws.cell(row=r, column=11)
                cell_n = summary_ws.cell(row=r, column=14)
                
                val_k = cell_k.value
                val_n = cell_n.value
                
                # Check for Red Strikethrough from Step 5 Outlier Logic
                strike_k = cell_k.font and cell_k.font.strike
                strike_n = cell_n.font and cell_n.font.strike
                
                # Add to list if it has a value AND (not in outlier mode OR not struck through)
                if val_k is not None:
                    if not use_outliers or not strike_k:
                        k_refs.append(f"K{r}")
                        
                if val_n is not None:
                    if not use_outliers or not strike_n:
                        n_refs.append(f"N{r}")

        # Helper to safely build formulas avoiding empty lists
        def build_formula(func, refs):
            if not refs: return ""
            return f"={func}({','.join(refs)})"

        # Write Values to the Yellow "All" Block using the pool of individual data cells
        if k_refs:
            summary_ws.cell(row=value_row, column=11, value=build_formula("AVERAGE", k_refs)).number_format = FMT_3_DEC
            summary_ws.cell(row=value_row, column=12, value=build_formula("STDEV", k_refs)).number_format = FMT_3_DEC
            summary_ws.cell(row=value_row, column=13, value=build_formula("COUNT", k_refs))
        
        if n_refs:
            summary_ws.cell(row=value_row, column=14, value=build_formula("AVERAGE", n_refs)).number_format = FMT_3_DEC
            summary_ws.cell(row=value_row, column=15, value=build_formula("STDEV", n_refs)).number_format = FMT_3_DEC
            summary_ws.cell(row=value_row, column=16, value=build_formula("COUNT", n_refs))

        for c_idx in range(11, 18):
            summary_ws.cell(row=value_row, column=c_idx).font = black_bold

        # --- NEW LOGIC: Force Conditional Formatting on these specific inserted STDEV cells ---
        if stdev_threshold is not None:
            thresh_str = str(stdev_threshold)
            
            # Column L (12) is Carbon Stdev
            if k_refs:
                rule_L = FormulaRule(formula=[f'AND(ISNUMBER(L{value_row}), L{value_row} > {thresh_str})'], fill=fill_error)
                summary_ws.conditional_formatting.add(f"L{value_row}", rule_L)
                
            # Column O (15) is Oxygen Stdev
            if n_refs:
                rule_O = FormulaRule(formula=[f'AND(ISNUMBER(O{value_row}), O{value_row} > {thresh_str})'], fill=fill_error)
                summary_ws.conditional_formatting.add(f"O{value_row}", rule_O)

        # Shift downstream group pointers
        for gi in group_infos:
            for key in ['dst_data_start', 'dst_data_end', 'dst_calc_mid']:
                if gi[key] is not None and gi[key] >= insert_at:
                    gi[key] += 3

        # Link Top Blue Box (Rows 5-8) to this new "All" block
        ref_base = action['ref_base']
        
        if ref_base and ref_base in std_row_index_map:
            target_r = std_row_index_map[ref_base]
            
            # Get the color for this standard
            color_f = std_bold_color_map.get(ref_base, black_bold)
            
            # Write formulas pointing to the new "All" block
            c = summary_ws.cell(row=target_r, column=11, value=f"=K{value_row}")
            c.font = color_f; c.number_format = FMT_3_DEC
            
            c = summary_ws.cell(row=target_r, column=14, value=f"=N{value_row}")
            c.font = color_f; c.number_format = FMT_3_DEC
    
    # --- 4. Fill in He/CO2 Box (Dynamic Row) ---
    # use the 'heco2_row' variable we calculated in the layout section
    
    if heco2_dst_ranges:
        calc_rows = []
        for g in heco2_dst_ranges:
            if g.get('dst_calc_mid'):
                calc_rows.append(g['dst_calc_mid'] + calc_value_offset)
        
        if calc_rows:
            colK, colL = get_column_letter(11), get_column_letter(12)
            colN, colO = get_column_letter(14), get_column_letter(15)

            def make_ref(col_let, rows):
                refs = [f"{col_let}{r}" for r in rows]
                if len(refs) == 1: return f"={refs[0]}"
                return f"=AVERAGE({','.join(refs)})"

            # Use heco2_row instead of hardcoded 9
            summary_ws.cell(row=heco2_row, column=11, value=make_ref(colK, calc_rows)).number_format = FMT_3_DEC
            summary_ws.cell(row=heco2_row, column=11).font = black_bold
            
            summary_ws.cell(row=heco2_row, column=12, value=make_ref(colL, calc_rows)).number_format = FMT_3_DEC
            summary_ws.cell(row=heco2_row, column=12).font = black_bold
            
            summary_ws.cell(row=heco2_row, column=14, value=make_ref(colN, calc_rows)).number_format = FMT_3_DEC
            summary_ws.cell(row=heco2_row, column=14).font = black_bold
            
            summary_ws.cell(row=heco2_row, column=15, value=make_ref(colO, calc_rows)).number_format = FMT_3_DEC
            summary_ws.cell(row=heco2_row, column=15).font = black_bold
    
    # --- Match widths ---
    for col in range(1, group_ws.max_column + 1):
        col_letter = get_column_letter(col)
        src_dim = group_ws.column_dimensions[col_letter]
        summary_ws.column_dimensions[col_letter].width = src_dim.width or 15
    summary_ws.column_dimensions['C'].width = 30
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['R'].width = 16
    # Dynamic Freeze Pane
    summary_ws.freeze_panes = f"B{dest_start_row + 1}"

    # --- Calculations Columns (Dynamic V, W, X...) ---
    
    # Store ranges for the Result Box summary
    # CHANGED: Dict now just stores rows: { "StdName": [row1, row2, ...] }
    standard_ranges = {} 
    col_N_str = get_column_letter(14)
    
    for gi in group_infos:
        ds = gi.get('dst_data_start')
        de = gi.get('dst_data_end')
        if not ds or not de or de < ds: continue

        for r in range(ds, de + 1):
            id_val = summary_ws.cell(row=r, column=3).value
            
            # Check for Header/Skip rows (HeCO2)
            skip_row = False
            if isinstance(id_val, str):
                normalized_val = id_val.strip().upper()
                if normalized_val.startswith("HECO2") or normalized_val.startswith("CO2"): skip_row = True
            
            # Write ID to Column S (Spacer)
            if not skip_row: summary_ws.cell(row=r, column=col_S, value=id_val)
            else: summary_ws.cell(row=r, column=col_S, value=None)
            
            if skip_row:
                for c_idx in group_col_map.values():
                    summary_ws.cell(row=r, column=c_idx, value=None)
                continue

            # 1. UNIVERSAL CALCULATION: Calculate Logic for EVERY Group/Column
            for group_key, col_idx in group_col_map.items():
                if group_key in group_slope_map:
                    slope_cell = group_slope_map[group_key]["slope_addr"]
                    intercept_cell = group_slope_map[group_key]["int_addr"]
                    
                    v_formula = f"=IF({col_N_str}{r}=\"\",\"\",{col_N_str}{r}*{slope_cell}+{intercept_cell})"
                    c = summary_ws.cell(row=r, column=col_idx, value=v_formula)
                    c.number_format = FMT_2_DEC

            # 2. REFERENCE MAPPING: Just track the row number
            ref = reference_base_key(id_val)
            if ref:
                if ref not in standard_ranges:
                    standard_ranges[ref] = []
                standard_ranges[ref].append(r)

        # Recalculate Average row location
        calc_mid = gi.get('dst_calc_mid')
        if calc_mid and isinstance(calc_mid, int): avg_row = calc_mid + 1
        else: avg_row = de + 2

        # Check if we need space for the STDEV row
        # We need avg_row-1 (Average) and avg_row (STDEV) to be available.
        if summary_ws.cell(row=avg_row, column=col_S).value is not None:
            summary_ws.insert_rows(avg_row, amount=1)
            # Update pointers for subsequent groups
            for gi2 in group_infos:
                for key in ['dst_data_start', 'dst_data_end', 'dst_calc_mid']:
                    if gi2.get(key) is not None and gi2[key] >= avg_row: gi2[key] += 1
            if de >= avg_row: de += 1
            if calc_mid and calc_mid >= avg_row: calc_mid += 1

        # --- ROW 1: AVERAGE ---
        summary_ws.cell(row=avg_row - 1, column=col_S, value="Average").font = Font(bold=True)
        
        # --- ROW 2: STDEV ---
        summary_ws.cell(row=avg_row, column=col_S, value="STDEV").font = Font(bold=True)
        
        # Calculate Per-Column Stats (Average on Top, STDEV Below)
        for group_key, col_idx in group_col_map.items():
            col_let = get_column_letter(col_idx)
            rng = f"{col_let}{ds}:{col_let}{de}"
            
            # Row 1: Average
            c = summary_ws.cell(row=avg_row - 1, column=col_idx, value=f"=IF(COUNT({rng})=0,\"\",AVERAGE({rng}))")
            c.font = Font(bold=True); c.number_format = FMT_2_DEC
            
            # Row 2: STDEV
            c = summary_ws.cell(row=avg_row, column=col_idx, value=f"=IF(COUNT({rng})=0,\"\",STDEV({rng}))")
            c.font = Font(bold=True); c.number_format = FMT_2_DEC
        
        # Block Summary Stats (Far Right)
        all_cols_range = f"{get_column_letter(start_calc_col)}{ds}:{get_column_letter(last_calc_col)}{de}"
        
        # Count (Row 1) - Keeps the total count of N
        summary_ws.cell(row=avg_row - 1, column=col_Count_Res, value=f'=COUNT({all_cols_range})').font = Font(bold=True)
        
        # Apply Blue Fill to BOTH summary rows
        for r_offset in [0, -1]: # avg_row (STDEV) and avg_row-1 (Average)
            for col_idx in range(col_S, col_Count_Res + 1):
                 summary_ws.cell(row=avg_row + r_offset, column=col_idx).fill = box_blue_fill
        
    # --- Conditional Formatting (Columns L and O) ---
    if stdev_threshold is not None:
        thresh_str = str(stdev_threshold)
        col_L = "L"
        col_O = "O"
        
        for gi in group_infos:
            ds = gi.get('dst_data_start')
            de = gi.get('dst_data_end')
            calc_row = gi.get('dst_calc_mid')
            
            if not ds or not de or not calc_row: continue
            if ds < 19: continue
            
            # --- FIX: Use FormulaRule with ISNUMBER to ignore text ---
            
            # 1. Data Rows
            rng_L = f"{col_L}{ds}:{col_L}{de}"
            rng_O = f"{col_O}{ds}:{col_O}{de}"
            
            # FormulaRule applies relative to the top-left cell of the range.
            rule_L_data = FormulaRule(
                formula=[f'AND(ISNUMBER({col_L}{ds}), {col_L}{ds} > {thresh_str})'],
                fill=fill_error
            )
            rule_O_data = FormulaRule(
                formula=[f'AND(ISNUMBER({col_O}{ds}), {col_O}{ds} > {thresh_str})'],
                fill=fill_error
            )
            summary_ws.conditional_formatting.add(rng_L, rule_L_data)
            summary_ws.conditional_formatting.add(rng_O, rule_O_data)
            
            # 2. Calculation Rows (Base + Offsets)
            # Apply to ALL possible calculation rows (0 to 5) to be safe
            for off in range(0, 6): 
                cr = calc_row + off
                if cr < summary_ws.max_row:
                    cell_L = f"{col_L}{cr}"
                    cell_O = f"{col_O}{cr}"
                    
                    rule_L_calc = FormulaRule(
                        formula=[f'AND(ISNUMBER({cell_L}), {cell_L} > {thresh_str})'],
                        fill=fill_error
                    )
                    rule_O_calc = FormulaRule(
                        formula=[f'AND(ISNUMBER({cell_O}), {cell_O} > {thresh_str})'],
                        fill=fill_error
                    )
                    summary_ws.conditional_formatting.add(cell_L, rule_L_calc)
                    summary_ws.conditional_formatting.add(cell_O, rule_O_calc)

    # --- Calibration Summary Boxes Logic (Dynamic) ---
    
    for norm_ref, rows in standard_ranges.items():
        # Find where this standard sits in the Summary Box (top right)
        cell_info = None
        
        # Direct Match
        if norm_ref in standard_cells_map:
            cell_info = standard_cells_map[norm_ref]
        else:
            # Fuzzy Match
            for k, v in standard_cells_map.items():
                if norm_ref.replace(" ","").lower() in k.replace(" ","").lower():
                    cell_info = v
                    break
        
        if not cell_info or not rows: 
            continue
        
        target_row = cell_info['row']
        target_stdev_col = cell_info['stdev_col']
        target_count_col = cell_info['count_col']
        
        # 1. FILL EVERY GROUP COLUMN (Average)
        # We iterate all available calculation columns (V, W, X...)
        for group_key, col_idx in group_col_map.items():
            col_letter = get_column_letter(col_idx)
            
            # Construct non-contiguous range for this specific column (e.g., V5, V8, V12)
            ranges_str = ",".join([f"{col_letter}{r}" for r in rows])
            
            # Write AVERAGE formula
            c = summary_ws.cell(row=target_row, column=col_idx)
            c.value = f"=IF(COUNT({ranges_str})=0,\"\",AVERAGE({ranges_str}))"
            c.number_format = FMT_2_DEC
        
        # 2. FILL STDEV & COUNT
        # We need to decide which column to use for the "Official" STDEV/Count.
        # Logic: Use the column corresponding to the group this standard belongs to.
        # If the standard isn't in a group (unlikely), default to the first column.
        
        stats_source_col_idx = start_calc_col # Default to first group
        
        # Find the group this standard actually belongs to
        for grp_list in slope_groups_settings:
            if any(norm_ref.replace(" ", "").upper() in m.replace(" ", "").upper() for m in grp_list):
                group_key = tuple(sorted(grp_list))
                if group_key in group_col_map:
                    stats_source_col_idx = group_col_map[group_key]
                break

        stats_col_letter = get_column_letter(stats_source_col_idx)
        stats_ranges_str = ",".join([f"{stats_col_letter}{r}" for r in rows])
        
        # Write STDEV
        c = summary_ws.cell(row=target_row, column=target_stdev_col)
        c.value = f"=IF(COUNT({stats_ranges_str})=0,\"\",STDEV({stats_ranges_str}))"
        c.number_format = FMT_2_DEC
        
        # Write Count
        c = summary_ws.cell(row=target_row, column=target_count_col)
        c.value = f"=COUNT({stats_ranges_str})"

        # Re-apply color to the whole row in the results box
        base_color = std_color_map.get(norm_ref)
        if base_color:
            for c_idx in range(col_R, col_Count_Res + 1):
                cell = summary_ws.cell(row=target_row, column=c_idx)
                cell.font = copy(base_color)
                cell.alignment = center_align

    # --- Final Polish & Save ---
    
    # Set the new sheet as active
    for ws in wb.worksheets: 
        ws.sheet_view.tabSelected = False
    summary_ws.sheet_view.tabSelected = True
    wb.active = wb.index(summary_ws)

    # Add Settings Popup Comment
    embed_settings_popup(summary_ws, "A2")

    # Adjust specific column widths
    summary_ws.column_dimensions["J"].width = 16 
    summary_ws.column_dimensions["C"].width = 30
    
    wb.save(file_path)
    print(f"✅ Step 6: Normalization completed on {file_path})")