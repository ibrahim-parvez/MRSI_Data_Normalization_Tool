import pandas as pd
import re
import statistics
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
import utils.settings as settings
from utils.common_utils import embed_settings_popup

# --- Helper Functions ---

def get_base_key(identifier):
    """
    Strips trailing run numbers (e.g., 'Sample r1', 'Sample r1.2') 
    to get the base name for grouping.
    """
    if identifier is None:
        return None
    s_id = str(identifier)
    return re.sub(r'\s+[rR]\d+(?:\.\d+)*(?:[a-zA-Z]*)?$', '', s_id).strip()

def extract_run_number(identifier):
    """
    Parses 'Sample r1.2' -> returns (1, 2).
    """
    if not identifier: return None, None
    match = re.search(r'[rR](\d+)(?:\.(\d+))?', str(identifier))
    if match:
        major = int(match.group(1))
        minor = int(match.group(2)) if match.group(2) else 0
        return major, minor
    return None, None

def _get_valid_co2_rows(rows_identifiers):
    """
    CO2 Logic: Skip R1, keep lowest minor for R2+.
    Returns indices of valid rows.
    """
    valid_indices = set()
    seen_majors = {} 

    for i, ident_val in enumerate(rows_identifiers):
        ident_raw = str(ident_val or "").strip()
        ident_lower = ident_raw.lower()

        if ident_lower.startswith("heco2"):
            ident_clean = "co2" + ident_raw[len("heco2"):]
        elif ident_lower.startswith("co2"):
            ident_clean = "co2" + ident_raw[len("co2"):]
        else:
            ident_clean = ident_raw
        
        major, minor = extract_run_number(ident_clean)
        
        if major is None or major == 1:
            continue
            
        if major not in seen_majors:
            seen_majors[major] = (minor, i)
        else:
            prev_minor, prev_i = seen_majors[major]
            if minor < prev_minor:
                seen_majors[major] = (minor, i)
    
    for _, idx in seen_majors.values():
        valid_indices.add(idx)
        
    return valid_indices

def get_material_colors(material_name, ref_materials):
    """
    Returns (ID_Font, Summary_Font) based on material settings.
    ID_Font: Color, No Bold (for Col C).
    Summary_Font: Color, Bold (for Stats).
    """
    COLOR_MAP = {
        "green": "008000",      # Dark Green for Text
        "red": "FF0000",        # Red
        "blue": "0000FF",       # Blue
        "lightblue": "00CCFF",  # Cyan-ish
        "darkblue": "00008B",   # Dark Blue
        "orange": "FF9900",     # Orange
        "yellow": "CCCC00",     # Darker Yellow for visibility
    }
    
    for mat in ref_materials:
        if mat.get("col_c") == material_name:
            c_name = mat.get("color", "").lower()
            hex_font = COLOR_MAP.get(c_name, "000000")
            
            # Font for ID Column: Color only, NO Bold
            id_font_obj = Font(bold=False, color=hex_font)
            
            # Font for Summary Text: Color + Bold
            summary_font_obj = Font(bold=True, color=hex_font)
            
            return id_font_obj, summary_font_obj
            
    return None, None

def copy_cell_style(src_cell, tgt_cell):
    if src_cell is None or tgt_cell is None: return
    if src_cell.has_style:
        try:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = src_cell.number_format 
            tgt_cell.alignment = copy(src_cell.alignment)
            tgt_cell.protection = copy(src_cell.protection)
        except: pass

# --- Main Function ---

def step5_group_carbonate(file_path: str):
    new_sheet_name = "Group_DNT"
    source_sheet_name = "Last 6_DNT"
    target_sibling_name = "Pre-Group_DNT"

    # --- 1. Load Settings ---
    # --- Configuration from Settings ---
    stdev_is_enabled = settings.get_setting("STDEV_THRESHOLD_ENABLED")
    
    # If disabled, set the variable to None so it bypasses conditional formatting
    if stdev_is_enabled:
        stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    else:
        stdev_threshold = None
    outlier_sigma = settings.get_setting("OUTLIER_SIGMA") or 2
    exclusion_mode = settings.get_setting("OUTLIER_EXCLUSION_MODE") or "Individual"


    # Load Reference Materials
    # Carbonate (for coloring and inclusion in Phase 1)
    ref_materials_carb = settings.get_setting("REFERENCE_MATERIALS", sub_key="Carbonate") or []
    carb_names = {m.get("col_c") for m in ref_materials_carb if m.get("col_c")}
    
    # Water (for exclusion from this sheet entirely)
    ref_materials_water = settings.get_setting("REFERENCE_MATERIALS", sub_key="Water") or []
    water_names = {m.get("col_c") for m in ref_materials_water if m.get("col_c")}
    
    # Styles
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") 
    gray_fill = PatternFill(start_color="E7E7E7", end_color="E7E7E7", fill_type="solid") 
    header_fill = PatternFill(start_color="8ED973", end_color="8ED973", fill_type="solid")
    dark_gray_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid") # Dark Gray for separator

    bold_font = Font(bold=True)
    normal_font = Font(bold=False) # Explicit non-bold
    strike_font = Font(strike=True, color="FF0000") 

    THREE_DECIMAL_FORMAT = "0.000"

    # --- 2. Load Workbook ---
    wb_values = load_workbook(file_path, data_only=True) 
    wb = load_workbook(file_path) 

    matched_source = next((s for s in wb.sheetnames if s.lower() == source_sheet_name.lower()), None)
    if matched_source is None:
        print(f"❌ Source sheet matching '{source_sheet_name}' not found.")
        return

    source_ws_vals = wb_values[matched_source]
    source_ws = wb[matched_source]

    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    if target_sibling_name in wb.sheetnames:
        idx = wb.sheetnames.index(target_sibling_name)
    else:
        idx = wb.index(source_ws) + 1
    
    new_ws = wb.create_sheet(new_sheet_name, idx)

    # --- 3. Define Columns ---
    EXCLUDED_COLS = {8, 9, 11, 12, 13, 26, 27, 14, 15}
    MAX_SOURCE_COL = 24
    
    col_map = {}
    dest_col_counter = 1
    source_cols_ordered = [] 

    for src_c in range(1, MAX_SOURCE_COL + 1):
        if src_c not in EXCLUDED_COLS:
            col_map[src_c] = dest_col_counter
            source_cols_ordered.append(src_c)
            dest_col_counter += 1
            
    col_C_meas = 11 # K
    col_O_meas = 14 # N
    col_id_dest = 3 

    # --- 4. Headers (Initial Copy) ---
    src_header_row = list(source_ws[1])
    for src_c, dest_c in col_map.items():
        if src_c <= len(src_header_row):
            src_cell = source_ws.cell(row=1, column=src_c)
            dest_cell = new_ws.cell(row=1, column=dest_c, value=src_cell.value)
            copy_cell_style(src_cell, dest_cell)
            if not src_cell.fill or src_cell.fill.patternType is None:
                dest_cell.fill = header_fill

    new_ws.column_dimensions[get_column_letter(3)].width = 22
    new_ws.column_dimensions[get_column_letter(11)].width = 12
    new_ws.column_dimensions[get_column_letter(14)].width = 12
    new_ws.freeze_panes = "A2"

    # --- 5. Group Data & Sort ---
    grouped_rows = {}
    max_row = source_ws.max_row
    
    for r in range(2, max_row + 1):
        if source_ws.row_dimensions[r].hidden: continue
        id_val = source_ws_vals.cell(row=r, column=3).value
        base_key = get_base_key(id_val)
        if base_key not in grouped_rows: grouped_rows[base_key] = []
        grouped_rows[base_key].append(r)

    # Split into Reference List (Phase 1) and Sample List (Phase 2)
    # This ensures separator is always between Refs and Samples.
    refs_list = []
    samples_list = []

    for base_key, row_indices in grouped_rows.items():
        k_str = str(base_key)
        k_lower = k_str.lower()
        
        # Exclude Water
        if k_str in water_names:
            continue
            
        # UPDATED LOGIC HERE:
        # Check if Reference: Either in Carb settings list OR contains "co2"/"heco2"
        is_co2_ref = "co2" in k_lower or "heco2" in k_lower
        
        if k_str in carb_names or is_co2_ref:
            refs_list.append((base_key, row_indices))
        else:
            samples_list.append((base_key, row_indices))

    # --- 6. Process Groups Loop (Phase 1 -> Separator -> Phase 2) ---
    dest_row = 2 
    
    # We iterate through phases: first Refs, then Samples
    phases = [("refs", refs_list), ("samples", samples_list)]

    for phase_name, group_list in phases:
        
        # INSERT SEPARATOR (Only before starting Samples phase)
        if phase_name == "samples":
             # 1. Six Blank Rows
            dest_row += 9

            # 2. Two Dark Gray Rows (Indefinite Width)
            for _ in range(2):
                new_ws.row_dimensions[dest_row].fill = dark_gray_fill
                dest_row += 1

            # 3. Manual Header Insertion
            # Mapping: A=Row, B=Time Code, C=Identifier 1, D=Comment, E=Identifier 2, 
            # F=Analysis, G=Preparation, H=Ampl 44, K=C avg, L=C stdev, N=O avg, O=O stdev, Q=Sum Area All
            manual_header_map = {
                1: "Row", 2: "Time Code", 3: "Identifier 1", 4: "Comment",
                5: "Identifier 2", 6: "Analysis", 7: "Preparation", 8: "Ampl 44",
                11: "C avg", 12: "C stdev", 14: "O avg", 15: "O stdev", 17: "Sum Area All"
            }

            # Apply Header Values & Style to ALL columns in the row (e.g., 1 to 24)
            for col_idx in range(1, MAX_SOURCE_COL + 1):
                cell = new_ws.cell(row=dest_row, column=col_idx)
                cell.fill = header_fill
                cell.font = normal_font # Explicitly No Bold
                
                if col_idx in manual_header_map:
                    cell.value = manual_header_map[col_idx]
            
            dest_row += 1

        # Process the list for the current phase
        for base_key, row_indices in group_list:
            
            # A. Get Style Info
            id_font, summary_font = get_material_colors(str(base_key), ref_materials_carb)
            if not summary_font:
                summary_font = bold_font

            # B. CO2 Validity
            group_identifiers = [source_ws_vals.cell(row=r, column=3).value for r in row_indices]
            is_co2_group = False
            valid_indices_set = set(range(len(row_indices))) 

            if base_key and ("co2" in base_key.lower() or "heco2" in base_key.lower()):
                is_co2_group = True
                valid_indices_set = _get_valid_co2_rows(group_identifiers)

            # C. Write Rows
            group_dest_rows = [] 
            
            for i, src_row in enumerate(row_indices):
                current_dest_row = dest_row
                group_dest_rows.append(current_dest_row)

                is_valid_run_idx = (i in valid_indices_set)

                for src_c in source_cols_ordered:
                    dest_c = col_map[src_c]
                    val = source_ws_vals.cell(row=src_row, column=src_c).value
                    dest_cell = new_ws.cell(row=dest_row, column=dest_c, value=val)
                    src_cell = source_ws.cell(row=src_row, column=src_c)
                    copy_cell_style(src_cell, dest_cell)

                    # Gray Fill for Valid CO2
                    if is_co2_group and is_valid_run_idx:
                        dest_cell.fill = gray_fill
                    
                    # Material Styling for ID (Font Only)
                    if dest_c == col_id_dest and id_font:
                        dest_cell.font = id_font
                
                    # Ensure Meas cols have gray background
                    if not (is_co2_group and is_valid_run_idx):
                        new_ws.cell(row=dest_row, column=col_C_meas).fill = gray_fill
                        new_ws.cell(row=dest_row, column=col_O_meas).fill = gray_fill
                
                dest_row += 1

            # D. Calculate Outliers
            c_vals = []
            o_vals = []
            valid_run_row_indices = [] # Indices relative to the group list

            for i in range(len(group_dest_rows)):
                if is_co2_group and i not in valid_indices_set:
                    c_vals.append(None)
                    o_vals.append(None)
                else:
                    valid_run_row_indices.append(i)
                    # Read from written sheet
                    c_v = new_ws.cell(row=group_dest_rows[i], column=col_C_meas).value
                    o_v = new_ws.cell(row=group_dest_rows[i], column=col_O_meas).value
                    c_vals.append(float(c_v) if isinstance(c_v, (int, float)) else None)
                    o_vals.append(float(o_v) if isinstance(o_v, (int, float)) else None)

            # Stats for Sigma Clipping (only using valid runs)
            valid_c_nums = [v for v in c_vals if v is not None]
            valid_o_nums = [v for v in o_vals if v is not None]
            
            mean_c = statistics.mean(valid_c_nums) if len(valid_c_nums) > 1 else 0
            stdev_c = statistics.stdev(valid_c_nums) if len(valid_c_nums) > 1 else 0
            mean_o = statistics.mean(valid_o_nums) if len(valid_o_nums) > 1 else 0
            stdev_o = statistics.stdev(valid_o_nums) if len(valid_o_nums) > 1 else 0

            # Calculate Bounds
            c_up, c_low = mean_c + (outlier_sigma * stdev_c), mean_c - (outlier_sigma * stdev_c)
            o_up, o_low = mean_o + (outlier_sigma * stdev_o), mean_o - (outlier_sigma * stdev_o)

            # Lists for Formulas
            all_candidate_cells_c = [] # Valid Runs (ignoring sigma outliers)
            all_candidate_cells_o = []
            
            final_valid_cells_c = []   # Valid Runs AND Not Sigma Outliers
            final_valid_cells_o = []

            c_let = get_column_letter(col_C_meas)
            o_let = get_column_letter(col_O_meas)

            for i in valid_run_row_indices:
                r_num = group_dest_rows[i]
                vc = c_vals[i]
                vo = o_vals[i]

                # Add to "All" list
                if vc is not None: all_candidate_cells_c.append(f"{c_let}{r_num}")
                if vo is not None: all_candidate_cells_o.append(f"{o_let}{r_num}")

                # Check Outliers
                is_c_out = False
                is_o_out = False
                
                if vc is not None and len(valid_c_nums) > 2:
                    if vc > c_up or vc < c_low: is_c_out = True
                
                if vo is not None and len(valid_o_nums) > 2:
                    if vo > o_up or vo < o_low: is_o_out = True
                
                exclude_c = False
                exclude_o = False

                if exclusion_mode == "Exclude Row":
                    if is_c_out or is_o_out:
                        exclude_c = True
                        exclude_o = True
                else:
                    if is_c_out: exclude_c = True
                    if is_o_out: exclude_o = True
                
                # Add to "Final" list
                if not exclude_c and vc is not None: final_valid_cells_c.append(f"{c_let}{r_num}")
                if not exclude_o and vo is not None: final_valid_cells_o.append(f"{o_let}{r_num}")
                
                # Strikeout
                if exclude_c and vc is not None:
                    new_ws.cell(row=r_num, column=col_C_meas).font = strike_font
                if exclude_o and vo is not None:
                    new_ws.cell(row=r_num, column=col_O_meas).font = strike_font

            # --- E. Write Dual Stats Blocks ---
            
            # 1. Block: "All" (Valid Runs)
            row_all = dest_row
            new_ws.cell(row=row_all, column=col_C_meas-1, value="--").font = bold_font
            
            for i, t in enumerate(["Average", "Stdev", "Count"]):
                # C Header
                c = new_ws.cell(row=row_all, column=col_C_meas + i, value=t)
                c.font = bold_font
                c.alignment = Alignment(horizontal='right')
                # O Header
                o = new_ws.cell(row=row_all, column=col_O_meas + i, value=t)
                o.font = bold_font
                o.alignment = Alignment(horizontal='right')
                
            row_all_calc = row_all + 1
            
            if all_candidate_cells_c:
                rng_c = ",".join(all_candidate_cells_c)
                new_ws.cell(row=row_all_calc, column=col_C_meas, value=f"=AVERAGE({rng_c})").number_format = THREE_DECIMAL_FORMAT
                new_ws.cell(row=row_all_calc, column=col_C_meas+1, value=f"=STDEV({rng_c})").number_format = THREE_DECIMAL_FORMAT
                new_ws.cell(row=row_all_calc, column=col_C_meas+2, value=f"=COUNT({rng_c})").number_format = "0"
                
            if all_candidate_cells_o:
                rng_o = ",".join(all_candidate_cells_o)
                new_ws.cell(row=row_all_calc, column=col_O_meas, value=f"=AVERAGE({rng_o})").number_format = THREE_DECIMAL_FORMAT
                new_ws.cell(row=row_all_calc, column=col_O_meas+1, value=f"=STDEV({rng_o})").number_format = THREE_DECIMAL_FORMAT
                new_ws.cell(row=row_all_calc, column=col_O_meas+2, value=f"=COUNT({rng_o})").number_format = "0"

            # Apply Color Text to All Block
            for c in range(col_C_meas, col_O_meas + 3):
                cell = new_ws.cell(row=row_all_calc, column=c)
                cell.font = summary_font

            # 2. Block: "Outlier Excl."
            row_filt = row_all_calc + 2
            new_ws.cell(row=row_filt, column=col_C_meas-1, value="Outlier Excl.").font = bold_font

            for i, t in enumerate(["Average", "Stdev", "Count"]):
                c = new_ws.cell(row=row_filt, column=col_C_meas + i, value=t)
                c.font = bold_font
                c.alignment = Alignment(horizontal='right')
                o = new_ws.cell(row=row_filt, column=col_O_meas + i, value=t)
                o.font = bold_font
                o.alignment = Alignment(horizontal='right')
                
            row_filt_calc = row_filt + 1
            
            if final_valid_cells_c:
                rng_c = ",".join(final_valid_cells_c)
                new_ws.cell(row=row_filt_calc, column=col_C_meas, value=f"=AVERAGE({rng_c})").number_format = THREE_DECIMAL_FORMAT
                new_ws.cell(row=row_filt_calc, column=col_C_meas+1, value=f"=STDEV({rng_c})").number_format = THREE_DECIMAL_FORMAT
                new_ws.cell(row=row_filt_calc, column=col_C_meas+2, value=f"=COUNT({rng_c})").number_format = "0"
                
            if final_valid_cells_o:
                rng_o = ",".join(final_valid_cells_o)
                new_ws.cell(row=row_filt_calc, column=col_O_meas, value=f"=AVERAGE({rng_o})").number_format = THREE_DECIMAL_FORMAT
                new_ws.cell(row=row_filt_calc, column=col_O_meas+1, value=f"=STDEV({rng_o})").number_format = THREE_DECIMAL_FORMAT
                new_ws.cell(row=row_filt_calc, column=col_O_meas+2, value=f"=COUNT({rng_o})").number_format = "0"

            # Apply Color Text to Filtered Block
            for c in range(col_C_meas, col_O_meas + 3):
                cell = new_ws.cell(row=row_filt_calc, column=c)
                cell.font = summary_font

            dest_row = row_filt_calc + 2 

    # --- 7. Conditional Formatting ---
    for s in wb.worksheets: s.sheet_view.tabSelected = False
    new_ws.sheet_view.tabSelected = True
    wb.active = wb.index(new_ws)
    new_ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    
    final_max_row = new_ws.max_row
    
    # Conditional Formatting for High Stdev
    if stdev_threshold is not None:
        thresh_str = str(stdev_threshold)
        rule_L = FormulaRule(formula=[f'AND(ISNUMBER(L2), L2 > {thresh_str})'], fill=fill_error)
        new_ws.conditional_formatting.add(f"L2:L{final_max_row}", rule_L)
        rule_O = FormulaRule(formula=[f'AND(ISNUMBER(O2), O2 > {thresh_str})'], fill=fill_error)
        new_ws.conditional_formatting.add(f"O2:O{final_max_row}", rule_O)

    # Add Settings Popup Comment
    embed_settings_popup(new_ws, "R1")

    wb.save(file_path)
    print(f"✅ Step 5: Group sheet '{new_sheet_name}'")