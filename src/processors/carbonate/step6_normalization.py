import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import re
import unicodedata
import statistics
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy 
import utils.settings as settings
from utils.common_utils import embed_settings_popup

# 💡 UPDATED IMPORT for FormulaRule
from openpyxl.formatting.rule import CellIsRule, FormulaRule 

# --- Helper Functions (omitted for brevity, assume unchanged) ---
def _normalize_text(text):
# ... (rest of helper function code)
    if not text:
        return ""
    text = str(text)
    text = unicodedata.normalize("NFKD", text)
    text = re.sub(r"[^A-Za-z0-9]+", "", text)
    return text.lower().strip()

def create_rich_text(parts):
    rt = CellRichText()
    for font, text in parts:
        rt.append(TextBlock(font, text))
    return rt

def extract_sample_base(identifier):
    if not identifier or not isinstance(identifier, str):
        return ""
    identifier = identifier.strip()
    base = re.sub(r"\s*r\d+(\.\d+)?$", "", identifier, flags=re.IGNORECASE)
    return base.strip()

def extract_run_number(identifier):
    if not identifier or not isinstance(identifier, str):
        return (9999, 0)
    m = re.search(r"r(\d+)(?:\.(\d+))?", identifier, flags=re.IGNORECASE)
    if m:
        major = int(m.group(1))
        minor = int(m.group(2)) if m.group(2) else 0
        return (major, minor)
    return (9999, 0)

def _make_fill(hex_color):
    c = hex_color.replace("#", "").upper()
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def _get_valid_co2_rows(rows, col_identifier1):
    valid_indices = []
    seen = {}
    for i, r in enumerate(rows):
        ident_raw = str(r[col_identifier1 - 1] or "").strip()
        ident_lower = ident_raw.lower()
        # Normalize explicit HeCO2 prefix -> co2 (preserve rest after prefix)
        if ident_lower.startswith("heco2"):
            ident_clean = "co2" + ident_raw[len("heco2"):]
        # If it starts with co2 (any casing) normalize to 'co2' prefix too
        elif ident_lower.startswith("co2"):
            ident_clean = "co2" + ident_raw[len("co2"):]
        else:
            ident_clean = ident_raw
        major, minor = extract_run_number(ident_clean)
        # if extraction failed, skip
        if major is None or minor is None:
            continue
        # skip R1.*
        if major == 1:
            continue
        # keep the lowest minor for each major
        if major not in seen:
            seen[major] = (minor, i)
        else:
            prev_minor, prev_i = seen[major]
            if minor < prev_minor:
                seen[major] = (minor, i)
    return sorted(idx for _, idx in seen.values())

def get_summary_num_format(base_name):
    return '0.000'

def draw_blue_box_structure(ws):
    """Step A: Draws the static structure (headers, colors) and returns row locations."""
    # --- Load Settings ---
    materials = settings.get_setting("REFERENCE_MATERIALS", sub_key="Carbonate")
    if not materials: materials = []
    
    slope_groups = settings.get_setting("SLOPE_INTERCEPT_GROUPS", sub_key="Carbonate")
    if not slope_groups: slope_groups = []
    
    num_materials = len(materials)
    
    # Rows
    data_start_row = 4
    data_end_row = 4 + num_materials - 1 if num_materials > 0 else 4

    # --- Styles ---
    thick = Side(border_style="thick", color="000000")
    medium = Side(border_style="medium", color="000000")
    blue_fill = PatternFill(start_color="DAE9F8", end_color="DAE9F8", fill_type="solid")
    black_bold = Font(color="000000", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    
    def get_style_font(color_name, is_bold):
        colors = {"black": "000000", "green": "008000", "red": "FF0000", 
                  "darkblue": "000080", "lightblue": "3399FF", "orange": "FF9900"}
        hex_code = colors.get(str(color_name).lower(), "000000")
        return Font(color=hex_code, bold=is_bold)

    # --- Header Info ---
    today_str = datetime.today().strftime("%Y-%m-%d")
    ws.cell(row=1, column=1, value=today_str).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=1, column=3, value="Normalization").font = black_bold
    ws.cell(row=1, column=3).alignment = center
    
    # === BOX 1: Reference Materials (Static Data) ===
    col_start = 3; col_end = 8
    ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)
    ws.cell(row=2, column=3, value="Reference Materials").font = black_bold
    ws.cell(row=2, column=3).alignment = center
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    ws.cell(row=2, column=6, value="Published (vs. VPDB)").font = black_bold
    ws.cell(row=2, column=6).alignment = center
    ws.cell(row=3, column=6, value="δ¹³C").alignment = center
    ws.cell(row=3, column=7, value="δ¹⁸O").alignment = center

    # Box 1 Formatting
    for r in range(2, 4):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_fill
            cell.border = Border(top=medium if r==2 else None, bottom=medium if r==3 else None, 
                                 left=medium if c==col_start else None, right=medium if c==col_end else None)

    # Write Reference Material Data
    for idx, mat in enumerate(materials):
        r = data_start_row + idx
        vals = [mat.get("col_c"), mat.get("col_d"), mat.get("col_e"), mat.get("col_f"), mat.get("col_g"), mat.get("col_h")]
        font_style = get_style_font(mat.get("color", "black"), mat.get("bold", False))
        
        for i, val in enumerate(vals):
            c = col_start + i
            if i in [3, 4]: # Columns F, G
                try: val = float(val)
                except: pass
            ws.cell(row=r, column=c, value=val).font = font_style
            ws.cell(row=r, column=c).alignment = center
            
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_fill
            cell.border = Border(bottom=medium if r==data_end_row else None, 
                                 left=medium if c==col_start else None, right=medium if c==col_end else None)

    # === BOX 2: Measured (Empty placeholders) ===
    ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=14)
    ws.cell(row=2, column=10, value="Measured (vs. Working Standard)").font = black_bold
    ws.cell(row=2, column=10).alignment = center
    ws.cell(row=3, column=11, value="δ¹³C").alignment = center
    ws.cell(row=3, column=14, value="δ¹⁸O").alignment = center
    
    for r in range(2, data_end_row + 1):
        for c in range(10, 15):
            cell = ws.cell(row=r, column=c)
            cell.fill = blue_fill
            cell.alignment = center
            top = thick if r == 2 else None
            bottom = thick if r == data_end_row else None
            left = thick if c == 10 else None
            right = thick if c == 14 else None
            if r == 3: bottom = thick
            if r == 4: top = None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # === Slope / Intercept Structure (Empty placeholders) ===
    red_if = InlineFont(color='00FF0000', b=True)
    blue_if = InlineFont(color='000000FF', b=True)
    green_if = InlineFont(color='008000', b=True)
    black_if = InlineFont(color='000000', b=True)

    def get_group_rich_text(group_list):
        parts = []
        for item in group_list:
            clean = str(item).strip().upper()
            font = black_if
            if "18" in clean: font = red_if
            elif "19" in clean: font = blue_if
            elif "603" in clean: font = green_if
            parts.append((font, clean.replace("NBS","").replace("IAEA","").strip() + " "))
        return create_rich_text(parts)

    current_row = data_end_row + 2 
    slope_info = []

    for group_list in slope_groups:
        ws.cell(row=current_row, column=10, value="Slope").font = black_bold
        ws.cell(row=current_row + 1, column=10, value="Intercept").font = black_bold
        
        ws.cell(row=current_row, column=9).value = get_group_rich_text(group_list)
        ws.cell(row=current_row, column=9).alignment = center
        
        # Center align the future formula cells
        for r_off in [0, 1]:
            ws.cell(row=current_row+r_off, column=11).alignment = center
            ws.cell(row=current_row+r_off, column=14).alignment = center

        slope_info.append({"slope_row": current_row, "intercept_row": current_row + 1})
        current_row += 3 

    # Return: (Last row of blue box, Header Row, Slope Info, Mat Row Map)
    # We rebuild mat_row_map here to pass to step 2
    mat_row_map = {}
    for idx, mat in enumerate(materials):
        if mat.get("col_c"): mat_row_map[str(mat.get("col_c")).strip().lower()] = data_start_row + idx

    return current_row - 1, current_row + 3, slope_info, mat_row_map

def populate_blue_box_math(ws, slope_info, mat_row_map):
    """Step B: Fills values and formulas AFTER data is written."""
    slope_groups = settings.get_setting("SLOPE_INTERCEPT_GROUPS", sub_key="Carbonate")
    if not slope_groups: slope_groups = []
    
    # 🔴 SETTINGS: Determine which Average to use (All vs Outlier Excluded)
    calc_mode = settings.get_setting("CALC_MODE_STEP7")
    use_outlier_excluded = (calc_mode == "Outliers Excluded") # False = Use "All Values"

    # Reload materials to get colors
    materials = settings.get_setting("REFERENCE_MATERIALS", sub_key="Carbonate")
    if not materials: materials = []

    def get_style_font(color_name, is_bold):
        colors = {"black": "000000", "green": "008000", "red": "FF0000", 
                  "darkblue": "000080", "lightblue": "3399FF", "orange": "FF9900"}
        hex_code = colors.get(str(color_name).lower(), "000000")
        return Font(color=hex_code, bold=is_bold)
    
    identifier_col = 3
    c_avg_col_ref = 11 # K
    o_avg_col_ref = 14 # N
    
    # 1. Find Data Rows (Scan the sheet now that it has data)
    found_map = {} 
    
    for r in range(1, ws.max_row + 1):
        # Look for "Average" in Column K
        val = ws.cell(row=r, column=c_avg_col_ref).value
        if val and str(val).strip().lower() == "average":
            
            # Check the label in Column J (10) to distinguish "All" vs "Outlier Excl."
            # In our new layout:
            # "All" block has "--" in Col J.
            # "Outlier Excl." block has "Outlier Excl." in Col J.
            label_val = str(ws.cell(row=r, column=10).value or "").strip()
            
            is_excluded_block = ("outlier" in label_val.lower())
            
            # Filter: If we want Excluded, skip "All". If we want All, skip "Excluded".
            if use_outlier_excluded and not is_excluded_block:
                continue
            if not use_outlier_excluded and is_excluded_block:
                continue
                
            avg_row = r + 1 # The formula row is 1 below the label
            id_row = r - 1
            ident = ""
            # Scan upwards for Identifier (limit scan to 25 rows to catch headers)
            for t in range(id_row, max(1, id_row - 25), -1):
                cellv = ws.cell(row=t, column=identifier_col).value
                if cellv:
                    ident = str(cellv).strip().lower()
                    break
            
            # Match with Reference Materials
            ident_clean = re.sub(r'[\s\-_]+', '', ident.upper())
            ident_no_std = ident_clean.replace("STD", "")
            
            for mat_name, target_row in mat_row_map.items():
                orig_mat = next((m for m in materials if str(m.get("col_c","")).strip().lower() == mat_name), None)
                if not orig_mat: continue
                
                std_clean = re.sub(r'[\s\-_]+', '', str(orig_mat.get("col_c", "")).upper())
                std_no_std = std_clean.replace("STD", "")
                
                is_match = False
                # 1. Exact clean match (e.g., MRSISTDW1 in MRSISTDW1)
                if std_clean in ident_clean:
                    is_match = True
                # 2. Lenient match ignoring "STD" (e.g., MRSIW1 in MRSIW1)
                elif len(std_no_std) >= 4 and std_no_std in ident_no_std:
                    is_match = True
                    
                if is_match:
                    color = orig_mat.get("color", "black")
                    bold = orig_mat.get("bold", False)
                    font_style = get_style_font(color, bold)

                    # Write formula AND apply font color
                    c_cell = ws.cell(row=target_row, column=11, value=f'=IFERROR({get_column_letter(c_avg_col_ref)}{avg_row},"")')
                    c_cell.font = font_style
                    
                    o_cell = ws.cell(row=target_row, column=14, value=f'=IFERROR({get_column_letter(o_avg_col_ref)}{avg_row},"")')
                    o_cell.font = font_style
                    
                    found_map[mat_name] = target_row
                    break

    # 2. Write Slope/Intercept Formulas (using Helper Columns)
    center = Alignment(horizontal="center", vertical="center")
    helper_col_idx = 100 
    
    for idx, group_list in enumerate(slope_groups):
        if idx >= len(slope_info): break 
        
        current_slope_row = slope_info[idx]["slope_row"]
        
        rows = []
        for n in group_list:
            key = str(n).strip().lower()
            if key in mat_row_map and key in found_map:
                rows.append(mat_row_map[key])
        
        if len(rows) >= 2:
            h_pub_c = helper_col_idx; h_meas_c = helper_col_idx + 1
            h_pub_o = helper_col_idx + 2; h_meas_o = helper_col_idx + 3
            
            h_start_row = 2
            for i, r_idx in enumerate(rows):
                ws.cell(row=h_start_row + i, column=h_pub_c, value=f'=F{r_idx}')
                ws.cell(row=h_start_row + i, column=h_meas_c, value=f'=K{r_idx}')
                ws.cell(row=h_start_row + i, column=h_pub_o, value=f'=G{r_idx}')
                ws.cell(row=h_start_row + i, column=h_meas_o, value=f'=N{r_idx}')
            
            h_end_row = h_start_row + len(rows) - 1
            
            rc_pub = f"{get_column_letter(h_pub_c)}{h_start_row}:{get_column_letter(h_pub_c)}{h_end_row}"
            rc_meas = f"{get_column_letter(h_meas_c)}{h_start_row}:{get_column_letter(h_meas_c)}{h_end_row}"
            ro_pub = f"{get_column_letter(h_pub_o)}{h_start_row}:{get_column_letter(h_pub_o)}{h_end_row}"
            ro_meas = f"{get_column_letter(h_meas_o)}{h_start_row}:{get_column_letter(h_meas_o)}{h_end_row}"
            
            ws.cell(row=current_slope_row, column=11, value=f'=SLOPE({rc_pub},{rc_meas})')
            ws.cell(row=current_slope_row+1, column=11, value=f'=INTERCEPT({rc_pub},{rc_meas})')
            ws.cell(row=current_slope_row, column=14, value=f'=SLOPE({ro_pub},{ro_meas})')
            ws.cell(row=current_slope_row+1, column=14, value=f'=INTERCEPT({ro_pub},{ro_meas})')
            
            helper_col_idx += 5

def create_rich_text(parts):
    richtext = CellRichText()
    for font, text in parts:
        richtext.append(TextBlock(font, text))
    return richtext

def draw_lower_boxes(ws, divider_top_row, blue_fill, black_bold, green_bold):
    slope_groups = settings.get_setting("SLOPE_INTERCEPT_GROUPS", sub_key="Carbonate") or []
    num_groups = len(slope_groups)
    if num_groups == 0:
        return
    
    # --- Dynamic Layout Math ---
    c_start = 19
    o_calc_start = c_start + num_groups + 1 
    o_arag_start = o_calc_start + num_groups + 1
    
    box1_start = c_start
    box1_end = o_arag_start + num_groups - 1
    
    box2_start = box1_end + 2
    vsmow_calc_start = box2_start
    vsmow_arag_start = vsmow_calc_start + num_groups + 1 
    box2_end = vsmow_arag_start + num_groups - 1

    thick = Side(style="thick")
    
    # ===============================
    # BOX 1: VPDB (Normalized)
    # ===============================
    for r in range(divider_top_row - 3, divider_top_row + 2):
        for c in range(box1_start, box1_end + 1):
            cell = ws.cell(r, c)
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            top = thick if r == divider_top_row - 3 else None
            bottom = thick if r == divider_top_row + 1 else None 
            left = thick if c == box1_start else None
            right = thick if c == box1_end else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Headers
    ws.cell(divider_top_row - 3, box1_start, "Normalized").font = black_bold
    ws.cell(divider_top_row - 2, box1_start, "VPDB").font = black_bold
    
    if num_groups > 1:
        ws.merge_cells(start_row=divider_top_row - 1, start_column=o_calc_start, end_row=divider_top_row - 1, end_column=o_calc_start + num_groups - 1)
        ws.merge_cells(start_row=divider_top_row - 1, start_column=o_arag_start, end_row=divider_top_row - 1, end_column=box1_end)
        
    ws.cell(divider_top_row - 1, o_calc_start, "Calcite").font = black_bold
    ws.cell(divider_top_row - 1, o_arag_start, "Aragonite").font = green_bold
    
    # Rich text generator for dynamic group names
    red_font = InlineFont(color='00FF0000', b=True)
    blue_font = InlineFont(color='000000FF', b=True)
    green_font = InlineFont(color='008000', b=True)
    black_font = InlineFont(color='000000', b=True)
    
    def get_rich_text_for_group(group_list):
        parts = []
        for item in group_list:
            clean = str(item).strip().upper()
            font = black_font
            if "18" in clean: font = red_font
            elif "19" in clean: font = blue_font
            elif "603" in clean: font = green_font
            parts.append((font, clean.replace("NBS","").replace("IAEA","").strip() + " "))
        return create_rich_text(parts)
        
    for i, grp in enumerate(slope_groups):
        rt = get_rich_text_for_group(grp)
        # δ¹³C
        ws.cell(divider_top_row, c_start + i, "δ¹³C")
        ws.cell(divider_top_row + 1, c_start + i).value = rt
        # δ¹⁸O Calcite
        ws.cell(divider_top_row, o_calc_start + i, "δ¹⁸O")
        ws.cell(divider_top_row + 1, o_calc_start + i).value = rt
        # δ¹⁸O Aragonite
        ws.cell(divider_top_row, o_arag_start + i, "δ¹⁸O")
        ws.cell(divider_top_row + 1, o_arag_start + i).value = rt

    # ===============================
    # BOX 2: VSMOW 
    # ===============================
    for r in range(divider_top_row - 2, divider_top_row + 2):
        for c in range(box2_start, box2_end + 1):
            cell = ws.cell(r, c)
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            top = thick if r == divider_top_row - 2 else None
            bottom = thick if r == divider_top_row + 1 else None
            left = thick if c == box2_start else None
            right = thick if c == box2_end else None
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    ws.cell(divider_top_row - 2, box2_start, "VSMOW").font = black_bold
    
    if num_groups > 1:
        ws.merge_cells(start_row=divider_top_row - 1, start_column=vsmow_calc_start, end_row=divider_top_row - 1, end_column=vsmow_calc_start + num_groups - 1)
        ws.merge_cells(start_row=divider_top_row - 1, start_column=vsmow_arag_start, end_row=divider_top_row - 1, end_column=box2_end)
        
    ws.cell(divider_top_row - 1, vsmow_calc_start, "Calcite").font = black_bold
    ws.cell(divider_top_row - 1, vsmow_arag_start, "Aragonite").font = green_bold
    
    for i, grp in enumerate(slope_groups):
        rt = get_rich_text_for_group(grp)
        # δ¹⁸O VSMOW Calcite
        ws.cell(divider_top_row, vsmow_calc_start + i, "δ¹⁸O")
        ws.cell(divider_top_row + 1, vsmow_calc_start + i).value = rt
        # δ¹⁸O VSMOW Aragonite
        ws.cell(divider_top_row, vsmow_arag_start + i, "δ¹⁸O")
        ws.cell(divider_top_row + 1, vsmow_arag_start + i).value = rt

def _detect_decimal_places_from_format(fmt: str):
    if not fmt or not isinstance(fmt, str):
        return None
    first_section = fmt.split(';', 1)[0]
    if '.' not in first_section:
        return None
    after_dot = first_section.split('.', 1)[1]
    count = 0
    for ch in after_dot:
        if ch in ('0', '#'):
            count += 1
        else:
            break
    return count if count > 0 else None

# --- Main Function ---
def step6_normalization_carbonate(file_path):
    # 🔴 Get the threshold setting
    # --- 1. Load Settings ---

    stdev_is_enabled = settings.get_setting("STDEV_THRESHOLD_ENABLED")
    
    # If disabled, set the variable to None so it bypasses conditional formatting
    if stdev_is_enabled:
        stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    else:
        stdev_threshold = None
    outlier_sigma = settings.get_setting("OUTLIER_SIGMA") or 2
    exclusion_mode = settings.get_setting("OUTLIER_EXCLUSION_MODE") or "Individual"

    
    strike_font = Font(strike=True, color="FF0000")

    # List of columns to EXCLUDE (1-based index)
    EXCLUDED_COLS = {8, 9, 11, 12, 13, 26, 27, 14, 15}
    MAX_SOURCE_COL = 24
    
    # Create a mapping of source column index -> destination column index
    col_map = {}
    dest_col = 1
    for src_col in range(1, MAX_SOURCE_COL + 1):
        if src_col not in EXCLUDED_COLS:
            col_map[src_col] = dest_col
            dest_col += 1
    
    # The identifier column (3, 'C') is preserved
    col_identifier1 = col_map.get(3, 0)
    if col_identifier1 != 3:
        raise Exception("Identifier column (3) was moved! Logic error in col_map.")
    
    # Use load_workbook from the import
    wb = openpyxl.load_workbook(file_path, data_only=False) 
    if "Last 6_DNT" not in wb.sheetnames:
        raise ValueError("Sheet 'Last 6' not found!")
    ws_last6 = wb["Last 6_DNT"]

    # Ensure Group sheet is recreated to the LEFT of "Group"
    if "Normalization_DNT" in wb.sheetnames:
        wb.remove(wb["Normalization_DNT"])
    pre_group_index = wb.sheetnames.index("Group_DNT")
    ws_group = wb.create_sheet("Normalization_DNT", pre_group_index)

    # make sure sheets are not grouped/selected together
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
            
    # mark Last 6 as the selected tab (prevents grouping with newly created sheet)
    try:
        ws_last6.sheet_view.tabSelected = True
        ws_group.sheet_view.tabSelected = False
    except Exception:
        pass
        
    blue_fill = _make_fill("DAE9F8")
    dark_fill = _make_fill("808080")
    gray_fill = _make_fill("E7E7E7")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    color_fonts = {
        "nbs18": Font(color="FF0000"),
        "nbs19": Font(color="000080"),
        "iaea603": Font(color="008000"),
        "lsvec": Font(color="3399FF"),
    }
    
    # --- Update MAX_GROUP_COL based on new exclusion list ---
    MAX_GROUP_COL = dest_col - 1
    
    # 1. Fill top blue area based on dynamic settings
    # 2. Add blue box and get layout info
    blue_box_bottom, header_row, slope_info, mat_row_map = draw_blue_box_structure(ws_group)
    
    # Fill blue background dynamically
    for row in ws_group.iter_rows(min_row=1, max_row=blue_box_bottom, min_col=1, max_col=MAX_GROUP_COL):
        for cell in row:
            cell.fill = blue_fill
            
    # Copy headers to the dynamic header row
    headers = []
    first_row_cells = list(ws_last6[1]) if ws_last6.max_row >= 1 else []
    
    for src_col, dest_col_idx in col_map.items():
        src_cell = first_row_cells[src_col - 1] if src_col <= len(first_row_cells) else None
        header_val = src_cell.value if src_cell else None
        headers.append(header_val)
        dest_cell = ws_group.cell(row=header_row, column=dest_col_idx, value=header_val)
        if src_cell:
            dest_cell.number_format = copy(src_cell.number_format)
            dest_cell.font = copy(src_cell.font)
            dest_cell.alignment = copy(src_cell.alignment)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)

    # --- Collect original row index along with values and formats ---
    data_rows_with_index = []
    # Iterate over cell objects (not just values) to get the row index
    for row_idx, row_cells in enumerate(ws_last6.iter_rows(min_row=2, max_col=MAX_SOURCE_COL), start=2):
        row_values_filtered = []
        source_cell_formats = {}
        
        for src_col, cell in enumerate(row_cells, start=1):
            if src_col in col_map:
                row_values_filtered.append(cell.value)
                
                # --- Store a deep copy of all style attributes ---
                source_cell_formats[src_col] = {
                    'format': copy(cell.number_format),
                    'font': copy(cell.font), 
                    'fill': copy(cell.fill), 
                    'alignment': copy(cell.alignment),
                    'border': copy(cell.border)
                }
        
        if any(row_values_filtered):
            # Ensure the row has the correct number of columns
            row_values_filtered = row_values_filtered + [None] * (MAX_GROUP_COL - len(row_values_filtered))
            data_rows_with_index.append((row_idx, tuple(row_values_filtered), source_cell_formats))

    groups = {}

    # --- Grouping now stores (original_row_idx, row_values_filtered, source_cell_formats) ---
    for original_row_idx, r_values, r_formats in data_rows_with_index:
        # col_identifier1 is 3 for 'C' (Identifier) in the new, filtered list.
        ident = r_values[col_identifier1 - 1]
        base = extract_sample_base(ident)
        norm = _normalize_text(base)
        
        if norm not in groups:
            groups[norm] = {"base": base, "rows": []}
        groups[norm]["rows"].append((original_row_idx, r_values, r_formats)) # Store index, values, and formats
        
    for g in groups.values():
        # Sort based on the identifier in the 'row_values' tuple
        g["rows"].sort(key=lambda item: extract_run_number(item[1][col_identifier1 - 1]))
        
    # Load materials from settings for dynamic reference checking
    carb_materials = settings.get_setting("REFERENCE_MATERIALS", sub_key="Carbonate") or []

    def is_reference_group(base_name):
        base_clean = re.sub(r'[\s\-_]+', '', str(base_name).upper())
        base_no_std = base_clean.replace("STD", "")
        
        # Always treat CO2/HeCO2 as reference groups
        if "CO2" in base_clean or "HECO2" in base_clean:
            return True
            
        for mat in carb_materials:
            std_name = mat.get("col_c")
            if not std_name: continue
            std_clean = re.sub(r'[\s\-_]+', '', str(std_name).upper())
            std_no_std = std_clean.replace("STD", "")
            
            if std_clean in base_clean:
                return True
            if len(std_no_std) >= 4 and std_no_std in base_no_std:
                return True
        return False

    ref_groups = []
    other_groups = []
    for norm, g in groups.items():
        if is_reference_group(g["base"]):
            ref_groups.append((norm, g))
        else:
            other_groups.append((norm, g))

    # --- DETERMINE PRESENT REFERENCE MATERIALS ---
    present_refs_norm = {norm for norm, _ in ref_groups}
    has_iaea_603 = any('603' in norm for norm in present_refs_norm)

    # --- DETERMINE PRESENT REFERENCE MATERIALS ---
    present_refs_norm = {norm for norm, _ in ref_groups}
    has_iaea_603 = 'iaea603' in present_refs_norm
            
    # Update Data Start Row
    current_row = header_row + 1
    n_arag_re = re.compile(r"\bn\.?\s*arag\b", flags=re.IGNORECASE)

    def _write_sample_output_cells(ws, r, is_arag, has_iaea_603, new_R_col, new_U_col, output_format, slope_info):
        num_groups = len(slope_info)
        if num_groups == 0:
            return 27 # fallback
        
        # --- Mirror the spacing built in draw_lower_boxes ---
        c_start = 19
        o_calc_start = c_start + num_groups + 1
        o_arag_start = o_calc_start + num_groups + 1
        
        box1_start = c_start
        box1_end = o_arag_start + num_groups - 1
        
        box2_start = box1_end + 2
        vsmow_calc_start = box2_start
        vsmow_arag_start = vsmow_calc_start + num_groups + 1
        
        col_AA_approx = vsmow_arag_start + num_groups - 1
        
        green_bold = Font(color="008000", bold=True)
        
        for i in range(num_groups):
            s_row = slope_info[i]["slope_row"]
            i_row = slope_info[i]["intercept_row"]
            
            # Dynamic Calibration Formulas
            c_formula = f'=IFERROR($K${s_row}*{get_column_letter(new_R_col)}{r}+$K${i_row},"")'
            o_formula = f'=IFERROR($N${s_row}*{get_column_letter(new_U_col)}{r}+$N${i_row},"")'
            
            # Columns Mapping
            c_col = c_start + i
            o_calc_col = o_calc_start + i
            o_arag_col = o_arag_start + i
            vsmow_calc_col = vsmow_calc_start + i
            vsmow_arag_col = vsmow_arag_start + i
            
            # 1. Write Delta 13C
            cell_c = ws.cell(row=r, column=c_col, value=c_formula)
            cell_c.number_format = output_format
            
            # 2. Write Delta 18O Based on Material Status
            if is_arag:
                ws.cell(row=r, column=o_calc_col, value=None)
                ws.cell(row=r, column=vsmow_calc_col, value=None)
                
                # Write Aragonite Values
                cell_o_arag = ws.cell(row=r, column=o_arag_col, value=o_formula)
                cell_o_arag.font = Font(bold=True)
                cell_o_arag.number_format = output_format
                
                vsmow_formula = f'=IFERROR((1.03092*{get_column_letter(o_arag_col)}{r})+30.92,"")'
                cell_vsmow_arag = ws.cell(row=r, column=vsmow_arag_col, value=vsmow_formula)
                cell_vsmow_arag.font = Font(bold=True)
                cell_vsmow_arag.number_format = output_format
                
                # Apply green Aragonite formatting
                for col_idx in (c_col, o_arag_col, vsmow_arag_col):
                    ws.cell(row=r, column=col_idx).font = green_bold
            else:
                ws.cell(row=r, column=o_arag_col, value=None)
                ws.cell(row=r, column=vsmow_arag_col, value=None)
                
                # Write Calcite Values
                cell_o_calc = ws.cell(row=r, column=o_calc_col, value=o_formula)
                cell_o_calc.number_format = output_format
                
                vsmow_formula = f'=IFERROR((1.03092*{get_column_letter(o_calc_col)}{r})+30.92,"")'
                cell_vsmow_calc = ws.cell(row=r, column=vsmow_calc_col, value=vsmow_formula)
                cell_vsmow_calc.font = Font(bold=True)
                cell_vsmow_calc.number_format = output_format
                
                # Apply standard bolding to populated entries
                for col_idx in (c_col, o_calc_col, vsmow_calc_col):
                    original_font = copy(ws.cell(row=r, column=col_idx).font)
                    original_font.bold = True
                    ws.cell(row=r, column=col_idx).font = original_font
                    
        return col_AA_approx

    # 🔴 UPDATED WRITE_GROUP FUNCTION (Outliers, Strikethrough, Dual Blocks)
    def write_group(norm, g, is_reference=True, has_iaea_603=False):
        nonlocal current_row
        base_name = _normalize_text(g["base"])
        rows_data = g["rows"]
        font_color = color_fonts.get(base_name)
        
        # CO2 Logic
        row_values_list = [item[1] for item in rows_data]
        valid_indices = []
        if base_name in ("co2", "heco2"):
            valid_indices = _get_valid_co2_rows(row_values_list, col_identifier1)
            
        row_map = [] # Track excel row numbers for calculation
        
        # --- 1. WRITE DATA ROWS ---
        c_vals = []; o_vals = [] # For python calculation of outliers
        valid_run_row_indices = [] # Indices of rows in 'rows_data' that are NOT skipped by CO2 logic

        for i, (source_row_idx, row_values_filtered, source_cell_formats) in enumerate(rows_data):
            excel_row = current_row
            
            # Write Cells
            for dest_col_idx in range(1, MAX_GROUP_COL + 1):
                src_col = next(s for s, d in col_map.items() if d == dest_col_idx)
                val = row_values_filtered[dest_col_idx - 1]
                dest_cell = ws_group.cell(row=excel_row, column=dest_col_idx, value=val)
                
                original_format = source_cell_formats.get(src_col, {})
                if original_format.get('format'): dest_cell.number_format = original_format['format']
                if original_format.get('font'): dest_cell.font = original_format['font']
                if original_format.get('fill'): dest_cell.fill = original_format['fill']
                if original_format.get('alignment'): dest_cell.alignment = original_format['alignment']
                if original_format.get('border'): dest_cell.border = original_format['border']

                if font_color and dest_col_idx == col_identifier1:
                    new_font = copy(dest_cell.font)
                    new_font.color = font_color.color
                    dest_cell.font = new_font
                    
            if base_name in ("co2", "heco2") and i in valid_indices:
                for col in range(1, MAX_GROUP_COL + 1):
                    ws_group.cell(row=excel_row, column=col).fill = gray_fill
                    
            if not is_reference:
                ident_val = str(ws_group.cell(row=excel_row, column=col_identifier1).value or "")
                is_arag = n_arag_re.search(ident_val)
                # Note: We pass slope_info here
                _write_sample_output_cells(ws_group, excel_row, is_arag, has_iaea_603, 11, 14, '0.00', slope_info)

            # Collect Value for Outlier Calc
            should_process = True
            if base_name in ("co2", "heco2") and i not in valid_indices:
                should_process = False
                c_vals.append(None); o_vals.append(None)
            
            if should_process:
                valid_run_row_indices.append(i)
                # Read back from sheet to get numeric values
                cv = row_values_filtered[11-1] # K is 11, index 10
                ov = row_values_filtered[14-1] # N is 14, index 13
                # Convert to float
                try: cv = float(cv)
                except: cv = None
                try: ov = float(ov)
                except: ov = None
                c_vals.append(cv); o_vals.append(ov)

            row_map.append(excel_row)
            current_row += 1
            
        # --- 2. OUTLIER CALCULATION ---
        # Only if we have enough data (References or Multi-run Samples)
        if is_reference or len(rows_data) > 1:
            valid_c_nums = [v for v in c_vals if v is not None]
            valid_o_nums = [v for v in o_vals if v is not None]

            mean_c = statistics.mean(valid_c_nums) if len(valid_c_nums) > 1 else 0
            stdev_c = statistics.stdev(valid_c_nums) if len(valid_c_nums) > 1 else 0
            mean_o = statistics.mean(valid_o_nums) if len(valid_o_nums) > 1 else 0
            stdev_o = statistics.stdev(valid_o_nums) if len(valid_o_nums) > 1 else 0

            c_up, c_low = mean_c + (outlier_sigma * stdev_c), mean_c - (outlier_sigma * stdev_c)
            o_up, o_low = mean_o + (outlier_sigma * stdev_o), mean_o - (outlier_sigma * stdev_o)

            all_runs_c = []; all_runs_o = []
            final_runs_c = []; final_runs_o = []

            c_col_let = "K"; o_col_let = "N"

            for i in valid_run_row_indices:
                r_num = row_map[i]
                vc = c_vals[i]; vo = o_vals[i]

                if vc is not None: all_runs_c.append(f"{c_col_let}{r_num}")
                if vo is not None: all_runs_o.append(f"{o_col_let}{r_num}")

                # Check Outlier
                is_c_out = (vc > c_up or vc < c_low) if (vc is not None and len(valid_c_nums) > 2) else False
                is_o_out = (vo > o_up or vo < o_low) if (vo is not None and len(valid_o_nums) > 2) else False

                exclude_c = False; exclude_o = False
                if exclusion_mode == "Exclude Row":
                    if is_c_out or is_o_out: exclude_c = True; exclude_o = True
                else:
                    if is_c_out: exclude_c = True
                    if is_o_out: exclude_o = True
                
                if not exclude_c and vc is not None: final_runs_c.append(f"{c_col_let}{r_num}")
                if not exclude_o and vo is not None: final_runs_o.append(f"{o_col_let}{r_num}")
                
                # Apply Strikethrough
                if exclude_c and vc is not None: ws_group.cell(row=r_num, column=11).font = strike_font
                if exclude_o and vo is not None: ws_group.cell(row=r_num, column=14).font = strike_font
        
            # --- 3. WRITE DUAL STATS BLOCKS ---
            # Define Summary Font
            summary_font = font_color or Font(bold=True)
            if is_reference and font_color:
                summary_font = copy(font_color); summary_font.bold = True
            
            data_fmt = get_summary_num_format(base_name); count_fmt = '0'

            # --- BLOCK A: ALL VALUES ---
            row_all = current_row
            ws_group.cell(row=row_all, column=10, value="--").font = Font(bold=True) # J
            for col, txt in zip([11, 12, 13, 14, 15, 16], ["Average", "Stdev", "Count", "Average", "Stdev", "Count"]):
                c = ws_group.cell(row=row_all, column=col, value=txt)
                c.font = summary_font; c.alignment = Alignment(horizontal="right")

            row_all_calc = row_all + 1
            if all_runs_c:
                rng_c = ",".join(all_runs_c)
                ws_group.cell(row=row_all_calc, column=11, value=f"=AVERAGE({rng_c})").number_format = data_fmt
                ws_group.cell(row=row_all_calc, column=12, value=f"=STDEV({rng_c})").number_format = data_fmt
                ws_group.cell(row=row_all_calc, column=13, value=f"=COUNT({rng_c})").number_format = count_fmt
            if all_runs_o:
                rng_o = ",".join(all_runs_o)
                ws_group.cell(row=row_all_calc, column=14, value=f"=AVERAGE({rng_o})").number_format = data_fmt
                ws_group.cell(row=row_all_calc, column=15, value=f"=STDEV({rng_o})").number_format = data_fmt
                ws_group.cell(row=row_all_calc, column=16, value=f"=COUNT({rng_o})").number_format = count_fmt
            
            # Apply Style to Calc Row
            for c in range(11, 17): ws_group.cell(row=row_all_calc, column=c).font = summary_font

            # Calc Normalization for All Block (If not reference)
            if not is_reference:
                is_arag_group = n_arag_re.search(g["base"])
                # Pass slope_info
                _write_sample_output_cells(ws_group, row_all_calc, is_arag_group, has_iaea_603, 11, 14, '0.00', slope_info)
                # Bold Output
                for c in range(11, 28): 
                    if ws_group.cell(row=row_all_calc, column=c).value: 
                        ws_group.cell(row=row_all_calc, column=c).font = Font(bold=True)

            # --- BLOCK B: OUTLIER EXCLUDED ---
            row_filt = row_all_calc + 1
            ws_group.cell(row=row_filt, column=10, value="Outlier Excl.").font = Font(bold=True) # J
            for col, txt in zip([11, 12, 13, 14, 15, 16], ["Average", "Stdev", "Count", "Average", "Stdev", "Count"]):
                c = ws_group.cell(row=row_filt, column=col, value=txt)
                c.font = summary_font; c.alignment = Alignment(horizontal="right")

            row_filt_calc = row_filt + 1
            if final_runs_c:
                rng_c = ",".join(final_runs_c)
                ws_group.cell(row=row_filt_calc, column=11, value=f"=AVERAGE({rng_c})").number_format = data_fmt
                ws_group.cell(row=row_filt_calc, column=12, value=f"=STDEV({rng_c})").number_format = data_fmt
                ws_group.cell(row=row_filt_calc, column=13, value=f"=COUNT({rng_c})").number_format = count_fmt
            if final_runs_o:
                rng_o = ",".join(final_runs_o)
                ws_group.cell(row=row_filt_calc, column=14, value=f"=AVERAGE({rng_o})").number_format = data_fmt
                ws_group.cell(row=row_filt_calc, column=15, value=f"=STDEV({rng_o})").number_format = data_fmt
                ws_group.cell(row=row_filt_calc, column=16, value=f"=COUNT({rng_o})").number_format = count_fmt

            # Apply Style to Calc Row
            for c in range(11, 17): ws_group.cell(row=row_filt_calc, column=c).font = summary_font

            # Calc Normalization for Filtered Block (If not reference)
            if not is_reference:
                is_arag_group = n_arag_re.search(g["base"])
                # Pass slope_info
                _write_sample_output_cells(ws_group, row_filt_calc, is_arag_group, has_iaea_603, 11, 14, '0.00', slope_info)
                # Bold Output
                for c in range(11, 28): 
                    if ws_group.cell(row=row_filt_calc, column=c).value: 
                        ws_group.cell(row=row_filt_calc, column=c).font = Font(bold=True)

            current_row = row_filt_calc + 2 # Space for next group
        else:
            current_row += 3
            
    # Write reference groups first
    for norm, g in ref_groups:
        write_group(norm, g, is_reference=True)
        
    # Divider
    if ref_groups:
        current_row += 8
        divider_top_row = current_row # store divider start row
        for _ in range(2):
            for col in range(1, 702):
                ws_group.cell(row=current_row, column=col).fill = dark_fill
            current_row += 1
            
        # Copy headers again below the divider, respecting the exclusion list AND copying styles (Row 18)
        current_header_col = 1
        for src_col, dest_col_idx in col_map.items():
            src_cell = first_row_cells[src_col - 1] if src_col <= len(first_row_cells) else None
            
            dest_cell = ws_group.cell(row=current_row, column=current_header_col, value=headers[current_header_col - 1])
            
            # Copy style properties for header row
            if src_cell:
                dest_cell.number_format = copy(src_cell.number_format)
                dest_cell.font = copy(src_cell.font)
                dest_cell.alignment = copy(src_cell.alignment)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)

            current_header_col += 1

        current_row += 1
        
        # Call the updated lower box drawing function with new column indices
        draw_lower_boxes(ws_group, divider_top_row, blue_fill, Font(bold=True, color="000000"), Font(bold=True, color="008000"))
        
    # Write non-reference groups
    # Pass the result of the reference check to the function call
    for norm, g in other_groups:
        write_group(norm, g, is_reference=False, has_iaea_603=has_iaea_603)
        
    populate_blue_box_math(ws_group, slope_info, mat_row_map)

    # --- FILL MEASURED COLUMNS WITH GRAY BACKGROUND (Dynamic Start) ---
    start_gray_row = header_row + 1 
    max_sheet_row = ws_group.max_row + 50
    
    for row in range(start_gray_row, max_sheet_row + 1):
        for col in (11, 14): # Measured C (K) and Measured O (N)
            ws_group.cell(row=row, column=col).fill = gray_fill
            
    # Set column widths
    ws_group.column_dimensions["A"].width = 13 
    ws_group.column_dimensions["C"].width = 22 
    ws_group.column_dimensions["H"].width = 15 
    ws_group.column_dimensions["J"].width = 16 
    
    # Conditional Formatting (Dynamic Start)
    max_data_row = ws_group.max_row
    threshold_str = str(stdev_threshold)
    
    ws_group.conditional_formatting.add(
        f"L{start_gray_row}:L{max_data_row}",
        FormulaRule(formula=[f'=AND(ISNUMBER(L{start_gray_row}), L{start_gray_row} > {threshold_str})'], fill=red_fill)
    )

    ws_group.conditional_formatting.add(
        f"O{start_gray_row}:O{max_data_row}",
        FormulaRule(formula=[f'=AND(ISNUMBER(O{start_gray_row}), O{start_gray_row} > {threshold_str})'], fill=red_fill)
    )

    ws_group.freeze_panes = f'B{start_gray_row}'

    # Add Settings Popup Comment
    embed_settings_popup(ws_group, "A2")

    wb.save(file_path)
    print(f"✅ Step 6: Normalization completed on {file_path}")
