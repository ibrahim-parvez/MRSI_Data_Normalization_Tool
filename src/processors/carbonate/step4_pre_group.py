import pandas as pd
import re
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule 
import utils.settings as settings
from utils.common_utils import embed_settings_popup

# --- Helper Functions ---

def get_base_key(identifier):
    """
    Strips trailing run numbers (e.g., 'Sample r1', 'Sample r1.2') 
    to get the base name for grouping.
    """
    if identifier is None:
        return None
    s_id = str(identifier)
    # Remove trailing ' r1', ' r1.1', etc.
    return re.sub(r'\s+[rR]\d+(?:\.\d+)*(?:[a-zA-Z]*)?$', '', s_id).strip()

def copy_cell_style(src_cell, tgt_cell):
    """Copies Font, Border, Fill, Alignment, and NumberFormat."""
    if src_cell is None or tgt_cell is None: return
    if src_cell.has_style:
        try:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = src_cell.number_format 
            tgt_cell.alignment = copy(src_cell.alignment)
            tgt_cell.protection = copy(src_cell.protection)
        except: pass

def step4_pre_group_carbonate(file_path: str):
    """
    Step 4: Pre-Group Carbonate
    - Linear Grouping (groups consecutively listed items).
    - Removes Blue Box/Normalization math.
    - Preserves Source Styling and Carbonate Column Mapping.
    - Adds Summary Stats (Avg/Stdev/Count) for C and O.
    - Conditional Formatting applied to entire columns L and O (ignoring text).
    """

    new_sheet_name = "Pre-Group_DNT"
    source_sheet_name = "Last 6_DNT"

    # --- 1. Settings & Styles ---
    # Get Threshold
    # --- Configuration from Settings ---
    stdev_is_enabled = settings.get_setting("STDEV_THRESHOLD_ENABLED")
    
    # If disabled, set the variable to None so it bypasses conditional formatting
    if stdev_is_enabled:
        stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    else:
        stdev_threshold = None
    
    # Fills
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red for high Stdev
    gray_fill = PatternFill(start_color="E7E7E7", end_color="E7E7E7", fill_type="solid")  # Gray for data cols
    header_fill = PatternFill(start_color="8ED973", end_color="8ED973", fill_type="solid") # Green for headers
    bold_font = Font(bold=True)

    # Number Format
    THREE_DECIMAL_FORMAT = "0.000"

    # --- 2. Load Workbook ---
    wb_values = load_workbook(file_path, data_only=True) # To read calculated values
    wb = load_workbook(file_path) # To read styles

    # Find source sheet
    matched_source = next((s for s in wb.sheetnames if s.lower() == source_sheet_name.lower()), None)
    if matched_source is None:
        print(f"❌ Source sheet matching '{source_sheet_name}' not found.")
        return

    source_ws_vals = wb_values[matched_source]
    source_ws = wb[matched_source]

    # Delete existing Pre-Group if exists
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    # Create new sheet to the left of Last 6
    idx = wb.index(source_ws)
    new_ws = wb.create_sheet(new_sheet_name, idx)

    # --- 3. Define Column Mapping (Carbonate Specific) ---
    # We use the Exclusion List logic from the original Group code to map columns.
    # Excluded Source Cols: 8, 9, 11, 12, 13, 26, 27, 14, 15
    EXCLUDED_COLS = {8, 9, 11, 12, 13, 26, 27, 14, 15}
    MAX_SOURCE_COL = 24
    
    # source_col_index -> dest_col_index
    col_map = {}
    dest_col_counter = 1
    source_cols_ordered = [] # Keep track for the loop

    for src_c in range(1, MAX_SOURCE_COL + 1):
        if src_c not in EXCLUDED_COLS:
            col_map[src_c] = dest_col_counter
            source_cols_ordered.append(src_c)
            dest_col_counter += 1
            
    # Key Columns in Destination (for Summaries)
    # Based on the mapping:
    # Col 11 (K): Measured C -> Stdev will be 12 (L)
    # Col 14 (N): Measured O -> Stdev will be 15 (O)
    col_C_meas = 11
    col_O_meas = 14
    
    # Identifier Column (Source col 3 usually maps to Dest col 3)
    col_id_dest = col_map.get(3, 3)

    # --- 4. Setup Headers ---
    # Copy headers from Source Row 1
    src_header_row = list(source_ws[1])
    
    for src_c, dest_c in col_map.items():
        if src_c <= len(src_header_row):
            # Get header value and style
            src_cell = source_ws.cell(row=1, column=src_c)
            dest_cell = new_ws.cell(row=1, column=dest_c, value=src_cell.value)
            
            # Apply Style or Default Green Header
            copy_cell_style(src_cell, dest_cell)
            # Ensure it has the standard header fill if the source didn't have one
            if not src_cell.fill or src_cell.fill.patternType is None:
                dest_cell.fill = header_fill

    # Set Column Widths (Approximate based on Carbonate needs)
    new_ws.column_dimensions[get_column_letter(3)].width = 22 # Identifier
    new_ws.column_dimensions[get_column_letter(11)].width = 12 # C Meas
    new_ws.column_dimensions[get_column_letter(14)].width = 12 # O Meas
    new_ws.column_dimensions[get_column_letter(10)].width = 16 # Q (Last 6 Column)

    new_ws.freeze_panes = "A2"

    # --- 5. Linear Processing Loop ---
    
    src_row = 2
    dest_row = 2 # Start writing data immediately at Row 2
    max_row = source_ws.max_row
    
    group_start_dest_row = dest_row 

    while src_row <= max_row:
        # Check if source row is hidden
        if source_ws.row_dimensions[src_row].hidden:
            src_row += 1
            continue

        # A. Read Identifier
        # Source Col 3 is Identifier
        id_val = source_ws_vals.cell(row=src_row, column=3).value
        current_base = get_base_key(id_val)

        # B. Copy Row to Destination
        for src_c in source_cols_ordered:
            dest_c = col_map[src_c]
            
            # Value from values-wb
            val = source_ws_vals.cell(row=src_row, column=src_c).value
            dest_cell = new_ws.cell(row=dest_row, column=dest_c, value=val)
            
            # Style from style-wb
            src_cell = source_ws.cell(row=src_row, column=src_c)
            copy_cell_style(src_cell, dest_cell)
            
        # Apply Gray Fill to Measured Columns (K/11 and N/14) for Data Rows
        new_ws.cell(row=dest_row, column=col_C_meas).fill = gray_fill
        new_ws.cell(row=dest_row, column=col_O_meas).fill = gray_fill

        # Copy Row Height
        try:
            rh = source_ws.row_dimensions[src_row].height
            if rh is not None:
                new_ws.row_dimensions[dest_row].height = rh
        except: pass

        # C. Look Ahead (Find next visible row)
        next_src_row = src_row + 1
        next_base = None
        
        while next_src_row <= max_row:
            if not source_ws.row_dimensions[next_src_row].hidden:
                next_val = source_ws_vals.cell(row=next_src_row, column=3).value
                next_base = get_base_key(next_val)
                break
            next_src_row += 1

        # D. Check Group End
        if current_base != next_base:
            # Group Finished. Insert Summary Block.
            group_end_dest_row = dest_row
            
            # Determine rows
            label_row = dest_row + 1
            calc_row = dest_row + 2
            
            # Define Columns for Carbon and Oxygen stats
            # Carbon: K(11)=Avg, L(12)=Stdev, M(13)=Count
            c_avg_let = get_column_letter(col_C_meas)     # K
            c_stdev_let = get_column_letter(col_C_meas+1) # L
            c_count_let = get_column_letter(col_C_meas+2) # M
            
            # Oxygen: N(14)=Avg, O(15)=Stdev, P(16)=Count
            o_avg_let = get_column_letter(col_O_meas)     # N
            o_stdev_let = get_column_letter(col_O_meas+1) # O
            o_count_let = get_column_letter(col_O_meas+2) # P
            
            # 1. Write Labels
            labels = ["Average", "Stdev", "Count"]
            
            # Carbon Labels
            for i, txt in enumerate(labels):
                cell = new_ws.cell(row=label_row, column=col_C_meas + i, value=txt)
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='right')
            
            # Oxygen Labels
            for i, txt in enumerate(labels):
                cell = new_ws.cell(row=label_row, column=col_O_meas + i, value=txt)
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='right')

            # 2. Write Formulas
            # Carbon
            rng_c = f"{c_avg_let}{group_start_dest_row}:{c_avg_let}{group_end_dest_row}"
            new_ws.cell(row=calc_row, column=col_C_meas, value=f"=AVERAGE({rng_c})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=calc_row, column=col_C_meas+1, value=f"=STDEV({rng_c})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=calc_row, column=col_C_meas+2, value=f"=COUNT({rng_c})").number_format = "0"
            
            # Oxygen
            rng_o = f"{o_avg_let}{group_start_dest_row}:{o_avg_let}{group_end_dest_row}"
            new_ws.cell(row=calc_row, column=col_O_meas, value=f"=AVERAGE({rng_o})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=calc_row, column=col_O_meas+1, value=f"=STDEV({rng_o})").number_format = THREE_DECIMAL_FORMAT
            new_ws.cell(row=calc_row, column=col_O_meas+2, value=f"=COUNT({rng_o})").number_format = "0"
            
            # Bold the calc row
            for c in range(col_C_meas, col_O_meas + 3):
                new_ws.cell(row=calc_row, column=c).font = bold_font

            # Update pointers
            # Current Dest is Data. Next is Label (+1). Next is Calc (+2). Next is Blank (+3). 
            # Next Data starts at +4.
            dest_row = calc_row + 2 
            group_start_dest_row = dest_row

        else:
            # Same group, just move down
            dest_row += 1

        src_row += 1

    # --- 6. Finalize & Conditional Formatting ---
    for s in wb.worksheets: s.sheet_view.tabSelected = False
    new_ws.sheet_view.tabSelected = True
    wb.active = wb.index(new_ws)
    new_ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    
    # --- Dynamic Conditional Formatting (Apply to Entire Column) ---
    # Formula logic: =AND(ISNUMBER(Cell), Cell > Threshold)
    # This prevents formatting text headers (like 'Average', 'Stdev') or empty strings.
    
    final_max_row = new_ws.max_row
    
    if stdev_threshold is not None:
        thresh_str = str(stdev_threshold)
        
        # Column L: Carbon Stdev (Col 12)
        # Apply from L2 to L<max>
        rule_L = FormulaRule(
            formula=[f'AND(ISNUMBER(L2), L2 > {thresh_str})'],
            fill=fill_error
        )
        new_ws.conditional_formatting.add(f"L2:L{final_max_row}", rule_L)
        
        # Column O: Oxygen Stdev (Col 15)
        # Apply from O2 to O<max>
        rule_O = FormulaRule(
            formula=[f'AND(ISNUMBER(O2), O2 > {thresh_str})'],
            fill=fill_error
        )
        new_ws.conditional_formatting.add(f"O2:O{final_max_row}", rule_O)

    # Add Settings Popup Comment
    embed_settings_popup(new_ws, "R1")

    wb.save(file_path)
    print(f"✅ Step 4: Carbonate Pre-Group sheet '{new_sheet_name}' created.")