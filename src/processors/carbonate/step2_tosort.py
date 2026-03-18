import os
import time
import traceback
from copy import copy
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection
from openpyxl.utils import get_column_letter
from utils.common_utils import embed_settings_popup

def _try_force_excel_recalc(file_path, timeout=5.0):
    """
    Try to open the workbook in Excel via xlwings, calculate, save and close.
    Returns True on success, False on failure (e.g. xlwings not installed or Excel not available).
    """
    try:
        import xlwings as xw
    except Exception:
        return False

    app = None
    try:
        app = xw.App(visible=False)
        # Give Excel a moment to start
        time.sleep(0.2)
        book = app.books.open(os.path.abspath(file_path))
        # Force a full recalculation
        book.app.calculate()
        # Wait a little for Excel to finish (best-effort)
        time.sleep(min(1.0, timeout))
        book.save()
        book.close()
        app.quit()
        return True
    except Exception:
        try:
            if app:
                app.quit()
        except Exception:
            pass
        return False


def step2_tosort_carbonate(file_path, filter_choice="Last 6"):
    """
    Step 2: TO SORT (refactored + style copying)

    - Copies rows from 'Data' into 'To Sort' but converts formulas into raw values in To Sort.
    - Values are taken from a data_only workbook so formulas do not get copied.
    - Formatting (fonts, fills, borders, alignment, number formats, merged cells,
      row heights, column widths) is copied from the original 'Data' worksheet.
    - Attempts to force Excel recalc (via xlwings) so cached values exist; if recalculation fails,
      the code will still copy whatever cached values exist (may be None for some formula cells).
    - Finally: applies autofilter on column Q and hides rows not matching filter (unless "All").

    DISPLAY RULES APPLIED:
    - Columns that display averages/stdevs (R,S,U,V) will be formatted with "0.000".
    - Sum area column X (24) will be formatted with "0.00".
    """

    source_sheet = "Data_DNT"
    new_sheet_name = "To Sort_DNT"

    # Try to force recalculation in Excel (best-effort)
    recalc_ok = _try_force_excel_recalc(file_path)
    if not recalc_ok:
        print("Warning: unable to force Excel recalculation (xlwings missing or failed).")
        print("If 'Data' contains formulas without cached values, 'To Sort' may have empty cells for those formulas.")

    # Load two workbook views:
    wb = load_workbook(file_path, data_only=False)
    wb_values = load_workbook(file_path, data_only=True)

    if source_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found in workbook. Run Step 1 first.")
    if source_sheet not in wb_values.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found in values workbook. Run Step 1 first.")

    # Remove old To Sort if present
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    ws_source = wb[source_sheet]                 
    ws_source_values = wb_values[source_sheet]   

    # Create To Sort sheet
    ws_new = wb.create_sheet(new_sheet_name, index=wb.index(ws_source))

    text_cols = {4, 5, 6}

    max_col_idx = ws_source.max_column or 0
    max_row_from_values = ws_source_values.max_row if (hasattr(ws_source_values, "max_row") and ws_source_values.max_row) else 0
    max_row_idx = max(max_row_from_values, ws_source.max_row or 0)

    if max_row_idx == 0 or max_col_idx == 0:
        ws_new.sheet_view.tabSelected = True
        wb.active = wb.index(ws_new)
        wb.save(file_path)
        wb.close()
        wb_values.close()
        print(f"Step 2: TO SORT created empty sheet '{new_sheet_name}' in {file_path}")
        return

    # Copy row heights
    for r in range(1, max_row_idx + 1):
        rd = ws_source.row_dimensions.get(r)
        if rd is not None and rd.height is not None:
            ws_new.row_dimensions[r].height = rd.height

    # Copy column widths
    for c in range(1, max_col_idx + 1):
        col_letter = get_column_letter(c)
        cd = ws_source.column_dimensions.get(col_letter)
        if cd is not None and cd.width is not None:
            ws_new.column_dimensions[col_letter].width = cd.width

    # Copy merged cells
    try:
        for merged in list(ws_source.merged_cells.ranges):
            try:
                ws_new.merge_cells(str(merged))
            except Exception:
                continue
    except Exception:
        pass

    # --- COPY CONDITIONAL FORMATTING RULES ---
    try:
        for cf_range in ws_source.conditional_formatting._cf_rules:
            rules = ws_source.conditional_formatting._cf_rules[cf_range]
            for rule in rules:
                new_rule = copy(rule)
                ws_new.conditional_formatting.add(cf_range, new_rule)
    except Exception as e:
        print("Warning: unable to copy conditional formatting rules:", e)

    # --- CORE: copy value from ws_source_values and style from ws_source ---
    for r in range(1, max_row_idx + 1):
        for c in range(1, max_col_idx + 1):
            new_cell = ws_new.cell(row=r, column=c)

            try:
                val = ws_source_values.cell(row=r, column=c).value
            except Exception:
                val = None

            if c in text_cols and val is not None:
                try:
                    val = str(val)
                except Exception:
                    pass
            new_cell.value = val

            try:
                src_cell = ws_source.cell(row=r, column=c)
                if hasattr(src_cell, "has_style") and src_cell.has_style:
                    if src_cell.font is not None:
                        new_cell.font = copy(src_cell.font)
                    if src_cell.fill is not None:
                        new_cell.fill = copy(src_cell.fill)
                    if src_cell.border is not None:
                        new_cell.border = copy(src_cell.border)
                    if src_cell.alignment is not None:
                        new_cell.alignment = copy(src_cell.alignment)
                    if src_cell.protection is not None:
                        new_cell.protection = copy(src_cell.protection)
                    try:
                        new_cell.number_format = src_cell.number_format
                    except Exception:
                        pass
            except Exception:
                continue

    last_row = max_row_idx

    # --- APPLY NUMBER FORMATTING OVERRIDES ---
    fmt_three = "0.000"
    fmt_two = "0.00"

    if last_row >= 2:
        for r in range(2, last_row + 1):
            try:
                if 18 <= max_col_idx: ws_new.cell(row=r, column=18).number_format = fmt_three
                if 19 <= max_col_idx: ws_new.cell(row=r, column=19).number_format = fmt_three
                if 21 <= max_col_idx: ws_new.cell(row=r, column=21).number_format = fmt_three
                if 22 <= max_col_idx: ws_new.cell(row=r, column=22).number_format = fmt_three
                if 24 <= max_col_idx: ws_new.cell(row=r, column=24).number_format = fmt_two
            except Exception:
                continue

    # Apply autofilter to keep dropdowns
    last_col_letter = get_column_letter(max_col_idx)
    try:
        ws_new.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    except Exception:
        pass

    filter_choice_norm = (filter_choice or "Last 6").strip().lower()

    # Hide rows not matching filter
    if filter_choice_norm != "all":
        for r in range(2, last_row + 1):
            try:
                val = ws_new.cell(row=r, column=17).value  
                if val is None:
                    ws_new.row_dimensions[r].hidden = True
                    continue
                try:
                    if str(val).strip().lower() != filter_choice_norm:
                        ws_new.row_dimensions[r].hidden = True
                    else:
                        ws_new.row_dimensions[r].hidden = False
                except Exception:
                    ws_new.row_dimensions[r].hidden = True
            except Exception:
                try:
                    ws_new.row_dimensions[r].hidden = True
                except Exception:
                    pass

    # Activate new sheet and set selection
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass
    try:
        ws_new.sheet_view.tabSelected = True
        wb.active = wb.index(ws_new)
        ws_new.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    except Exception:
        pass

    # Add Settings Popup Comment
    embed_settings_popup(ws_new, "AB1")
    ws_new.column_dimensions["Q"].width = 16

    # Save and close workbooks securely
    try:
        wb.save(file_path)
        wb.close()
        wb_values.close()
        print(f"✅ Step 2: To Sort completed on {file_path}")
        if not recalc_ok:
            print("Note: xlwings recalculation was not run. If To Sort contains blanks in R–AA,")
            print("open the workbook in Excel and save once (or enable auto-calc), then re-run Step 2.")
    except Exception as e:
        print("Error: failed to save workbook after creating 'To Sort' sheet.")
        traceback.print_exc()
        raise e