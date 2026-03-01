from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import utils.settings as settings
from utils.common_utils import embed_settings_popup

def _detect_decimal_places_from_format(fmt: str):
    """
    Inspect an Excel number format string and return the number of decimal
    digit placeholders after the decimal point if detectable (int), otherwise None.
    """
    if not fmt or not isinstance(fmt, str):
        return None
    first_section = fmt.split(';', 1)[0]
    if '.' not in first_section:
        return None
    after_dot = first_section.split('.', 1)[1]
    count = 0
    for ch in after_dot:
        if ch in ('0', '#'):
            count += 1
        else:
            break
    return count if count > 0 else None


def step3_last6_carbonate(file_path):
    """
    Step 3: LAST 6 — Fully Refactored
    Copies rows from 'To Sort' based on the calculation mode setting.
    - If "Last 6": Copies rows where col Q is "last 6".
    - If "Last 6 Outliers Excl.": Copies rows where col Q is "outliers excl.".
    Preserves all formatting.
    """

    source_sheet = "To Sort_DNT"
    new_sheet_name = "Last 6_DNT"

    # --- 1. Determine Filter Target based on Settings ---
    calc_mode = settings.get_setting("CALC_MODE_STEP3")
    if calc_mode == "Last 6 Outliers Excl.":
        filter_target = "last 6 outliers excl."
    else:
        filter_target = "last 6"
        
    print(f"   ℹ️ Carbonate Step 3 Mode: {calc_mode} -> Filtering for '{filter_target}'")

    wb = load_workbook(file_path, data_only=False)

    if source_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{source_sheet}' not found. Run Step 2 first.")

    # Remove old sheet if exists
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    ws_source = wb[source_sheet]

    # Create Last 6 sheet to the left
    ws_new = wb.create_sheet(new_sheet_name, index=wb.index(ws_source))

    max_row = ws_source.max_row or 1
    max_col = ws_source.max_column or 1

    # ------------------------------------------------------------
    # COPY HEADER (row 1, including formatting)
    # ------------------------------------------------------------
    header_map = {}
    for c in range(1, max_col + 1):
        src = ws_source.cell(row=1, column=c)
        tgt = ws_new.cell(row=1, column=c)
        tgt.value = src.value

        header_val = str(src.value).strip().lower() if src.value else ""
        header_map[header_val] = c

        if src.has_style:
            tgt.font = copy(src.font)
            tgt.fill = copy(src.fill)
            tgt.border = copy(src.border)
            tgt.alignment = copy(src.alignment)
            tgt.number_format = src.number_format
            tgt.protection = copy(src.protection)

    # Special text-enforced columns
    special_headers = {"comment", "identifier 2", "analysis"}
    special_cols = {idx for name, idx in header_map.items() if name in special_headers}

    # ------------------------------------------------------------
    # COPY ROW HEIGHTS, COLUMN WIDTHS
    # ------------------------------------------------------------
    for r in range(1, max_row + 1):
        src_dim = ws_source.row_dimensions.get(r)
        if src_dim and src_dim.height:
            ws_new.row_dimensions[r].height = src_dim.height

    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        src_dim = ws_source.column_dimensions.get(col_letter)
        if src_dim and src_dim.width:
            ws_new.column_dimensions[col_letter].width = src_dim.width

    # ------------------------------------------------------------
    # COPY MERGED CELLS
    # ------------------------------------------------------------
    try:
        for merged_range in ws_source.merged_cells.ranges:
            try:
                ws_new.merge_cells(str(merged_range))
            except Exception:
                pass
    except Exception:
        pass

    # ------------------------------------------------------------
    # COPY ONLY ROWS MATCHING THE FILTER TARGET
    # ------------------------------------------------------------
    col_q = 17
    new_row = 2
    mapping_new_to_src = {}

    for r in range(2, max_row + 1):
        v = ws_source.cell(row=r, column=col_q).value
        # Check against dynamic filter target
        if str(v).strip().lower() != filter_target:
            continue

        mapping_new_to_src[new_row] = r

        # Copy row values + formatting
        for c in range(1, max_col + 1):
            src = ws_source.cell(row=r, column=c)
            tgt = ws_new.cell(row=new_row, column=c)

            # value copy
            if c in special_cols and src.value is not None:
                tgt.value = str(src.value)
            else:
                tgt.value = src.value

            # formatting copy
            if src.has_style:
                tgt.font = copy(src.font)
                tgt.fill = copy(src.fill)
                tgt.border = copy(src.border)
                tgt.alignment = copy(src.alignment)
                tgt.number_format = src.number_format
                tgt.protection = copy(src.protection)

        new_row += 1

    last_row = ws_new.max_row
    last_col = ws_new.max_column

    # ------------------------------------------------------------
    # DECIMAL PLACE DISPLAY RULES (NO ROUNDING)
    # ------------------------------------------------------------
    fmt_three = "0.000"
    fmt_two = "0.00"

    for new_r, src_r in mapping_new_to_src.items():
        for c in range(1, last_col + 1):
            tgt = ws_new.cell(row=new_r, column=c)
            src = ws_source.cell(row=src_r, column=c)

            sval = src.value
            if sval is None or isinstance(sval, str):
                continue

            if isinstance(sval, (int, float)):
                # Detect decimal placeholders
                dec = _detect_decimal_places_from_format(src.number_format)

                if dec is None:
                    # Fallback check actual decimals in the stored float
                    s = f"{sval:.10f}".rstrip("0").rstrip(".")
                    if "." in s:
                        actual = len(s.split(".", 1)[1])
                        tgt.number_format = fmt_three if actual >= 3 else fmt_two
                else:
                    tgt.number_format = fmt_three if dec >= 3 else fmt_two

    # -------------------- CONDITIONAL FORMATTING --------------------
    threshold = settings.get_setting("STDEV_THRESHOLD")

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Column S (19) - Carbon Stdev
    ws_new.conditional_formatting.add(
        f"S2:S{ws_new.max_row}",
        CellIsRule(
            operator="greaterThan",
            formula=[str(threshold)],
            fill=red_fill
        )
    )

    # Column V (22) - Oxygen Stdev
    ws_new.conditional_formatting.add(
        f"V2:V{ws_new.max_row}",
        CellIsRule(
            operator="greaterThan",
            formula=[str(threshold)],
            fill=red_fill
        )
    )

    # ------------------------------------------------------------
    # SELECT A1, MAKE SHEET ACTIVE
    # ------------------------------------------------------------
    for s in wb.worksheets:
        try:
            s.sheet_view.tabSelected = False
        except Exception:
            pass

    ws_new.sheet_view.tabSelected = True
    wb.active = wb.index(ws_new)
    ws_new.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Add Settings Popup Comment
    embed_settings_popup(ws_new, "AB1")

    # Set column widths
    ws_new.column_dimensions["Q"].width = 16 

    wb.save(file_path)
    print(f"Step 3: LAST 6 completed on {file_path}")