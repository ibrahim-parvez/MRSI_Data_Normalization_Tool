import pandas as pd
import statistics
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, NamedStyle
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.views import Selection
import utils.settings as settings
from utils.common_utils import embed_settings_popup


def step1_data_carbonate(file_path, sheet_name='ExportGB2.wke'):
    """
    Step 1: DATA for Carbonate
    Reads the Excel file, generates padded data rows,
    summary metrics, formulas, and conditional formatting.
    
    FIXES:
    1. Robust Header Detection (Fixes KeyError: 'Row').
    2. Dynamic Formula Generation for 'Last 6 Outliers Excl.'.
    3. Visual Red Strikethrough matches calculation logic.
    """

    # --- Configuration from Settings ---
    stdev_threshold = settings.get_setting("STDEV_THRESHOLD")
    sigma_mult = settings.get_setting("OUTLIER_SIGMA") or 2
    exclusion_mode = settings.get_setting("OUTLIER_EXCLUSION_MODE") or "Individual"

    new_sheet_name = 'Data_DNT'

    # --- READ DATA WITH ROBUST HEADER FINDING ---
    try:
        # Read first few rows to locate the actual header row
        df_preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=10, engine='openpyxl')
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}")

    header_row_idx = None
    for idx, row in df_preview.iterrows():
        # Check if row contains key columns (case-insensitive)
        row_str = row.astype(str).str.lower().tolist()
        if 'row' in row_str or 'line' in row_str:
            header_row_idx = idx
            break
    
    if header_row_idx is None:
        header_row_idx = 0

    # Read full dataframe
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx, engine='openpyxl')

    # Normalize 'Row' column for grouping
    if 'Row' not in df.columns:
        if 'Line' in df.columns:
            df['Row'] = df['Line']
        elif 'Identifier 1' in df.columns:
            df['Row'] = df['Identifier 1']
        else:
            # Fallback if no grouping column found
            df['Row'] = range(1, len(df) + 1)

    # Data sheet headers:
    headers = [
        'Row', 'Time Code', 'Identifier 1', 'Comment', 'Identifier 2', 'Analysis',
        'Preparation', 'Peak Nr', 'Rt', 'Ampl 44', 'Area All',
        'd 13C/12C', 'd 18O/16O',
        '', '', '', '',  # spacer columns
        'C avg', 'C stdev', '', 'O avg', 'O stdev', '',
        'Sum area all', 'area peaks', 'funny peaks', 'min intensity'
    ]

    wb = load_workbook(file_path)
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]

    first_index = wb.index(wb[sheet_name])
    ws = wb.create_sheet(new_sheet_name, first_index)

    # Select tab
    for s in wb.worksheets:
        try: s.sheet_view.tabSelected = False
        except: pass
    ws.sheet_view.tabSelected = True
    wb.active = wb.index(ws)
    ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

    # Write headers
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    # --- COLOR HEADER ROW ---
    green_fill = PatternFill(start_color="8ed973", end_color="8ed973", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Columns needing yellow fill
    yellow_cols = ["H", "I", "K", "L", "M", "Z", "AA"]

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx)
        if col_letter in yellow_cols:
            cell.fill = yellow_fill
        else:
            cell.fill = green_fill

    cur_row = 3

    # Header → column index
    col_map = {h: i + 1 for i, h in enumerate(headers) if h}

    # Normalization helper
    def normalize_name(s): return ' '.join(str(s).split()).lower() if s else ''
    df_cols_norm = {normalize_name(c): c for c in df.columns}

    # Map headers to df columns
    header_to_dfcol = {}
    for h in headers:
        if not h:
            header_to_dfcol[h] = None
            continue
        nh = normalize_name(h)
        if nh in df_cols_norm:
            header_to_dfcol[h] = df_cols_norm[nh]
            continue
        # Fuzzy matching
        nh_join = nh.replace(' ', '')
        match = None
        for dc_norm, dc in df_cols_norm.items():
            if dc_norm.replace(' ', '') == nh_join:
                match = dc
                break
        if match:
            header_to_dfcol[h] = match
            continue
        for dc_norm, dc in df_cols_norm.items():
            if nh in dc_norm or dc_norm in nh:
                match = dc
                break
        header_to_dfcol[h] = match

    # Key column indices
    col_area = col_map.get('Area All')
    col_c = col_map.get('d 13C/12C')
    col_o = col_map.get('d 18O/16O')
    col_ampl = col_map.get('Ampl 44')
    col_funny = col_map.get('funny peaks')
    col_minint = col_map.get('min intensity')

    check_headers = ['Rt', 'Ampl 44', 'Area All', 'd 13C/12C', 'd 18O/16O']
    check_df_cols = [header_to_dfcol.get(h) for h in check_headers if header_to_dfcol.get(h)]

    col_letter_area = get_column_letter(col_area)
    col_letter_c = get_column_letter(col_c)
    col_letter_o = get_column_letter(col_o)
    col_letter_ampl = get_column_letter(col_ampl)

    # Summary row layout - Added "last 6 Outliers Excl."
    summary_layout = [
        ("ref avg", 0),
        ("all", 3),
        ("last 6", 0),
        ("last 6 outliers excl.", 0), # Added Row
        ("start", 1),
        ("end", 0),
        ("delta", 0),
    ]

    # Summary column offsets
    col_label = 17  # Q
    col_c_avg = col_label + 1  # R
    col_c_stdev = col_label + 2  # S
    col_o_avg = col_label + 4  # U
    col_o_stdev = col_label + 5  # V
    col_sum_area = col_label + 7  # X

    # Formatting
    fill_label = PatternFill(start_color="cdffcc", end_color="cdffcc", fill_type="solid")
    fill_funny_min = PatternFill(start_color="cdfeff", end_color="cdfeff", fill_type="solid")
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_outlier = Font(color="FF0000", strike=True)

    group_highlight_ranges = []
    group_spacer_rows = []
    all_delta_rows = []

    grouped = df.groupby('Row', sort=False)

    # PROCESS GROUPS
    for Row, group in grouped:

        # Blank line between groups
        if cur_row != 3:
            cur_row += 1

        first_data_row = cur_row
        row_count = max(11, len(group))

        # Pad group rows to 11
        padded_rows = []
        for i in range(row_count):
            if i < len(group):
                padded_rows.append(group.iloc[i].to_dict())
            else:
                base_vals = group.iloc[0].to_dict() if len(group) else {}
                new_blank = {col: None for col in df.columns}
                for nm in ["Row", "Time Code", "Identifier 1"]:
                    new_blank[nm] = base_vals.get(nm)
                new_blank["Peak Nr"] = i + 1
                padded_rows.append(new_blank)

        # Write data rows
        for row_dict in padded_rows:
            for h in headers:
                if not h or h == "Sum area all":
                    continue
                excel_col = col_map[h]
                source_col = header_to_dfcol.get(h)
                val = row_dict.get(source_col) if source_col else None
                cell = ws.cell(row=cur_row, column=excel_col, value=val)
                if h in ("Identifier 2", "Analysis") and val is not None:
                    cell.number_format = '@'
            cur_row += 1

        last_data_row = cur_row - 1
        last7_start = max(first_data_row, last_data_row - 6)
        
        # --- FIXED ROW INDEX: Must match for Formula AND Formatting ---
        last6_start = max(first_data_row, last_data_row - 5)
        
        # Define Ranges for Formulas
        range_c_last6 = f"{col_letter_c}{last6_start}:{col_letter_c}{last_data_row}"
        range_o_last6 = f"{col_letter_o}{last6_start}:{col_letter_o}{last_data_row}"

        summary_row = first_data_row
        row_positions = {}

        data_count = group[check_df_cols].dropna(how='all').shape[0]

        # --- DYNAMIC FORMULA CONSTRUCTION LOGIC ---
        # 1. Gather Data for analysis
        c_values = []
        o_values = []
        row_indices = []
        
        # Determine absolute start row for Last 6
        start_list_idx = last6_start - first_data_row
        end_list_idx = last_data_row - first_data_row
        
        for i in range(start_list_idx, end_list_idx + 1):
            row_data = padded_rows[i]
            # Get numeric values safely
            try: 
                vc = float(row_data.get(header_to_dfcol['d 13C/12C']))
            except (ValueError, TypeError): vc = None
            
            try: 
                vo = float(row_data.get(header_to_dfcol['d 18O/16O']))
            except (ValueError, TypeError): vo = None
            
            c_values.append(vc)
            o_values.append(vo)
            row_indices.append(first_data_row + i) # Absolute Excel Row

        # 2. Calculate Stats for Bounds
        valid_c = [x for x in c_values if x is not None]
        valid_o = [x for x in o_values if x is not None]
        
        mean_c = statistics.mean(valid_c) if len(valid_c) > 0 else 0
        std_c = statistics.stdev(valid_c) if len(valid_c) > 1 else 0
        mean_o = statistics.mean(valid_o) if len(valid_o) > 0 else 0
        std_o = statistics.stdev(valid_o) if len(valid_o) > 1 else 0
        
        c_up = mean_c + (sigma_mult * std_c)
        c_low = mean_c - (sigma_mult * std_c)
        o_up = mean_o + (sigma_mult * std_o)
        o_low = mean_o - (sigma_mult * std_o)
        
        # 3. Build List of Valid Cells
        valid_c_cells = []
        valid_o_cells = []
        
        for idx, r_num in enumerate(row_indices):
            vc = c_values[idx]
            vo = o_values[idx]
            
            is_c_out = False
            is_o_out = False
            
            if vc is not None:
                if vc > c_up or vc < c_low: is_c_out = True
            if vo is not None:
                if vo > o_up or vo < o_low: is_o_out = True
                
            if exclusion_mode == "Exclude Row":
                # If EITHER is bad, EXCLUDE BOTH from list
                if not is_c_out and not is_o_out:
                    if vc is not None: valid_c_cells.append(f"{col_letter_c}{r_num}")
                    if vo is not None: valid_o_cells.append(f"{col_letter_o}{r_num}")
            else:
                # Individual Logic
                if not is_c_out and vc is not None: valid_c_cells.append(f"{col_letter_c}{r_num}")
                if not is_o_out and vo is not None: valid_o_cells.append(f"{col_letter_o}{r_num}")

        # 4. Construct Formula Strings
        if valid_c_cells:
            c_range_str = ",".join(valid_c_cells)
            f_c_avg = f"=AVERAGE({c_range_str})"
            f_c_std = f"=STDEV({c_range_str})" if len(valid_c_cells) > 1 else "0"
        else:
            f_c_avg = "No Data"
            f_c_std = "0"
            
        if valid_o_cells:
            o_range_str = ",".join(valid_o_cells)
            f_o_avg = f"=AVERAGE({o_range_str})"
            f_o_std = f"=STDEV({o_range_str})" if len(valid_o_cells) > 1 else "0"
        else:
            f_o_avg = "No Data"
            f_o_std = "0"

        # SUMMARY FORMULAS WRITING
        for label, spacing in summary_layout:
            summary_row += spacing
            row_positions[label] = summary_row

            ws.cell(summary_row, col_label, value=label)

            if label == "last 6":
                target_col = col_c_stdev + 1
                flag_msg = ""
                if data_count < 11:
                    flag_msg = f"< 11: {data_count}"
                elif data_count > 11:
                    flag_msg = f"> 11: {data_count}"
                
                if flag_msg:
                    c = ws.cell(summary_row, target_col, value=flag_msg)
                    c.fill = fill_error
                    c.font = Font(bold=True)
                    
                    # DUPLICATE FLAG TO "last 6 Outliers Excl." ROW
                    outlier_row_idx = summary_row + 1 # Logic from layout (+1 row down)
                    # We write it later when loop hits that label, or store it now
                    # Safer to just write it when we process 'last 6 Outliers Excl.'

            # --- ref avg ---
            if label == "ref avg":
                idx1, idx2, idx4 = first_data_row, first_data_row + 1, first_data_row + 3
                ws.cell(summary_row, col_c_avg, value=f"=AVERAGE({col_letter_c}{idx1},{col_letter_c}{idx2},{col_letter_c}{idx4})").number_format = "0.000"
                ws.cell(summary_row, col_c_stdev, value=f"=STDEV({col_letter_c}{idx1},{col_letter_c}{idx2},{col_letter_c}{idx4})").number_format = "0.000"
                ws.cell(summary_row, col_o_avg, value=f"=AVERAGE({col_letter_o}{idx1},{col_letter_o}{idx2},{col_letter_o}{idx4})").number_format = "0.000"
                ws.cell(summary_row, col_o_stdev, value=f"=STDEV({col_letter_o}{idx1},{col_letter_o}{idx2},{col_letter_o}{idx4})").number_format = "0.000"

            # --- all ---
            elif label == "all":
                ws.cell(summary_row, col_c_avg, value=f"=AVERAGE({col_letter_c}{last7_start}:{col_letter_c}{last_data_row})").number_format = "0.000"
                ws.cell(summary_row, col_c_stdev, value=f"=STDEV({col_letter_c}{last7_start}:{col_letter_c}{last_data_row})").number_format = "0.000"
                ws.cell(summary_row, col_o_avg, value=f"=AVERAGE({col_letter_o}{last7_start}:{col_letter_o}{last_data_row})").number_format = "0.000"
                ws.cell(summary_row, col_o_stdev, value=f"=STDEV({col_letter_o}{last7_start}:{col_letter_o}{last_data_row})").number_format = "0.000"
                ws.cell(summary_row, col_sum_area, value=f"=SUM({col_letter_area}{last7_start}:{col_letter_area}{last_data_row})").number_format = "0.00"

            # --- last 6 ---
            elif label == "last 6":
                ws.cell(summary_row, col_c_avg, value=f"=AVERAGE({range_c_last6})").number_format = "0.000"
                ws.cell(summary_row, col_c_stdev, value=f"=STDEV({range_c_last6})").number_format = "0.000"
                ws.cell(summary_row, col_o_avg, value=f"=AVERAGE({range_o_last6})").number_format = "0.000"
                ws.cell(summary_row, col_o_stdev, value=f"=STDEV({range_o_last6})").number_format = "0.000"
                ws.cell(summary_row, col_sum_area, value=f"=SUM({col_letter_area}{last6_start}:{col_letter_area}{last_data_row})").number_format = "0.00"

            # --- last 6 Outliers Excl. ---
            elif label == "last 6 outliers excl.":
                ws.cell(summary_row, col_c_avg, value=f_c_avg).number_format = "0.000"
                ws.cell(summary_row, col_c_stdev, value=f_c_std).number_format = "0.000"
                ws.cell(summary_row, col_o_avg, value=f_o_avg).number_format = "0.000"
                ws.cell(summary_row, col_o_stdev, value=f_o_std).number_format = "0.000"
                
                # DUPLICATE SUM AREA
                ws.cell(summary_row, col_sum_area, value=f"=SUM({col_letter_area}{last6_start}:{col_letter_area}{last_data_row})").number_format = "0.00"
                
                # DUPLICATE FLAG (Column T)
                # T is col_c_stdev + 1 in Carbonate layout (R, S, -> T is next)
                # No wait, in Carbonate:
                # Q=Label, R=C_avg, S=C_stdev, T=Flag?
                # The headers define T as 'Flag' (Spacer) in the summary layout?
                # col_c_stdev is S (19). T is 20.
                target_col = col_c_stdev + 1
                flag_msg = ""
                if data_count < 11: flag_msg = f"< 11: {data_count}"
                elif data_count > 11: flag_msg = f"> 11: {data_count}"
                
                if flag_msg:
                    c = ws.cell(summary_row, target_col, value=flag_msg)
                    c.fill = fill_error
                    c.font = Font(bold=True)

            # --- start ---
            elif label == "start":
                ws.cell(summary_row, col_c_avg, value=f"={col_letter_c}{last6_start}").number_format = "0.000"
                ws.cell(summary_row, col_o_avg, value=f"={col_letter_o}{last6_start}").number_format = "0.000"

            # --- end ---
            elif label == "end":
                ws.cell(summary_row, col_c_avg, value=f"={col_letter_c}{last_data_row}").number_format = "0.000"
                ws.cell(summary_row, col_o_avg, value=f"={col_letter_o}{last_data_row - 1}").number_format = "0.000"

            # --- delta ---
            elif label == "delta":
                sr = row_positions["start"]
                er = row_positions["end"]
                ws.cell(summary_row, col_c_avg, value=f"={get_column_letter(col_c_avg)}{er}-{get_column_letter(col_c_avg)}{sr}").number_format = "0.000"
                ws.cell(summary_row, col_o_avg, value=f"={get_column_letter(col_o_avg)}{er}-{get_column_letter(col_o_avg)}{sr}").number_format = "0.000"

            summary_row += 1

        # Track areas for highlighting
        delta_row = row_positions["delta"]
        all_delta_rows.append(delta_row)
        highlight_end = max(last_data_row, delta_row)
        group_highlight_ranges.append((first_data_row, highlight_end))
        group_spacer_rows.append(last_data_row + 1)

        # Funny/min-intensity calculations
        if col_letter_ampl and col_funny and col_minint:
            for i in range(row_count):
                r = first_data_row + i
                if i < 4:
                    ws.cell(r, col_funny, value="ref")
                else:
                    ws.cell(r, col_funny, value=f'=IF({col_letter_ampl}{r}>{col_letter_ampl}{r+1},IF({col_letter_ampl}{r+1}<{col_letter_ampl}{r},"ok","check"),"check")')

                if i == 0:
                    ws.cell(r, col_minint, value="if 44<1000")
                elif i < 4:
                    ws.cell(r, col_minint, value="")
                else:
                    ws.cell(r, col_minint, value=f'=IF({col_letter_ampl}{r}<400,"check","ok")')


        # -------------------- CONDITIONAL FORMATTING --------------------
        last6_row = row_positions.get("last 6")
        
        # 1. Stdev Threshold (Orange Fill)
        if last6_row:
            for col in [col_c_stdev, col_o_stdev]:
                ws.conditional_formatting.add(
                    f"{get_column_letter(col)}{last6_row}",
                    CellIsRule(operator="greaterThan", formula=[str(stdev_threshold)], fill=fill_error)
                )
        
        # 2. Outlier Strikethrough (Red Text)
        if last6_row:
            # Absolute references for CF logic
            c_mean_abs = f"${get_column_letter(col_c_avg)}${last6_row}"
            c_stdev_abs = f"${get_column_letter(col_c_stdev)}${last6_row}"
            o_mean_abs = f"${get_column_letter(col_o_avg)}${last6_row}"
            o_stdev_abs = f"${get_column_letter(col_o_stdev)}${last6_row}"
            
            c_logic = f"OR({col_letter_c}{last6_start}>({c_mean_abs}+({sigma_mult}*{c_stdev_abs})), {col_letter_c}{last6_start}<({c_mean_abs}-({sigma_mult}*{c_stdev_abs})))"
            o_logic = f"OR({col_letter_o}{last6_start}>({o_mean_abs}+({sigma_mult}*{o_stdev_abs})), {col_letter_o}{last6_start}<({o_mean_abs}-({sigma_mult}*{o_stdev_abs})))"
            
            if exclusion_mode == "Exclude Row":
                final_c_logic = f"OR({c_logic}, {o_logic})"
                final_o_logic = f"OR({c_logic}, {o_logic})"
            else:
                final_c_logic = c_logic
                final_o_logic = o_logic
            
            # Apply Rules
            ws.conditional_formatting.add(range_c_last6, FormulaRule(formula=[final_c_logic], font=font_outlier))
            ws.conditional_formatting.add(range_o_last6, FormulaRule(formula=[final_o_logic], font=font_outlier))

    # Apply background fills
    max_row = ws.max_row
    for s, e in group_highlight_ranges:
        s = max(2, s)
        e = min(max_row, e)
        for r in range(s, e + 1):
            ws.cell(r, col_label).fill = fill_label
            ws.cell(r, col_funny).fill = fill_funny_min
            ws.cell(r, col_minint).fill = fill_funny_min

    # Clear spacer rows
    for spacer in group_spacer_rows:
        for c in (col_label, col_funny, col_minint):
            ws.cell(spacer, c).fill = PatternFill(fill_type=None)
        
    # Add Settings Popup Comment
    embed_settings_popup(ws, "AB1")
            
    # Set column width for Q (Calculations Label)
    ws.column_dimensions["Q"].width = 18

    wb.save(file_path)
    print(f"Step 1: Data completed on {file_path}")