import os
import shutil
import re
import time
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
import tempfile
from copy import copy
from datetime import datetime
import xlwings as xw
from openpyxl.comments import Comment
import sys
import subprocess

from PyQt6.QtCore import QThread, pyqtSignal
import utils.settings as settings

# ---- Import carbonate modules ----
from processors.carbonate.step1_data import step1_data_carbonate
from processors.carbonate.step2_tosort import step2_tosort_carbonate
from processors.carbonate.step3_last6 import step3_last6_carbonate
from processors.carbonate.step4_pre_group import step4_pre_group_carbonate
from processors.carbonate.step5_group import step5_group_carbonate
from processors.carbonate.step6_normalization import step6_normalization_carbonate
from processors.carbonate.step7_report import step7_report_carbonate

class CarbonateCombineWorker(QThread):
    log = pyqtSignal(str, str)
    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal()
    error = pyqtSignal(str)
    stopped_early = pyqtSignal()

    def __init__(self, params):
        super().__init__()
        self.params = params
        self._is_running = True

    def stop(self):
        self._is_running = False

    def copy_cell_exact(self, src_cell, tgt_cell):
        tgt_cell.value = src_cell.value
        
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = copy(src_cell.number_format)
            tgt_cell.alignment = copy(src_cell.alignment)
            
        if src_cell.comment:
            tgt_cell.comment = Comment(src_cell.comment.text, src_cell.comment.author)

    def get_base_reference_name(self, identifier, ref_settings):
        if not identifier: return None
        raw_text = str(identifier).upper().strip()
        text_clean = re.sub(r'[\s\-_]+', '', raw_text)
        text_no_std = text_clean.replace("STD", "")

        if "CO2" in text_clean or "HECO2" in text_clean or "HEC02" in text_clean or "C02" in text_clean:
            return "HeCO2"

        for std in ref_settings:
            std_name = std.get("col_c")
            if not std_name: continue
            
            std_clean = re.sub(r'[\s\-_]+', '', str(std_name).upper())
            std_no_std = std_clean.replace("STD", "")
            
            if std_clean in text_clean:
                return std_name
            if len(std_no_std) >= 4 and std_no_std in text_no_std:
                return std_name
        return None

    def parse_timestamp(self, ts_val):
        if isinstance(ts_val, datetime):
            return ts_val
        if isinstance(ts_val, str):
            try:
                match = re.search(r'(\d{4}[-/]\d{2}[-/]\d{2}\s+\d{2}:\d{2}:\d{2})', ts_val)
                if match:
                    return datetime.strptime(match.group(1).replace('-', '/'), "%Y/%m/%d %H:%M:%S")
            except:
                pass
        return datetime.min

    def run(self):
        temp_dir = None
        app = None
        try:
            mode = "carbonate"
            files_data = self.params["file_list"]
            output_path = self.params["output_path"]
            protect_originals = self.params["protect_originals"]
            
            ref_settings = settings.get_setting("REFERENCE_MATERIALS", sub_key=mode.capitalize()) or []
            
            total_steps = len(files_data) * 8 + 2
            current_step = 0

            self.log.emit(f"Starting Carbonate Process & Combine for {len(files_data)} files...", "white")

            combined_data = {} 
            
            if protect_originals:
                temp_dir = tempfile.mkdtemp(prefix="mrsi_combine_tmp_")
                self.log.emit("Working on temporary copies to protect originals.", "white")

            self.log.emit("Starting background Excel engine...", "white")
            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False

            def local_refresh(file_path):
                wb = app.books.open(os.path.abspath(file_path))
                wb.app.calculate()
                time.sleep(1.0)
                wb.save()
                wb.close()

            for idx, file_info in enumerate(files_data):
                if not self._is_running: return self.stopped_early.emit()

                raw_file = file_info["path"]
                sheet_name = file_info["sheet"]
                target_file = raw_file
                
                if protect_originals and temp_dir:
                    filename = os.path.basename(raw_file)
                    target_file = os.path.join(temp_dir, f"{idx}_{filename}")
                    shutil.copy(raw_file, target_file)

                self.log.emit("-" * 40, "white")
                self.log.emit(f"⚙️ Processing File {idx+1}/{len(files_data)}: {os.path.basename(raw_file)}", "white")
                
                step2_carbonate = lambda: step2_tosort_carbonate(target_file, "")
                
                step_order = [
                    ("Step 1: Data", lambda: step1_data_carbonate(target_file, sheet_name)),
                    ("Step 2: To Sort", lambda: (local_refresh(target_file), step2_carbonate())), 
                    ("Step 3: Last 6", lambda: step3_last6_carbonate(target_file)),
                    ("Step 4: Pre-Group", lambda: step4_pre_group_carbonate(target_file)),
                    ("Step 5: Group", lambda: step5_group_carbonate(target_file)),
                    ("Step 6: Normalization", lambda: step6_normalization_carbonate(target_file)),
                    ("Step 7: Report", lambda: step7_report_carbonate(target_file)),
                ]

                for name, func in step_order:
                    if not self._is_running: return self.stopped_early.emit()
                    self.log.emit(f"▶  Running {name}...", "white")
                    time.sleep(1.0)
                    try:
                        func()
                        self.log.emit(f"✔  {name} Completed", "green")
                    except Exception as e:
                        raise Exception(f"{name} Failed on file {os.path.basename(target_file)}: {str(e)}")
                        
                    current_step += 1
                    self.progress.emit(current_step, total_steps, f"{os.path.basename(target_file)} - {name}")

                self.log.emit(f"🔄 Preparing final calculations for {os.path.basename(target_file)}...", "white")
                local_refresh(target_file)
                time.sleep(1.5)

                self.log.emit(f"Extracting standards from {os.path.basename(target_file)}...", "white")
                
                wb = openpyxl.load_workbook(target_file, data_only=True)
                if "Normalization_DNT" not in wb.sheetnames:
                    raise Exception(f"Sheet 'Normalization_DNT' was not created in {os.path.basename(target_file)}")
                    
                ws = wb["Normalization_DNT"]
                
                
                data_header_row = 1
                for r in range(1, ws.max_row + 1):
                    val = str(ws.cell(row=r, column=3).value or "").strip().lower()
                    if "identifier" in val or "time code" in str(ws.cell(row=r, column=2).value or "").strip().lower():
                        data_header_row = r
                        break

                # --- CARBONATE EXTRACTION LOGIC ---
                end_blue_box = max(1, data_header_row - 1)
                file_blue_box = []
                for r in range(1, end_blue_box + 1):
                    row_cells = [ws.cell(row=r, column=c) for c in range(1, 16)] 
                    file_blue_box.append(row_cells)
                    
                file_data_header = [ws.cell(row=data_header_row, column=c) for c in range(1, 18)]
                
                file_blocks = {}
                current_mat = None
                recording = False
                
                for r in range(data_header_row + 1, ws.max_row + 1):
                    cell_c = ws.cell(row=r, column=3).value
                    cell_b = ws.cell(row=r, column=2).value 
                    
                    if cell_c is not None and str(cell_c).strip() != "":
                        base_mat = self.get_base_reference_name(cell_c, ref_settings)
                        if base_mat:
                            current_mat = base_mat
                            recording = True 
                            
                            if current_mat not in file_blocks:
                                file_blocks[current_mat] = {
                                    'filename': os.path.basename(target_file),
                                    'timestamp': None,
                                    'rows': []
                                }
                                
                            if file_blocks[current_mat]['timestamp'] is None and cell_b:
                                ts = self.parse_timestamp(cell_b)
                                if ts != datetime.min:
                                    file_blocks[current_mat]['timestamp'] = ts
                        else:
                            recording = False
                            
                    if recording:
                        row_cells = [ws.cell(row=r, column=c) for c in range(1, 18)] 
                        file_blocks[current_mat]['rows'].append(row_cells)
                        
                # Trim trailing blank rows
                for mat, block_data in file_blocks.items():
                    rows = block_data['rows']
                    while rows:
                        is_empty = True
                        for cell in rows[-1]:
                            if cell.value is not None and str(cell.value).strip() != "":
                                is_empty = False
                                break
                        
                        if is_empty:
                            rows.pop() 
                        else:
                            break 
                            
                for mat, block_data in file_blocks.items():
                    if mat not in combined_data:
                        combined_data[mat] = []
                    combined_data[mat].append({
                        'filename': block_data['filename'],
                        'timestamp': block_data['timestamp'],
                        'blue_box': file_blue_box,
                        'data_header': file_data_header,
                        'block_rows': block_data['rows']
                    })

                wb.close()
                current_step += 1
                self.progress.emit(current_step, total_steps, f"Extracted {os.path.basename(target_file)}")

            if not self._is_running: return self.stopped_early.emit()
            self.log.emit("=" * 40, "white")
            self.log.emit(f"Compiling {len(combined_data)} Standard sheets...", "white")
            
            out_wb = openpyxl.Workbook()
            out_wb.remove(out_wb.active) 
            
            grey_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            black_font = openpyxl.styles.Font(bold=True, color="000000")
            
            color_hex_map = {
                "red": "FF0000", "blue": "0000FF", "darkblue": "00008B", 
                "orange": "E46C0A", "green": "00B050", "lightblue": "5DADE2", 
                "black": "000000"
            }

            # --- CARBONATE COMPILATION LOGIC ---
            for mat_name, files_data_list in combined_data.items():
                clean_title = str(mat_name)[:31] 
                ws_out = out_wb.create_sheet(title=clean_title)
                
                files_data_list.sort(key=lambda x: x['timestamp'] if x['timestamp'] else datetime.min)
                
                mat_color = "000000"
                for std in ref_settings:
                    if std.get("col_c") == mat_name:
                        c_name = std.get("color", "black").lower()
                        mat_color = color_hex_map.get(c_name, "000000")
                        break
                
                if files_data_list and files_data_list[0]['data_header']:
                    for c_idx, src_cell in enumerate(files_data_list[0]['data_header']):
                        tgt_cell = ws_out.cell(row=1, column=1 + c_idx)
                        self.copy_cell_exact(src_cell, tgt_cell)
                
                ws_out.freeze_panes = "A2"
                current_out_row = 2
                
                file_data_chunks = []
                
                for file_data in files_data_list:
                    ws_out.merge_cells(start_row=current_out_row, start_column=1, end_row=current_out_row, end_column=17)
                    div_cell = ws_out.cell(row=current_out_row, column=1, value=f"Data from: {file_data['filename']}")
                    div_cell.fill = grey_fill
                    div_cell.font = black_font
                    div_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                    
                    for c in range(1, 18):
                        ws_out.cell(row=current_out_row, column=c).fill = grey_fill
                        
                    current_out_row += 1
                    start_data_row = current_out_row
                    
                    for r_idx, row_cells in enumerate(file_data['block_rows']):
                        for c_idx, src_cell in enumerate(row_cells):
                            tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=1 + c_idx)
                            self.copy_cell_exact(src_cell, tgt_cell)
                            
                    for r_idx, row_cells in enumerate(file_data['blue_box']):
                        for c_idx, src_cell in enumerate(row_cells):
                            tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=19 + c_idx)
                            self.copy_cell_exact(src_cell, tgt_cell)
                            
                    end_data_row = start_data_row + len(file_data['block_rows']) - 1
                    if end_data_row >= start_data_row:
                        file_data_chunks.append((start_data_row, end_data_row))
                            
                    max_written = max(
                        end_data_row,
                        start_data_row + len(file_data['blue_box']) - 1
                    )
                    
                    current_out_row = max_written + 3 
                    
                for c in range(1, 35):
                    ws_out.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15

                # --- CARBONATE LINE CHARTS GENERATION ---
                columns_to_graph = [
                    (11, "δ¹³C RAW"),
                    (14, "δ¹⁸O RAW")
                ]
                
                chart_col_offset = 3 
                chart_row = current_out_row + 2
                
                for col_idx, chart_title in columns_to_graph:
                    chart = ScatterChart()
                    chart.title = chart_title
                    chart.style = 13 
                    chart.x_axis.title = "Time"
                    chart.y_axis.title = chart_title
                    chart.scatterStyle = "line"
                    chart.legend = None 
                    
                    has_data = False
                    
                    for chunk in file_data_chunks:
                        s_row, e_row = chunk
                        
                        valid_ranges = []
                        current_start = None
                        
                        for r in range(s_row, e_row + 1):
                            val_b = ws_out.cell(row=r, column=2).value
                            val_y = ws_out.cell(row=r, column=col_idx).value
                            
                            is_valid_time = isinstance(val_b, datetime) or (isinstance(val_b, str) and re.search(r'\d{4}[-/]\d{2}', val_b))
                            is_valid_y = isinstance(val_y, (int, float))
                            
                            if is_valid_time and is_valid_y:
                                if current_start is None:
                                    current_start = r
                            else:
                                if current_start is not None:
                                    valid_ranges.append((current_start, r - 1))
                                    current_start = None
                        
                        if current_start is not None:
                            valid_ranges.append((current_start, e_row))
                            
                        for sub_s, sub_e in valid_ranges:
                            xvalues = Reference(ws_out, min_col=2, min_row=sub_s, max_row=sub_e)
                            yvalues = Reference(ws_out, min_col=col_idx, min_row=sub_s, max_row=sub_e)
                            
                            series = Series(values=yvalues, xvalues=xvalues, title_from_data=False)
                            
                            line_prop = series.graphicalProperties.line
                            line_prop.solidFill = mat_color
                            series.marker.symbol = "circle"
                            series.marker.graphicalProperties.solidFill = mat_color
                            series.marker.graphicalProperties.line.solidFill = mat_color
                            
                            chart.series.append(series)
                            has_data = True
                            
                    if has_data:
                        col_letter = openpyxl.utils.get_column_letter(chart_col_offset)
                        ws_out.add_chart(chart, f"{col_letter}{chart_row}")
                        chart_col_offset += 6 

            out_wb.save(output_path)
            self.log.emit("-" * 50, "white")
            self.log.emit(f"✅ Combine Complete! Saved to: {output_path}", "green")
            
            if self.params.get("open_on_complete") and os.path.exists(output_path):
                self.log.emit("Opening combined file...", "white")
                try:
                    if sys.platform == "win32":
                        os.startfile(output_path)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", output_path])
                    else:
                        subprocess.call(["xdg-open", output_path])
                except Exception as e:
                    self.log.emit(f"Warning: Could not automatically open file: {e}", "white")

            self.progress.emit(total_steps, total_steps, "Done")
            self.finished.emit()

        except Exception as e:
            self.log.emit(f"❌ Critical Error in Combine: {str(e)}", "red")
            self.error.emit(str(e))
            
        finally:
            if app:
                try:
                    app.quit()
                except:
                    pass
                    
            if temp_dir and os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    self.log.emit("Cleaned up temporary workspace.", "white")
                except Exception as e:
                    self.log.emit(f"Warning: Could not completely delete temp directory: {e}", "white")