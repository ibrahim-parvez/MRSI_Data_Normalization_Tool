import os
import shutil
import re
import time
import openpyxl
import tempfile
from copy import copy
from datetime import datetime
import xlwings as xw
from openpyxl.comments import Comment

import sys
import subprocess
from PyQt6.QtGui import QPainter, QColor, QPen, QFont, QIcon, QDesktopServices
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QFileDialog, QGroupBox, QLineEdit, QAbstractItemView, 
    QMessageBox, QRadioButton, QButtonGroup, QTableWidget, QTableWidgetItem, QHeaderView, QStyle, QCheckBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QStandardPaths, QUrl

import utils.settings as settings

# ---- Import step modules ----
from processors.carbonate.step1_data import step1_data_carbonate
from processors.carbonate.step2_tosort import step2_tosort_carbonate
from processors.carbonate.step3_last6 import step3_last6_carbonate
from processors.carbonate.step4_pre_group import step4_pre_group_carbonate
from processors.carbonate.step5_group import step5_group_carbonate
from processors.carbonate.step6_normalization import step6_normalization_carbonate
from processors.carbonate.step7_report import step7_report_carbonate

from processors.water.step1_data import step1_data_water
from processors.water.step2_tosort import step2_tosort_water
from processors.water.step3_last6 import step3_last6_water
from processors.water.step4_pre_group import step4_pre_group_water
from processors.water.step5_group import step5_group_water
from processors.water.step6_normalization import step6_normalization_water
from processors.water.step7_report import step7_report_water

class DragDropBox(QGroupBox):
    filesDropped = pyqtSignal(list) # Changed to emit a list of paths

    def __init__(self, title, parent=None):
        super().__init__(title, parent)
        self.setAcceptDrops(True)
        self.drag_active = False

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                if url.toLocalFile().lower().endswith(('.xlsx', '.xls')):
                    event.accept()
                    self.drag_active = True
                    self.update() # Trigger repaint for animation
                    return
        event.ignore()

    def dragLeaveEvent(self, event):
        self.drag_active = False
        self.update()

    def dropEvent(self, event):
        self.drag_active = False
        self.update()
        
        valid_files = []
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if path.lower().endswith(('.xlsx', '.xls')):
                valid_files.append(path)
                
        if valid_files:
            self.filesDropped.emit(valid_files)
            event.accept()

    def paintEvent(self, event):
        # 1. Draw the standard GroupBox look
        super().paintEvent(event)

        # 2. If a file is hovering, draw the "Cool Upload" overlay
        if self.drag_active:
            painter = QPainter(self)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)
            
            # Create semi-transparent overlay
            rect = self.contentsRect()
            overlay_color = QColor("#E3F2FD") 
            overlay_color.setAlpha(200) # Transparency
            
            painter.setBrush(overlay_color)
            
            # Dashed Border
            pen = QPen(QColor("#2196F3"))
            pen.setWidth(3)
            pen.setStyle(Qt.PenStyle.DashLine)
            painter.setPen(pen)
            
            # Draw rounded rect
            painter.drawRoundedRect(rect.adjusted(5, 5, -5, -5), 10, 10)
            
            # Draw Text
            painter.setPen(QColor("#0D47A1"))
            font = QFont("Arial", 16, QFont.Weight.Bold)
            painter.setFont(font)
            painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, "📂 Drop Excel File(s) Here")


class FileTableWidget(QTableWidget):
    def paintEvent(self, event):
        super().paintEvent(event)
        
        # Only draw the big text when empty. No overlapping when files are present!
        if self.rowCount() == 0:
            painter = QPainter(self.viewport())
            rect = self.viewport().rect()
            painter.setPen(QColor("#888888"))
            font = self.font()
            font.setPointSize(14)
            font.setItalic(True)
            font.setBold(True)
            painter.setFont(font)
            painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, "📥 Drag & Drop File(s) Here")

            
class CombineTab(QWidget):
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)  # Enable Drag and Drop
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10) # Added slight margins for breathing room

        # Add the warning label FIRST so it sits right above the buttons
        self.mode_warning_label = QLabel("Please select either water or carbonate")
        self.mode_warning_label.setStyleSheet("color: #d32f2f; font-size: 11px; font-weight: bold;")
        self.mode_warning_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.mode_warning_label)

        # --- 0. Mode Configuration (Sleek Toggles) ---
        mode_layout = QHBoxLayout()
        
        self.btn_water = QPushButton("Water")
        self.btn_carbonate = QPushButton("Carbonate")
        self.btn_water.setCheckable(True)
        self.btn_carbonate.setCheckable(True)
        
        # 💡 Removed self.btn_water.setChecked(True) so neither is selected by default
        
        self.btn_water.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_carbonate.setCursor(Qt.CursorShape.PointingHandCursor)
        
        # 💡 Removed QButtonGroup entirely to allow manual toggling off
        
        # Unified styling matching the Reset button aesthetic
        toggle_style = """
            QPushButton {
                background-color: #f3f3f3;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 4px 12px;
                color: #333;
                min-width: 80px; /* Prevents text cutoff when it turns bold */
            }
            QPushButton:hover {
                background-color: #e5e5e5;
            }
            QPushButton:checked {
                font-weight: bold;
                border: 2px solid #333;
                background-color: #e5e5e5; /* Keeps it slightly darker so it looks pressed */
            }
        """
        self.btn_water.setStyleSheet(toggle_style)
        self.btn_carbonate.setStyleSheet(toggle_style)
        
        # Connect to our new custom manual toggle handler
        self.btn_water.clicked.connect(self._on_mode_clicked)
        self.btn_carbonate.clicked.connect(self._on_mode_clicked)

        # The stretch on both sides is what forces the buttons to center
        mode_layout.addStretch()
        mode_layout.addWidget(self.btn_water)
        mode_layout.addWidget(self.btn_carbonate)
        mode_layout.addStretch()
        layout.addLayout(mode_layout)
        
        # Adding a bit of spacing
        #layout.addSpacing(10)

        # --- 1. File Handling (Horizontal) ---
        copy_group = QGroupBox("File Handling")
        copy_layout = QHBoxLayout()
        
        self.handling_group = QButtonGroup(self)
        
        self.radio_temp_copy = QRadioButton("Process data on temp files")
        self.radio_modify_orig = QRadioButton("Process data on original files")
        self.radio_temp_copy.setChecked(True) # Safer default
        
        self.handling_group.addButton(self.radio_temp_copy)
        self.handling_group.addButton(self.radio_modify_orig)
        
        copy_layout.addWidget(self.radio_temp_copy)
        copy_layout.addWidget(self.radio_modify_orig)
        copy_layout.addStretch() # Pushes radio buttons to the left
        
        copy_group.setLayout(copy_layout)
        layout.addWidget(copy_group)

        # --- 2. File List Section (Custom DragDropBox) ---
        list_group = DragDropBox("Raw Files to Combine")
        list_group.filesDropped.connect(self._add_files_to_table)
        
        list_layout = QVBoxLayout()
        list_layout.setSpacing(2) # 💡 Keeps the gap between table and footer razor-thin
        
        # Top Table Controls
        top_controls = QHBoxLayout()
        
        self.browse_files_btn = QPushButton(" Browse Files")

        # Add standard folder icon
        folder_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_DirIcon)
        self.browse_files_btn.setIcon(folder_icon)

        self.browse_files_btn.clicked.connect(self.add_files)
        self.browse_files_btn.setStyleSheet("padding: 5px 15px;")
        
        self.clear_btn = QPushButton(" Clear All")

        # Add standard trash/delete icon
        trash_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_TrashIcon)
        self.clear_btn.setIcon(trash_icon)

        self.clear_btn.clicked.connect(self.clear_all)
        self.clear_btn.setStyleSheet("padding: 5px 15px;")

        top_controls.addWidget(self.browse_files_btn)
        top_controls.addStretch() # Pushes Browse left and Clear right
        top_controls.addWidget(self.clear_btn)
        list_layout.addLayout(top_controls)
        
        # Table Setup (Using our custom class!)
        self.file_table = FileTableWidget(0, 3)
        self.file_table.setHorizontalHeaderLabels(["File Name", "Default Sheet Name", ""])
        
        # Column resizing to prevent cutoff
        self.file_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.file_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.file_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.file_table.setColumnWidth(2, 40) # Small width for the red 'X' button
        
        self.file_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection) 
        self.file_table.setAlternatingRowColors(True)
        list_layout.addWidget(self.file_table)
        
        # 💡 NEW: Tiny, permanent footer label right under the table. No layout jumps!
        self.footer_hint = QLabel("Drag & drop more files anywhere in this box...")
        self.footer_hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.footer_hint.setStyleSheet("color: #999999; font-style: italic; font-size: 11px;")
        self.footer_hint.setContentsMargins(0, 0, 0, 0) # Strips out any hidden margins
        self.footer_hint.hide()
        list_layout.addWidget(self.footer_hint)

        list_group.setLayout(list_layout)
        layout.addWidget(list_group)
        

        # --- 3. Output Configuration ---
        output_group = QGroupBox("Final Combined Output")
        output_layout = QVBoxLayout()
        
        # 💡 Squeeze out the excessive vertical space
        output_layout.setContentsMargins(10, 8, 10, 8) 
        output_layout.setSpacing(5) # Tighter gap between the two rows
        
        row_out = QHBoxLayout()
        self.output_path_input = QLineEdit()
        
        # Set default to Desktop
        desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        default_out_path = os.path.join(desktop_dir, "Combined_Normalization_Data.xlsx")
        self.output_path_input.setText(default_out_path)
        
        self.browse_out_btn = QPushButton(" Browse Files")
        self.browse_out_btn.setFixedWidth(130) # 💡 Lock the width to match the Open button
        
        # Add standard folder icon
        folder_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_DirIcon)
        self.browse_out_btn.setIcon(folder_icon)
        self.browse_out_btn.clicked.connect(self.browse_output)
        
        row_out.addWidget(QLabel("Output File:"))
        row_out.addWidget(self.output_path_input)
        row_out.addWidget(self.browse_out_btn)
        
        output_layout.addLayout(row_out)

        action_row = QHBoxLayout()
        
        self.open_checkbox = QCheckBox("Open file upon completion of processing")
        self.open_checkbox.setChecked(True) # Remains enabled by default
        self.open_checkbox.setCursor(Qt.CursorShape.PointingHandCursor)
        
        self.btn_open_file = QPushButton(" Open File")
        self.btn_open_file.setFixedWidth(130) # 💡 Match the exact width of the Browse button
        self.btn_open_file.setCursor(Qt.CursorShape.PointingHandCursor)
        
        open_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_FileIcon)
        self.btn_open_file.setIcon(open_icon)
        
        self.btn_open_file.clicked.connect(self.open_combined_file)
        
        action_row.addWidget(self.open_checkbox)
        action_row.addStretch() # This pushes the Open button to the far right edge
        action_row.addWidget(self.btn_open_file)
        
        output_layout.addLayout(action_row)

        output_group.setLayout(output_layout)
        layout.addWidget(output_group)
        
    # --- Drag and Drop Events ---
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
    
    def dropEvent(self, event):
        files = []
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.endswith(('.xlsx', '.xls')):
                files.append(file_path)
        self._add_files_to_table(files)

    def _add_files_to_table(self, files):
        # Grab full paths from UserRole to check for existing files correctly
        existing_files = []
        for i in range(self.file_table.rowCount()):
            item = self.file_table.item(i, 0)
            if item:
                existing_files.append(item.data(Qt.ItemDataRole.UserRole))
                
        if self.btn_water.isChecked():
            default_sheet = "ExportGB1.wke"
        elif self.btn_carbonate.isChecked():
            default_sheet = "ExportGB2.wke"
        else:
            default_sheet = ""
        
        for f in files:
            if f not in existing_files:
                row = self.file_table.rowCount()
                self.file_table.insertRow(row)
                
                # Col 0: File Name (but store full path internally)
                filename = os.path.basename(f)
                path_item = QTableWidgetItem(filename)
                path_item.setData(Qt.ItemDataRole.UserRole, f) # 💡 Secretly store full path here
                path_item.setFlags(path_item.flags() & ~Qt.ItemFlag.ItemIsEditable) 
                self.file_table.setItem(row, 0, path_item)
                
                # Col 1: Default Sheet Name
                sheet_item = QTableWidgetItem(default_sheet)
                self.file_table.setItem(row, 1, sheet_item)
                
                # Col 2: Delete Button
                del_btn = QPushButton("−") # Using your exact minus character
                del_btn.setFixedSize(24, 24)
                del_btn.setToolTip("Remove this file")
                del_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                del_btn.setStyleSheet("""
                    QPushButton { 
                        background-color: #ff4d4d; 
                        color: white; 
                        border: none; 
                        border-radius: 4px; 
                        font-weight: bold; 
                        font-size: 16px; 
                        padding: 0px; 
                    }
                    QPushButton:hover { 
                        background-color: #d32f2f; 
                    }
                """)
                del_btn.clicked.connect(lambda _, r=path_item: self._remove_specific_row(r))
                
                # Wrapper to keep it perfectly centered in the table cell
                cell_widget = QWidget()
                cell_layout = QHBoxLayout(cell_widget)
                cell_layout.setContentsMargins(0, 0, 0, 0)
                cell_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                cell_layout.addWidget(del_btn)
                
                self.file_table.setCellWidget(row, 2, cell_widget)
            
            self._update_footer_visibility()

    def _remove_specific_row(self, item):
        # Find the row of the item and remove it
        row = self.file_table.row(item)
        if row >= 0:
            self.file_table.removeRow(row)
        self._update_footer_visibility()

    def update_default_sheets_in_table(self):
        if self.btn_water.isChecked():
            default_sheet = "ExportGB1.wke"
        elif self.btn_carbonate.isChecked():
            default_sheet = "ExportGB2.wke"
        else:
            default_sheet = "" # Blank if neither is selected

        for row in range(self.file_table.rowCount()):
            current_text = self.file_table.item(row, 1).text()
            # If it's one of the defaults or blank, update it
            if current_text in ["ExportGB1.wke", "ExportGB2.wke", ""]:
                self.file_table.item(row, 1).setText(default_sheet)

    def add_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Raw Excel Files", "", "Excel Files (*.xlsx *.xls)"
        )
        if files:
            self._add_files_to_table(files)

    def remove_selected(self):
        # Remove from bottom to top to avoid index shifting issues
        rows = sorted(set(index.row() for index in self.file_table.selectedIndexes()), reverse=True)
        for row in rows:
            self.file_table.removeRow(row)

    def clear_all(self):
        self.file_table.setRowCount(0)
        self._update_footer_visibility()

    def browse_output(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Combined File", "Combined_Normalization_Data.xlsx", "Excel Files (*.xlsx)"
        )
        if path:
            self.output_path_input.setText(path)

    def get_run_parameters(self):
        file_list = []
        for row in range(self.file_table.rowCount()):
            file_list.append({
                "path": self.file_table.item(row, 0).data(Qt.ItemDataRole.UserRole),
                "sheet": self.file_table.item(row, 1).text().strip()
            })
            
        return {
            "mode": "water" if self.btn_water.isChecked() else "carbonate",
            "protect_originals": self.radio_temp_copy.isChecked(),
            "file_list": file_list,
            "output_path": self.output_path_input.text().strip(),
            "open_on_complete": self.open_checkbox.isChecked() # 💡 NEW
        }

    def _on_mode_clicked(self, checked):
        sender = self.sender()
        
        # 💡 Restrict unselecting: If they try to uncheck the active button, force it back on!
        if not checked:
            sender.setChecked(True)
            return

        # If we reach here, a new selection was made.
        # Ensure the OTHER button is turned off.
        if sender == self.btn_water:
            self.btn_carbonate.setChecked(False)
        else:
            self.btn_water.setChecked(False)
            
        # Hide the warning text permanently
        self.mode_warning_label.hide()
            
        # Trigger the table update
        self.update_default_sheets_in_table()

    def open_combined_file(self):
        path = self.output_path_input.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "File Not Found", "The combined file has not been created yet or the path is invalid.")
            return
        
        # Safely open the file cross-platform
        QDesktopServices.openUrl(QUrl.fromLocalFile(path))
    
    def _update_footer_visibility(self):
        if self.file_table.rowCount() == 0:
            self.footer_hint.hide()
        else:
            self.footer_hint.show()
    

# =========================================================================
# ALL-IN-ONE COMBINE WORKER
# =========================================================================

class CombineWorker(QThread):
    log = pyqtSignal(str, str)
    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal()
    error = pyqtSignal(str)
    stopped_early = pyqtSignal()

    def __init__(self, params):
        super().__init__()
        self.params = params
        self._is_running = True

    def stop(self):
        self._is_running = False

    def copy_cell_exact(self, src_cell, tgt_cell):
        # This extracts the evaluated numerical value because we loaded the 
        # workbook with `data_only=True` right after xlwings calculated it.
        tgt_cell.value = src_cell.value
        
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = copy(src_cell.number_format)
            tgt_cell.alignment = copy(src_cell.alignment)
            
        # Copy comments exactly as requested
        if src_cell.comment:
            tgt_cell.comment = Comment(src_cell.comment.text, src_cell.comment.author)

    def get_base_reference_name(self, identifier, mode, ref_settings):
        if not identifier: return None
        raw_text = str(identifier).upper().strip()
        text_clean = re.sub(r'[\s\-_]+', '', raw_text)
        text_no_std = text_clean.replace("STD", "")

        # Catch HeCO2/CO2 typos and apply globally
        if "CO2" in text_clean or "HECO2" in text_clean or "HEC02" in text_clean or "C02" in text_clean:
            return "HeCO2"

        for std in ref_settings:
            std_name = std.get("col_c")
            if not std_name: continue
            
            std_clean = re.sub(r'[\s\-_]+', '', str(std_name).upper())
            std_no_std = std_clean.replace("STD", "")
            
            if std_clean in text_clean:
                return std_name
            if len(std_no_std) >= 4 and std_no_std in text_no_std:
                return std_name
        return None

    def parse_timestamp(self, ts_val):
        """Robustly extracts datetime from cell value for chronological sorting."""
        if isinstance(ts_val, datetime):
            return ts_val
        if isinstance(ts_val, str):
            try:
                # Catch standard timestamp formats like "2025/10/30 14:22:47" or "2025-10-30 14:22:47"
                match = re.search(r'(\d{4}[-/]\d{2}[-/]\d{2}\s+\d{2}:\d{2}:\d{2})', ts_val)
                if match:
                    return datetime.strptime(match.group(1).replace('-', '/'), "%Y/%m/%d %H:%M:%S")
            except:
                pass
        # Fallback to absolute minimum if timestamp isn't readable
        return datetime.min

    def run(self):
        temp_dir = None
        app = None  # Holds our persistent Excel engine
        try:
            mode = self.params["mode"]
            files_data = self.params["file_list"]
            output_path = self.params["output_path"]
            protect_originals = self.params["protect_originals"]
            
            ref_settings = settings.get_setting("REFERENCE_MATERIALS", sub_key=mode.capitalize()) or []
            
            total_steps = len(files_data) * 8 + 2
            current_step = 0

            self.log.emit(f"Starting {mode.title()} Process & Combine for {len(files_data)} files...", "white")

            combined_data = {} 
            header_row_data = None
            summary_box_cells = [] 
            summary_box_max_row = 0
            summary_box_max_col = 0
            
            # Setup temporary directory if protecting originals
            if protect_originals:
                temp_dir = tempfile.mkdtemp(prefix="mrsi_combine_tmp_")
                self.log.emit("Working on temporary copies to protect originals.", "white")

            # ========================================================
            # LAUNCH EXCEL ONCE FOR THE ENTIRE BATCH
            # ========================================================
            self.log.emit("Starting background Excel engine...", "white")
            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False

            def local_refresh(file_path):
                wb = app.books.open(os.path.abspath(file_path))
                wb.app.calculate()
                time.sleep(1.0)
                wb.save()
                wb.close()

            for idx, file_info in enumerate(files_data):
                if not self._is_running: return self.stopped_early.emit()

                raw_file = file_info["path"]
                sheet_name = file_info["sheet"]
                target_file = raw_file
                
                # 1. Handle Copying
                if protect_originals and temp_dir:
                    filename = os.path.basename(raw_file)
                    target_file = os.path.join(temp_dir, f"{idx}_{filename}")
                    shutil.copy(raw_file, target_file)

                # ========================================================
                # 2. RUN EXACT PROCESSING PIPELINE
                # ========================================================
                self.log.emit("-" * 40, "white")
                self.log.emit(f"⚙️ Processing File {idx+1}/{len(files_data)}: {os.path.basename(raw_file)}", "white")
                
                step2_carbonate = lambda: step2_tosort_carbonate(target_file, "")
                step2_water = lambda: step2_tosort_water(target_file, "")
                
                if mode == "carbonate":
                    step_order = [
                        ("Step 1: Data", lambda: step1_data_carbonate(target_file, sheet_name)),
                        ("Step 2: To Sort", lambda: (local_refresh(target_file), step2_carbonate())), 
                        ("Step 3: Last 6", lambda: step3_last6_carbonate(target_file)),
                        ("Step 4: Pre-Group", lambda: step4_pre_group_carbonate(target_file)),
                        ("Step 5: Group", lambda: step5_group_carbonate(target_file)),
                        ("Step 6: Normalization", lambda: step6_normalization_carbonate(target_file)),
                        ("Step 7: Report", lambda: step7_report_carbonate(target_file)),
                    ]
                else:
                    step_order = [
                        ("Step 1: Data", lambda: step1_data_water(target_file, sheet_name)),
                        ("Step 2: To Sort", lambda: (local_refresh(target_file), step2_water())),
                        ("Step 3: Last 6", lambda: step3_last6_water(target_file)),
                        ("Step 4: Pre-Group", lambda: (local_refresh(target_file), step4_pre_group_water(target_file))),
                        ("Step 5: Group", lambda: (local_refresh(target_file), step5_group_water(target_file))),
                        ("Step 6: Normalization", lambda: (local_refresh(target_file), step6_normalization_water(target_file))),
                        ("Step 7: Report", lambda: (local_refresh(target_file), step7_report_water(target_file))),
                    ]

                for name, func in step_order:
                    if not self._is_running: return self.stopped_early.emit()
                    self.log.emit(f"▶  Running {name}...", "white")
                    time.sleep(1.0)
                    try:
                        func()
                        self.log.emit(f"✔  {name} Completed", "green")
                    except Exception as e:
                        raise Exception(f"{name} Failed on file {os.path.basename(target_file)}: {str(e)}")
                        
                    current_step += 1
                    self.progress.emit(current_step, total_steps, f"{os.path.basename(target_file)} - {name}")

                # --- CRITICAL FIX FOR ISSUE 1 ---
                # Force Excel to calculate and save the very last round of formulas 
                # (written by Step 6/7) so data_only=True can read the pure numerical values!
                self.log.emit(f"🔄 Preparing final calculations for {os.path.basename(target_file)}...", "white")
                local_refresh(target_file)
                time.sleep(1.5)

                # ========================================================
                # 3. EXTRACTION PHASE (DUAL LOGIC)
                # ========================================================
                self.log.emit(f"Extracting standards from {os.path.basename(target_file)}...", "white")
                
                # data_only=True grabs purely the evaluated numerical values
                wb = openpyxl.load_workbook(target_file, data_only=True)
                if "Normalization_DNT" not in wb.sheetnames:
                    raise Exception(f"Sheet 'Normalization_DNT' was not created in {os.path.basename(target_file)}")
                    
                ws = wb["Normalization_DNT"]
                
                # Locate Data Header Row
                data_header_row = 1
                for r in range(1, ws.max_row + 1):
                    val = str(ws.cell(row=r, column=3).value or "").strip().lower()
                    if "identifier" in val or "time code" in str(ws.cell(row=r, column=2).value or "").strip().lower():
                        data_header_row = r
                        break

                if mode == "water":
                    # --- WATER EXTRACTION LOGIC ---
                    
                    # 1. Extract Green Box (A-Z) up to row before data headers
                    end_green_box = max(15, data_header_row - 1)
                    file_green_box = []
                    for r in range(1, end_green_box + 1):
                        row_cells = [ws.cell(row=r, column=c) for c in range(1, 27)] # A to Z (1 to 26)
                        file_green_box.append(row_cells)
                        
                    # 2. Extract Data Header (A-Q)
                    file_data_header = [ws.cell(row=data_header_row, column=c) for c in range(1, 18)]
                    
                    # 3. Extract Reference Block Rows Only
                    file_blocks = {}
                    current_mat = None
                    recording = False
                    
                    for r in range(data_header_row + 1, ws.max_row + 1):
                        cell_c = ws.cell(row=r, column=3).value
                        cell_b = ws.cell(row=r, column=2).value # Timestamp
                        
                        # Only check if it's a new identifier. Calculation rows have empty col C, so recording stays True.
                        if cell_c is not None and str(cell_c).strip() != "":
                            base_mat = self.get_base_reference_name(cell_c, mode, ref_settings)
                            if base_mat:
                                current_mat = base_mat
                                recording = True # Start capturing this standard and its calculation rows
                                
                                # Setup block if it's the first time we see this material in this file
                                if current_mat not in file_blocks:
                                    file_blocks[current_mat] = {
                                        'filename': os.path.basename(target_file),
                                        'timestamp': None,
                                        'rows': []
                                    }
                                
                                # Capture the very first timestamp for sorting purposes
                                if file_blocks[current_mat]['timestamp'] is None and cell_b:
                                    ts = self.parse_timestamp(cell_b)
                                    if ts != datetime.min:
                                        file_blocks[current_mat]['timestamp'] = ts
                            else:
                                # It's a sample! Stop recording.
                                recording = False
                                
                        if recording:
                            row_cells = [ws.cell(row=r, column=c) for c in range(1, 18)] # A to Q
                            file_blocks[current_mat]['rows'].append(row_cells)
                            
                    # --- CRITICAL FIX FOR ISSUE 2 ---
                    # Trim trailing blank rows (which includes the dark grey divider since it has no values)
                    for mat, block_data in file_blocks.items():
                        rows = block_data['rows']
                        while rows:
                            # Check if the entire row (A to Q) is completely empty of values
                            is_empty = True
                            for cell in rows[-1]:
                                if cell.value is not None and str(cell.value).strip() != "":
                                    is_empty = False
                                    break
                            
                            if is_empty:
                                rows.pop() # Delete the blank/divider row
                            else:
                                break # Stop deleting once we hit actual calculations/data
                            
                    # Map the extracted file-specific data into the global combined dict
                    for mat, block_data in file_blocks.items():
                        if mat not in combined_data:
                            combined_data[mat] = []
                        combined_data[mat].append({
                            'filename': block_data['filename'],
                            'timestamp': block_data['timestamp'],
                            'green_box': file_green_box,
                            'data_header': file_data_header,
                            'block_rows': block_data['rows']
                        })
                else:
                    # --- NEW CARBONATE EXTRACTION LOGIC ---
                    
                    # 1. Extract Blue Box (A-O) up to row before data headers
                    end_blue_box = max(1, data_header_row - 1)
                    file_blue_box = []
                    for r in range(1, end_blue_box + 1):
                        row_cells = [ws.cell(row=r, column=c) for c in range(1, 16)] # A to O (1 to 15)
                        file_blue_box.append(row_cells)
                        
                    # 2. Extract Data Header (A-Q)
                    file_data_header = [ws.cell(row=data_header_row, column=c) for c in range(1, 18)]
                    
                    # 3. Extract Reference Block Rows Only
                    file_blocks = {}
                    current_mat = None
                    recording = False
                    
                    for r in range(data_header_row + 1, ws.max_row + 1):
                        cell_c = ws.cell(row=r, column=3).value
                        cell_b = ws.cell(row=r, column=2).value # Timestamp
                        
                        # Only check if it's a new identifier. Calculation rows have empty col C, so recording stays True.
                        if cell_c is not None and str(cell_c).strip() != "":
                            base_mat = self.get_base_reference_name(cell_c, mode, ref_settings)
                            if base_mat:
                                current_mat = base_mat
                                recording = True # Start capturing this standard and its calculation rows
                                
                                # Setup block if it's the first time we see this material in this file
                                if current_mat not in file_blocks:
                                    file_blocks[current_mat] = {
                                        'filename': os.path.basename(target_file),
                                        'timestamp': None,
                                        'rows': []
                                    }
                                    
                                # Capture the very first timestamp for sorting purposes
                                if file_blocks[current_mat]['timestamp'] is None and cell_b:
                                    ts = self.parse_timestamp(cell_b)
                                    if ts != datetime.min:
                                        file_blocks[current_mat]['timestamp'] = ts
                            else:
                                # It's a sample! Stop recording.
                                recording = False
                                
                        if recording:
                            row_cells = [ws.cell(row=r, column=c) for c in range(1, 18)] # A to Q
                            file_blocks[current_mat]['rows'].append(row_cells)
                            
                    # Trim trailing blank rows (which includes the dark grey divider since it has no values)
                    for mat, block_data in file_blocks.items():
                        rows = block_data['rows']
                        while rows:
                            # Check if the entire row (A to Q) is completely empty of values
                            is_empty = True
                            for cell in rows[-1]:
                                if cell.value is not None and str(cell.value).strip() != "":
                                    is_empty = False
                                    break
                            
                            if is_empty:
                                rows.pop() # Delete the blank/divider row
                            else:
                                break # Stop deleting once we hit actual calculations/data
                                
                    # Map the extracted file-specific data into the global combined dict
                    for mat, block_data in file_blocks.items():
                        if mat not in combined_data:
                            combined_data[mat] = []
                        combined_data[mat].append({
                            'filename': block_data['filename'],
                            'timestamp': block_data['timestamp'],
                            'blue_box': file_blue_box,
                            'data_header': file_data_header,
                            'block_rows': block_data['rows']
                        })

                wb.close()
                current_step += 1
                self.progress.emit(current_step, total_steps, f"Extracted {os.path.basename(target_file)}")

            # ========================================================
            # 4. COMPILATION PHASE (DUAL LOGIC)
            # ========================================================
            if not self._is_running: return self.stopped_early.emit()
            self.log.emit("=" * 40, "white")
            self.log.emit(f"Compiling {len(combined_data)} Standard sheets...", "white")
            
            out_wb = openpyxl.Workbook()
            out_wb.remove(out_wb.active) 
            
            grey_fill = openpyxl.styles.PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            black_font = openpyxl.styles.Font(bold=True, color="000000")

            if mode == "water":
                # --- NEW WATER COMPILATION LOGIC ---
                for mat_name, files_data_list in combined_data.items():
                    clean_title = str(mat_name)[:31] 
                    ws_out = out_wb.create_sheet(title=clean_title)
                    
                    # Sort blocks chronologically based on their parsed timestamp
                    files_data_list.sort(key=lambda x: x['timestamp'] if x['timestamp'] else datetime.min)
                    
                    # Write the Global Header exactly ONCE at the top
                    if files_data_list and files_data_list[0]['data_header']:
                        for c_idx, src_cell in enumerate(files_data_list[0]['data_header']):
                            tgt_cell = ws_out.cell(row=1, column=1 + c_idx)
                            self.copy_cell_exact(src_cell, tgt_cell)
                    
                    # Freeze Pane so the header stays visible when scrolling
                    ws_out.freeze_panes = "A2"
                    
                    current_out_row = 2
                    
                    for file_data in files_data_list:
                        # 1. Sleek Grey Divider Row with the Origin File Name
                        ws_out.merge_cells(start_row=current_out_row, start_column=1, end_row=current_out_row, end_column=17)
                        div_cell = ws_out.cell(row=current_out_row, column=1, value=f"Data from: {file_data['filename']}")
                        div_cell.fill = grey_fill
                        div_cell.font = black_font
                        div_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                        
                        # Pad the rest of the merged area with grey just to be safe
                        for c in range(1, 18):
                            ws_out.cell(row=current_out_row, column=c).fill = grey_fill
                            
                        current_out_row += 1
                        start_data_row = current_out_row
                        
                        # 2. Write Data Block (A-Q)
                        for r_idx, row_cells in enumerate(file_data['block_rows']):
                            for c_idx, src_cell in enumerate(row_cells):
                                tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=1 + c_idx)
                                self.copy_cell_exact(src_cell, tgt_cell)
                                
                        # 3. Write Green Box Summary (S-AR) anchored neatly next to the data block
                        for r_idx, row_cells in enumerate(file_data['green_box']):
                            for c_idx, src_cell in enumerate(row_cells):
                                tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=19 + c_idx)
                                self.copy_cell_exact(src_cell, tgt_cell)
                                
                        # Calc the bottommost row used by either the data block or the green box
                        max_written = max(
                            start_data_row + len(file_data['block_rows']),
                            start_data_row + len(file_data['green_box'])
                        )
                        
                        # Set starting point for next file loop (+ 2 blank rows spacing)
                        current_out_row = max_written + 2
                        
                    # Apply general column formatting widths
                    for c in range(1, 45):
                        ws_out.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15

            else:
                # --- NEW CARBONATE COMPILATION LOGIC ---
                for mat_name, files_data_list in combined_data.items():
                    clean_title = str(mat_name)[:31] 
                    ws_out = out_wb.create_sheet(title=clean_title)
                    
                    # Sort blocks chronologically based on their parsed timestamp
                    files_data_list.sort(key=lambda x: x['timestamp'] if x['timestamp'] else datetime.min)
                    
                    # Write the Global Header exactly ONCE at the top
                    if files_data_list and files_data_list[0]['data_header']:
                        for c_idx, src_cell in enumerate(files_data_list[0]['data_header']):
                            tgt_cell = ws_out.cell(row=1, column=1 + c_idx)
                            self.copy_cell_exact(src_cell, tgt_cell)
                    
                    # Freeze Pane so the header stays visible when scrolling
                    ws_out.freeze_panes = "A2"
                    
                    current_out_row = 2
                    
                    for file_data in files_data_list:
                        # 1. Sleek Grey Divider Row with the Origin File Name
                        ws_out.merge_cells(start_row=current_out_row, start_column=1, end_row=current_out_row, end_column=17)
                        div_cell = ws_out.cell(row=current_out_row, column=1, value=f"Data from: {file_data['filename']}")
                        div_cell.fill = grey_fill
                        div_cell.font = black_font
                        div_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                        
                        # Pad the rest of the merged area with grey just to be safe
                        for c in range(1, 18):
                            ws_out.cell(row=current_out_row, column=c).fill = grey_fill
                            
                        current_out_row += 1
                        start_data_row = current_out_row
                        
                        # 2. Write Data Block (A-Q)
                        for r_idx, row_cells in enumerate(file_data['block_rows']):
                            for c_idx, src_cell in enumerate(row_cells):
                                tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=1 + c_idx)
                                self.copy_cell_exact(src_cell, tgt_cell)
                                
                        # 3. Write Blue Box Summary (Cols A-O) placed on the right (starting at Col S / 19)
                        for r_idx, row_cells in enumerate(file_data['blue_box']):
                            for c_idx, src_cell in enumerate(row_cells):
                                tgt_cell = ws_out.cell(row=start_data_row + r_idx, column=19 + c_idx)
                                self.copy_cell_exact(src_cell, tgt_cell)
                                
                        # Calc the bottommost row used by either the data block or the blue box
                        max_written = max(
                            start_data_row + len(file_data['block_rows']),
                            start_data_row + len(file_data['blue_box'])
                        )
                        
                        # Set starting point for next file loop (+ 2 blank rows spacing)
                        current_out_row = max_written + 2
                        
                    # Apply general column formatting widths (A to AH to cover the blue box on the right)
                    for c in range(1, 35):
                        ws_out.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15

            out_wb.save(output_path)
            self.log.emit("-" * 50, "white")
            self.log.emit(f"✅ Combine Complete! Saved to: {output_path}", "green")
            
            if self.params.get("open_on_complete") and os.path.exists(output_path):
                self.log.emit("Opening combined file...", "white")
                try:
                    if sys.platform == "win32":
                        os.startfile(output_path)
                    elif sys.platform == "darwin":
                        subprocess.call(["open", output_path])
                    else:
                        subprocess.call(["xdg-open", output_path])
                except Exception as e:
                    self.log.emit(f"Warning: Could not automatically open file: {e}", "white")

            self.progress.emit(total_steps, total_steps, "Done")
            self.finished.emit()

        except Exception as e:
            self.log.emit(f"❌ Critical Error in Combine: {str(e)}", "red")
            self.error.emit(str(e))
            
        finally:
            if app:
                try:
                    app.quit()
                except:
                    pass
                    
            if temp_dir and os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    self.log.emit("Cleaned up temporary workspace.", "white")
                except Exception as e:
                    self.log.emit(f"Warning: Could not completely delete temp directory: {e}", "white")